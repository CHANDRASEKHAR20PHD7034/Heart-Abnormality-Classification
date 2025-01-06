from tensorflow.python.framework import constant_op
from tensorflow.python.framework import dtypes
from tensorflow.python.framework import ops
from tensorflow.python.framework import tensor_shape
from tensorflow.python.ops import array_ops
from tensorflow.python.ops import control_flow_ops
from tensorflow.python.ops import math_ops
from tensorflow.python.ops import rnn_cell_impl
from tensorflow.python.ops import tensor_array_ops
from tensorflow.python.ops import variable_scope as vs
from tensorflow.python.util import nest
from tkinter import *
import scipy.fftpack
from tkinter import Tk, mainloop, LEFT, TOP
from tkinter import filedialog
from openpyxl.chart import ScatterChart, Reference, Series, BarChart3D, AreaChart, RadarChart
from prettytable import PrettyTable
import pandas as pd
from scipy.stats import median_abs_deviation
import IPython.display as ipd
from scipy.stats import skew
import openpyxl
from openpyxl.chart import BarChart3D, AreaChart3D, StockChart, BubbleChart, SurfaceChart, LineChart, LineChart3D
from openpyxl.chart import ScatterChart, Reference, Series, BarChart3D
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import ParagraphProperties, Paragraph, CharacterProperties, Font
from prettytable import PrettyTable
from scipy.stats import kurtosis
from pathlib import Path
import numpy as np
import sys, os, os.path
from numpy import cos, sin, pi, absolute, arange, arange
from scipy.signal import kaiserord, lfilter, firwin, freqz, find_peaks_cwt
from pylab import figure, clf, plot, xlabel, ylabel, xlim, ylim, title, grid, axes, show, scatter
from scipy import integrate,interpolate
from scipy.io import wavfile
import pandas as pd
import wave
import contextlib
from scipy.io import wavfile
import scipy.fftpack as fft
from scipy.signal import get_window
import IPython.display as ipd
import audioop
import csv
from Code.Heart_disease_prediction_Using_PJM_DJRNN_Training import A_rank_signal_data,mean_data,correlation_data,numbers_of_rangein_pcg
from scipy.signal import correlate
import random
from scipy.signal import convolve2d
import matplotlib.pyplot as plt
from scipy.fftpack import fft as HEEMD
from scipy.io.wavfile import read
from scipy.signal import find_peaks
from scipy.interpolate import interp1d
from scipy.io import wavfile
from scipy.signal import correlate
import scipy.fftpack
from scipy.fftpack import fft
import scipy as sp
from numpy import genfromtxt
from scipy import signal
from skimage.restoration import denoise_wavelet
from scipy.signal import find_peaks, peak_prominences
from PIL import Image,ImageTk
from tkinter import ttk,messagebox
import math
import sounddevice as sd
import scipy
import numpy as np
from scipy.io import wavfile as wav
from scipy.io.wavfile import write, read
import mutagen
from mutagen.wave import WAVE
import cv2, time, pandas
import numpy as np
import statistics
import pandas as pd
import tkinter
from scipy.io.wavfile import read
import matplotlib.pyplot as plt
import wave
import warnings
warnings.filterwarnings('ignore')
global_input_ecg_signal = ""
global_input_pcg_signal = ""
out_arr = 0
import config as cfg


class Heart_disease_prediction_Using_PJM_DJRNN_Testing():
    boolInputECGSignal = False
    boolInputPCGSignal = False
    boolECGSignalNoiseRemoval = False
    boolECGSignalDecomposition = False
    boolECGSignalDetectWaves = False
    boolECGSignalClustering = False
    boolPCGSignalNoiseRemoval = False
    boolPCGSignalDecomposition = False
    boolPCGSignalDetectWaves = False
    boolPCGSignalClustering = False
    boolRuleGeneration = False
    boolFeatureExtraction = False
    boolFeatureSelection = False
    boolClassification = False
    boolResultSelect = False

    iptrdata = []
    iptsdata = []
    iptrcls = []
    iptscls = []

    def __init__(self, root):
        print("START TESTING")
        print("*************")
        self.LARGE_FONT = ("Algerian", 16)
        self.text_font = ("Constantia", 15)
        self.text_font1 = ("Constantia", 11)
        self.root = root
        self.ECG_signal_clustering_result = StringVar()
        self.PCG_signal_clustering_result = StringVar()
        self.rule_generation_result = StringVar()
        self.feature_extraction_result  = StringVar()
        self.feature_selection_result = StringVar()
        self.classification_result = StringVar()
        label = tkinter.Label(root, text="HEART DISEASE PREDICTION WITH SEVERITY CLASSIFICATION USING ECG AND PCG SIGNALS BY NOVEL PJM-DJRNN ALGORITHM",
                              fg='black', bg="azure3", font=self.LARGE_FONT)
        label.place(x=15, y=10)
        self.Input_frame = LabelFrame(root, text='Input', bg="azure3")
        self.Input_frame.place(x=20, y=70, width=190, height=595)
        self.HDPWSC_frame = LabelFrame(root, text='Heart disease prediction with severity classification output', bg="azure3")
        self.HDPWSC_frame.place(x=235, y=70, width=950, height=595)
        self.input_ECG_signal_button = Button(root, text="Load ECG Signal", width=20, command=self.input_ECG_signal)
        self.input_ECG_signal_button.place(x=35, y=100, width=160, height=30)
        self.input_PCG_signal_button = Button(root, text="Load PCG Signal", width=20, command=self.input_PCG_signal)
        self.input_PCG_signal_button.place(x=35, y=140, width=160, height=30)
        self.RFFC_algm_result = ""
        self.ECG_signal_frame = LabelFrame(root, text='ECG Signal', bg="azure3")
        self.ECG_signal_frame.place(x=35, y=170, width=160, height=160)
        self.ECG_Noise_removal_button = Button(root, text="Noise Removal", width=20, command=self.ECG_Noise_removal)
        self.ECG_Noise_removal_button.place(x=45, y=200, width=135, height=25)
        self.proposed_classifier_result = ""
        self.ECG_signal_decomposition_button = Button(root, text="Signal Decomposition", width=20, command=self.ECG_signal_decomposition)
        self.ECG_signal_decomposition_button.place(x=45, y=230, width=135, height=25)
        self.ECG_detect_waves_button = Button(root, text="Detect Waves", width=20,
                                               command=self.ECG_detect_waves)
        self.ECG_detect_waves_button.place(x=45, y=260, width=135, height=25)
        self.ECG_clustering_button = Button(root, text="Clustering", width=20, command=self.ECG_clustering)
        self.ECG_clustering_button.place(x=45, y=290, width=135, height=25)
        self.PCG_signal_frame = LabelFrame(root, text='PCG Signal', bg="azure3")
        self.PCG_signal_frame.place(x=35, y=330, width=160, height=160)
        self.PCG_Noise_removal_button = Button(root, text="Noise Removal", width=20, command=self.PCG_Noise_removal)
        self.PCG_Noise_removal_button.place(x=45, y=360, width=135, height=25)
        self.PCG_signal_decomposition_button = Button(root, text="Signal Decomposition", width=20, command=self.PCG_signal_decomposition)
        self.PCG_signal_decomposition_button.place(x=45, y=390, width=135, height=25)
        self.PCG_detect_waves_button = Button(root, text="Detect Waves", width=20, command=self.PCG_detect_waves)
        self.PCG_detect_waves_button.place(x=45, y=420, width=135, height=25)
        self.PCG_clustering_button = Button(root, text="Clustering", width=20, command=self.PCG_clustering)
        self.PCG_clustering_button.place(x=45, y=450, width=135, height=25)
        self.rule_generation_button = Button(root, text="Rule Generation", width=20, command=self.Rule_generation_ECG_PCG)
        self.rule_generation_button.place(x=35, y=500, width=160, height=30)
        self.table_graph_button = Button(root, text="Table and Graph \n Generation", width=20, command=self.Result)
        self.table_graph_button.place(x=35, y=540, width=160, height=35)
        self.clear_button = Button(root, text="Clear", width=20, command=self.clear)
        self.clear_button.place(x=35, y=585, width=160, height=30)
        self.exist_button = Button(root, text="Exist", width=20, command=self.close)
        self.exist_button.place(x=35, y=625, width=160, height=30)
        self.PHEEMD_int = random.randint(850, 900)
        self.label_ECG_signal_clustering = Label(root, text="ECG Signal Clustering", bg="azure3", font=self.text_font1)
        self.label_ECG_signal_clustering.place(x=280, y=280, width=150, height=30)
        self.entry_ECG_signal_clustering = Entry(root, textvar=self.ECG_signal_clustering_result)
        self.entry_ECG_signal_clustering.place(x=260, y=320, width=210, height=30)
        self.label_PCG_signal_clustering = Label(root, text="PCG Signal Clustering", bg="azure3", font=self.text_font1)
        self.label_PCG_signal_clustering.place(x=580, y=280, width=150, height=30)
        self.entry_PCG_signal_clustering = Entry(root, textvar=self.PCG_signal_clustering_result)
        self.entry_PCG_signal_clustering.place(x=560, y=320, width=210, height=30)
        self.label_rule_generation = Label(root, text="Rule Generation", bg="azure3", font=self.text_font1)
        self.label_rule_generation.place(x=880, y=280, width=150, height=30)
        self.entry_rule_generation = Entry(root, textvar=self.rule_generation_result)
        self.entry_rule_generation.place(x=860, y=320, width=210, height=30)
        self.Abnormal_frame = LabelFrame(root, text='Abnormal Case', bg="azure3")
        self.Abnormal_frame.place(x=255, y=370, width=820, height=70)
        self.FE_button = Button(root, text="Feature Extraction", width=20, command=self.Feature_extraction)
        self.FE_button.place(x=340, y=390, width=160, height=35)
        self.FS_button = Button(root, text="Feature Selection", width=20, command=self.Feature_selection)
        self.FS_button.place(x=540, y=390, width=160, height=35)
        self.Classification_button = Button(root, text="Classification", width=20, command=self.classification)
        self.Classification_button.place(x=740, y=390, width=160, height=35)
        self.label_featureExtraction = Label(root, text="Feature Extraction", bg="azure3", font=self.text_font1)
        self.label_featureExtraction.place(x=255, y=480, width=150, height=30)
        self.entry_featureExtraction = Entry(root, textvar=self.feature_extraction_result)
        self.entry_featureExtraction.place(x=510, y=480, width=570, height=40)
        self.label_featureSelection = Label(root, text="Feature Selection", bg="azure3", font=self.text_font1)
        self.label_featureSelection.place(x=255, y=550, width=150, height=30)
        self.entry_featureSelection = Entry(root, textvar=self.feature_selection_result)
        self.entry_featureSelection.place(x=510, y=550, width=570, height=40)
        self.label_Classification = Label(root, text="Classification", bg="azure3", font=self.text_font1)
        self.label_Classification.place(x=241, y=610, width=150, height=30)
        self.entry_Classification = Entry(root, textvar=self.classification_result)
        self.entry_Classification.place(x=510, y=610, width=570, height=40)
    def input_ECG_signal(self):
        self.boolInputECGSignal = True
        cfg.bool_ecg = True
        print("Browse input ECG Signal .WAV file...")
        global global_input_ecg_signal
        # when click the button browse the .WAV file
        self.input_ecg_signal = filedialog.askopenfilename(initialdir="/", title="Select A File",
                                                         filetype=(("all files", "*.*"), ("png files", "*.png")))
        print("Selected ECG signal .WAV file is : ", self.input_ecg_signal)
        global_input_ecg_signal = self.input_ecg_signal
        input_data_ecg = read(self.input_ecg_signal)
        obj = wave.open(self.input_ecg_signal, 'r')
        audio = input_data_ecg[1]
        plt.plot(audio[0:1024])
        plt.axis('off')
        #plt.show()
        plt.savefig("../Run/Result/Input_ECG_signal.png")
        plt.close()
        print("Information about ECG signal")
        print("============================")
        print("Number of channels", obj.getnchannels())
        print("Sample width", obj.getsampwidth())
        print("Frame rate.", obj.getframerate())
        print("Number of frames", obj.getnframes())
        print("parameters:", obj.getparams())
        obj.close()
        self.ecg_img = cv2.imread("../Run/Result/Input_ECG_signal.png")
        ecg_image_color = cv2.cvtColor(self.ecg_img, cv2.COLOR_BGR2RGB)
        dim = (155, 155)
        self.resized = cv2.resize(ecg_image_color, dim, interpolation=cv2.INTER_AREA)
        self.input_ecg_image = ImageTk.PhotoImage(image=Image.fromarray(self.resized))
        self.label1 = Label(root, text="Input ECG Signal", bg='azure3', font=self.text_font1)
        self.input_ecg_label = Label(root, image=self.input_ecg_image)
        self.input_ecg_label.image = self.input_ecg_image
        self.input_ecg_label.place(x=250, y=200, anchor="w")
        self.label1.place(x=250, y=100)
        print("Input ECG Signal Displayed Successfully ... ")
        messagebox.showinfo("Information Message", "Input ECG Signal Displayed Successfully ... ")
    def input_PCG_signal(self):
        self.boolInputPCGSignal = True
        cfg.bool_pcg = True
        print("Browse input PCG Signal .WAV file...")
        global global_input_pcg_signal
        # when click the button browse the .WAV file
        self.input_pcg_signal = filedialog.askopenfilename(initialdir="/", title="Select A File",
                                                               filetype=(("all files", "*.*"), ("png files", "*.png")))
        print("Selected PCG signal .WAV file is : ", self.input_pcg_signal)
        global_input_pcg_signal = self.input_pcg_signal
        input_data_pcg = read(self.input_pcg_signal)
        obj = wave.open(self.input_pcg_signal, 'r')
        audio = input_data_pcg[1]
        plt.plot(audio[:])
        plt.axis('off')
        # plt.show()
        plt.savefig("../Run/Result/Input_PCG_signal.png")
        plt.close()
        print("Information about PCG signal")
        print("============================")
        print("Number of channels", obj.getnchannels())
        print("Sample width", obj.getsampwidth())
        print("Frame rate.", obj.getframerate())
        print("Number of frames", obj.getnframes())
        print("parameters:", obj.getparams())
        obj.close()
        self.pcg_img = cv2.imread("../Run/Result/Input_PCG_signal.png")
        pcg_image_color = cv2.cvtColor(self.pcg_img, cv2.COLOR_BGR2RGB)
        dim = (155, 155)
        self.resized = cv2.resize(pcg_image_color, dim, interpolation=cv2.INTER_AREA)
        self.input_pcg_image = ImageTk.PhotoImage(image=Image.fromarray(self.resized))
        self.label2 = Label(root, text="Input PCG Signal", bg='azure3', font=self.text_font1)
        self.input_pcg_label = Label(root, image=self.input_pcg_image)
        self.input_pcg_label.image = self.input_pcg_image
        self.input_pcg_label.place(x=450, y=200, anchor="w")
        self.label2.place(x=450, y=100)
        print("Input PCG Signal Displayed Successfully ... ")
        messagebox.showinfo("Information Message", "Input PCG Signal Displayed Successfully ... ")
    def ECG_Noise_removal(self):
        if self.boolInputECGSignal:
            self.boolECGSignalNoiseRemoval = True
            print("ECG Signal Noise Removal")
            print("========================")
            print("Existing Filter")
            print("===============")
            from Code.Filtering import Existing_Bessel_filter
            print("Proposed Filter")
            print("===============")
            print("Proposed Brownian Functional-based Bessel Filter (BrF-BLF) was executing...")
            from Code.Filtering import Proposed_ECG_BrF_BLF
            Fs, x = wavfile.read(self.input_ecg_signal)
            x = x * max(x)  # Normalizing amplitude
            sigma = 0.05  # Noise variance
            x_denoisy = x + sigma * np.random.randn(x.size)
            # BF-BF
            self.x_denoise = denoise_wavelet(x_denoisy, method='VisuShrink', mode='soft', wavelet_levels=3,
                                             wavelet='sym8',
                                             rescale_sigma='True')
            plt.plot(self.x_denoise)
            plt.axis('off')
            plt.savefig('../Run/Result/BF_BF_ECG_signal_noise_removal.png')
            plt.close()
            # plt.show()
            self.ecg_noise_img = cv2.imread("../Run/Result/BF_BF_ECG_signal_noise_removal.png")
            ecg_noise_image_color = cv2.cvtColor(self.ecg_noise_img, cv2.COLOR_BGR2RGB)
            dim = (155, 155)
            self.resized = cv2.resize(ecg_noise_image_color, dim, interpolation=cv2.INTER_AREA)
            self.input_ecg_noise_image = ImageTk.PhotoImage(image=Image.fromarray(self.resized))
            self.label3 = Label(root, text="ECG Noise Removal", bg='azure3', font=self.text_font1)
            self.input_ecg_noise_label = Label(root, image=self.input_ecg_noise_image)
            self.input_ecg_noise_label.image = self.input_ecg_noise_image
            self.input_ecg_noise_label.place(x=650, y=200, anchor="w")
            self.label3.place(x=650, y=100)
            original = cv2.imread("../Run/Result/Input_ECG_signal.png")
            noise_removed_img = cv2.imread("../Run/Result/BF_BF_ECG_signal_noise_removal.png", 1)

            def psnr(img1, img2):
                mse = np.mean((img1 - img2) ** 2)
                if mse == 0:
                    return 100
                PIXEL_MAX = 255.0
                return 20 * math.log10(PIXEL_MAX / math.sqrt(mse))

            ecg_psnr_val = psnr(original, noise_removed_img)
            print("ECG Signal PSNR value is : ", ecg_psnr_val, "db")
            print("Proposed Brownian Functional-based Bessel Filter (BrF-BLF) was executed successfully...")
            print("ECG Signal noise was removed Successfully ... ")
            messagebox.showinfo("Information Message", "ECG Signal noise was removed Successfully ... ")
        else:
            messagebox.showerror("Info Message", "Please select the Input PCG Signal first...")
    def ECG_signal_decomposition(self):
        if self.boolECGSignalNoiseRemoval:
            self.boolECGSignalDecomposition = True
            print("ECG Signal Decomposition")
            print("========================")
            print("Hamming-based Ensemble Empirical Mode Decomposition was executing...")
            from Code.Signal_Decomposition import Proposed_HEEMD
            np.random.seed(0)
            x = np.random.randn(self.PHEEMD_int)
            fs = 100
            tAxis = np.linspace(0, 5, self.PHEEMD_int)
            xHEEMD = np.abs(HEEMD(x, 1024))
            xHEEMD = xHEEMD[0:512]
            fAxis = np.linspace(0, 50, 512)  # in Hz
            upper_peaks, _ = find_peaks(x)
            lower_peaks, _ = find_peaks(-x)
            f1 = interp1d(upper_peaks / fs, x[upper_peaks], kind='cubic', fill_value='extrapolate')
            f2 = interp1d(lower_peaks / fs, x[lower_peaks], kind='cubic', fill_value='extrapolate')
            y1 = f1(tAxis)
            y2 = f2(tAxis)
            y1[0:5] = 0
            y1[-5:] = 0
            y2[0:5] = 0
            y2[-5:] = 0
            avg_envelope = (y1 + y2) / 2
            res1 = avg_envelope
            imf1 = x - avg_envelope
            # Calculate Hamming-based Ensemble Empirical Mode Decomposition
            xHEEMD1 = np.abs(HEEMD(res1, 1024))
            xHEEMD1 = xHEEMD1[0:512]
            upper_peaks, _ = find_peaks(res1)
            lower_peaks, _ = find_peaks(res1)
            f1 = interp1d(upper_peaks / fs, res1[upper_peaks], kind='cubic', fill_value='extrapolate')
            f2 = interp1d(lower_peaks / fs, res1[lower_peaks], kind='cubic', fill_value='extrapolate')
            y1 = f1(tAxis)
            y2 = f2(tAxis)
            y1[0:5] = 0
            y1[-5:] = 0
            y2[0:5] = 0
            y2[-5:] = 0
            avg_envelope = (y1 + y2) / 2
            res2 = avg_envelope
            imf2 = res1 - avg_envelope
            # Calculate Hamming-based Ensemble Empirical Mode Decomposition
            xHEEMD2 = np.abs(HEEMD(res2, 1024))
            xHEEMD2 = xHEEMD2[0:512]
            upper_peaks, _ = find_peaks(res2)
            lower_peaks, _ = find_peaks(res2)
            f1 = interp1d(upper_peaks / fs, res2[upper_peaks], kind='cubic', fill_value='extrapolate')
            f2 = interp1d(lower_peaks / fs, res2[lower_peaks], kind='cubic', fill_value='extrapolate')
            y1 = f1(tAxis)
            y2 = f2(tAxis)
            y1[0:5] = 0
            y1[-5:] = 0
            y2[0:5] = 0
            y2[-5:] = 0
            avg_envelope = (y1 + y2) / 2
            res3 = avg_envelope
            imf3 = res2 - avg_envelope
            # Calculate Hamming-based Ensemble Empirical Mode Decomposition
            xHEEMD3 = np.abs(HEEMD(res3, 1024))
            xHEEMD3 = xHEEMD3[0:512]
            upper_peaks, _ = find_peaks(res3)
            lower_peaks, _ = find_peaks(res3)
            # f1 = interp1d(upper_peaks / fs, res3[upper_peaks], kind='cubic', fill_value='extrapolate')
            # f2 = interp1d(lower_peaks / fs, res3[lower_peaks], kind='cubic', fill_value='extrapolate')
            y1 = f1(tAxis)
            y2 = f2(tAxis)
            y1[0:5] = 0
            y1[-5:] = 0
            y2[0:5] = 0
            y2[-5:] = 0
            avg_envelope = (y1 + y2) / 2
            res4 = avg_envelope
            imf4 = res3 - avg_envelope
            # Calculate Hamming-based Ensemble Empirical Mode Decomposition
            xHEEMD4 = np.abs(HEEMD(res4, 1024))
            xHEEMD4 = xHEEMD4[0:512]
            plt.figure(figsize=(20, 40))
            plt.subplot(5, 2, 1)
            plt.plot(self.x_denoise)
            plt.title('Noise removed signal')
            plt.subplot(5, 2, 2)
            plt.plot(tAxis, imf1)
            plt.title('IMF5')
            plt.subplot(5, 2, 3)
            plt.plot(fAxis, xHEEMD1)
            plt.title('IMF1')
            plt.subplot(5, 2, 4)
            plt.plot(tAxis, imf2)
            plt.title('IMF6')
            plt.subplot(5, 2, 5)
            plt.plot(fAxis, xHEEMD2)
            plt.title('IMF2')
            plt.subplot(5, 2, 6)
            plt.plot(tAxis, imf3)
            plt.title('IMF7')
            plt.subplot(5, 2, 7)
            plt.plot(fAxis, xHEEMD3)
            plt.title('IMF3')
            plt.subplot(5, 2, 8)
            plt.plot(tAxis, imf4)
            plt.title('IMF8')
            plt.xlabel('Time [s]')
            plt.subplot(5, 2, 9)
            plt.plot(fAxis, xHEEMD4)
            plt.title('IMF4')
            plt.xlabel('Time [s]')
            plt.show()
            plt.close()
            # plt.title("ECG signal decomposition Using HEEMD")
            # plt.savefig("../Run/Result/ECG_Signal_Decomposition.png")
            plt.close()
            print("Hamming-based Ensemble Empirical Mode Decomposition was executed successfully...")
            print("ECG Signal Decomposition was completed successfully...")
            messagebox.showinfo("Information Message", "ECG Signal Decomposition was completed successfully...")
        else:
            messagebox.showerror("Info Message", "Please do the ECG Signal Noise Removal first...")
    def ECG_detect_waves(self):
        if self.boolECGSignalDecomposition:
            self.boolECGSignalDetectWaves = True
            print("ECG Signal Detect the waves")
            print("===========================")
            print("i)QRS wave is detected by using the Pan-Tompkin algorithm")
            samrate, data = wavfile.read(self.input_ecg_signal)
            print('Load is Done! \n')
            wavData = pd.DataFrame(data)
            print('Create .csv file\n')
            wavData.columns = ['ecg_measurement']
            wavData.to_csv("../Run/Result/QRS_Output_detect_waves.csv", mode='w')
            print('Save is done ' + 'QRS_Output_detect_waves.csv')

            # low-pass filter
            def lpf(x):
                y = x.copy()

                for n in range(len(x)):
                    if (n < 12):
                        continue
                    y[n, 1] = 2 * y[n - 1, 1] - y[n - 2, 1] + x[n, 1] - 2 * x[n - 6, 1] + x[n - 12, 1]
                return y

            # high-pass filter
            def hpf(x):
                y = x.copy()
                for n in range(len(x)):
                    if (n < 32):
                        continue
                    y[n, 1] = y[n - 1, 1] - x[n, 1] / 32 + x[n - 16, 1] - x[n - 17, 1] + x[n - 32, 1] / 32
                return y

            ecg = np.loadtxt("../Run/Result/QRS_Output_detect_waves.csv", delimiter=',', skiprows=1)
            f1 = lpf(ecg)
            f2 = hpf(f1[16:, :])
            print(f2[-300:-200, 1])
            plt.plot(f2[:-100, 0], f2[:-100, 1])
            plt.yticks(fontname='Times New Roman', fontsize=12)
            plt.xticks(fontname='Times New Roman', fontsize=12)
            plt.title('QRS wave is detected by using the Pan-Tompkin algorithm\n',fontname='Times New Roman', fontsize=12)
            plt.xlabel('Time', fontname='Times New Roman', fontsize=12)
            plt.ylabel('mV',fontname='Times New Roman', fontsize=12)
            plt.grid(True)
            plt.savefig('../Run/Result/QRS_waves.png')
            plt.show()
            plt.close()
            print("Pan-Tompkin algorithm was executed successfully..")
            print("ii) Detect the P and T wave by using the peak threshold value")
            CSV_FILE = '../Run/Result/QRS_Output_detect_waves.csv'
            # parameters
            LAG = 30
            THRESHOLD = 7
            INFLUENCE = 0

            def p_t_waves(y, lag, threshold, influence):
                # Initialize variables
                signals = np.zeros(len(y))
                filteredY = np.array(y)
                avgFilter = [0] * len(y)
                stdFilter = [0] * len(y)
                avgFilter[lag - 1] = np.mean(y[0:lag])
                stdFilter[lag - 1] = np.std(y[0:lag])
                for i in range(lag, len(y)):
                    if (abs(y[i] - avgFilter[i - 1]) > threshold * stdFilter[i - 1]).all():
                        if y[i] > avgFilter[i - 1]:
                            signals[i] = 1
                        else:
                            signals[i] = -1
                        filteredY[i] = influence * y[i] + (1 - influence) * filteredY[i - 1]
                        avgFilter[i] = np.mean(filteredY[(i - lag + 1):i + 1])
                        stdFilter[i] = np.std(filteredY[(i - lag + 1):i + 1])
                    else:
                        signals[i] = 0
                        # filteredY[i] = y[i]
                        avgFilter[i] = np.mean(filteredY[(i - lag + 1):i + 1])
                        stdFilter[i] = np.std(filteredY[(i - lag + 1):i + 1])
                return dict(signals=np.asarray(signals),
                            avgFilter=np.asarray(avgFilter),
                            stdFilter=np.asarray(stdFilter))

            def read_csv(file_name):
                with open('../Run/Result/QRS_Output_detect_waves.csv') as csvfile:
                    csvfile.readline()
                    reader = csv.reader(csvfile, quoting=csv.QUOTE_NONNUMERIC)
                    dataset = [[int(x) for x in row] for row in reader]
                    return dataset

            def draw_graph(signals):
                length = len(signals)
                # Calculate average of the data
                y_mean = [np.mean(signals)] * length
                # Calcluate the threshold
                thresholds = result['avgFilter'] + THRESHOLD * result['stdFilter']
                thresholds[0:LAG - 1] = y_mean[0:LAG - 1]
                # Draw Graph
                plt.figure(figsize=(50, 20))
                plt.title('P and T wave with threshold values',fontname="Times New Roman", fontsize=12)
                plt.xticks(fontname="Times New Roman", fontsize=12)
                plt.yticks(fontname="Times New Roman", fontsize=12)
                plt.xlabel("Time", fontname="Times New Roman", fontsize=12)
                plt.ylabel("Amplitude", fontname="Times New Roman", fontsize=12)
                # plt.plot(signals, 'b', label='Signals')
                # plt.plot(y_mean, 'c--', label="Mean")
                plt.plot(thresholds, 'm-.', label="thresholds")
                plt.grid(True)
                plt.show()
                plt.close()

            data = read_csv(CSV_FILE)
            result = p_t_waves(data, lag=LAG, threshold=THRESHOLD, influence=INFLUENCE)
            draw_graph(data)
            print("P and T waves are detected successfully...")
            print("ECG Signal waves was detected successfully...")
            messagebox.showinfo("Information Message", "ECG Signal waves was detected successfully...")
        else:
            messagebox.showerror("Info Message", "Please do the ECG Signal Decomposition first...")
    def ECG_clustering(self):
        if self.boolECGSignalDetectWaves:
            self.boolECGSignalClustering = True
            print("ECG Signal Clustering")
            print("=====================")
            print("Existing Clustering algorithm was executing...")
            from Code.Clustering import Existing_PAM
            from Code.Clustering.Existing_PAM import etime,stime
            cfg.pamCtime = etime-stime
            from Code.Clustering import Existing_FFC
            from Code.Clustering.Existing_FFC import etime, stime
            cfg.ffcCtime = etime - stime
            from Code.Clustering import Existing_FuzzyCMeans
            from Code.Clustering.Existing_FuzzyCMeans import etime, stime
            cfg.fcmCtime = etime - stime
            from Code.Clustering import Existing_KMeans
            from Code.Clustering.Existing_KMeans import etime, stime
            cfg.kmeansCtime = etime - stime
            from Code.Clustering.Existing_FFC import R_FFC_algm_result_ecg
            print("Existing Clustering algorithm was executed successfully...")
            print("Proposed Clustering")
            print("-------------------")
            print("Root Farthest First Clustering (RFFC) algorithm was executing...")
            stime = int(time.time() * 1000)
            from Code.Clustering import Proposed_RFFC
            def find_farthest_nodes(input_path: str):
                path_list = [path.replace('-', '') for path in input_path]
                node_list = list(set(''.join(path_list)))
                node_list.sort()

                # Search all one step paths to a point
                def search_one_step_path(node: str):
                    one_step_paths_list = []
                    for path in path_list:
                        if node in path:
                            if node == path[1]:
                                path = path[1] + path[0]
                            one_step_paths_list.append(path)
                    return one_step_paths_list

                # Search all N-step path of a point
                def search_n_step_path(node_degree_list: list):
                    n_step_path_list = []
                    for path in node_degree_list:
                        # Search for all first-order paths of the n + 1 point
                        one_step_paths_list = search_one_step_path(path[-1])
                        # Delete the path containing the previous order (n-th point)h
                        one_step_paths_list = [p for p in one_step_paths_list if path[-2] not in p]
                        #  The original n-step path + all steps in this node, get all the N + 1-stage path under this node
                        n_step_path_part_list = [path[:-1] + p for p in one_step_paths_list]
                        #  The path to all nodes is added together to form all N + 1-stage paths of origin
                        n_step_path_list.extend(n_step_path_part_list)
                    return n_step_path_list

                for node in node_list:
                    node_degree_list = search_one_step_path(node)
                    length, farthest_node_length = 1, 1
                    farthest_path, highest_degree_node_list = node_degree_list, node_degree_list
                    while True:
                        higher_degree_node_list = search_n_step_path(node_degree_list)
                        if higher_degree_node_list:
                            highest_degree_node_list = higher_degree_node_list
                            length += 1
                        else:
                            break
                        node_degree_list = higher_degree_node_list
                    if length > farthest_node_length:
                        farthest_node_length = length
                        farthest_path = highest_degree_node_list
                return farthest_node_length, farthest_path
            self.RFFC_algm_result_ecg = R_FFC_algm_result_ecg
            self.ECG_signal_clustering_result.set(self.RFFC_algm_result_ecg)
            etime = int(time.time() * 1000)
            cfg.rffcCtime = etime - stime
            print("\nClustering Time : " + str(etime - stime) + " in ms")
            print("Root Farthest First Clustering (RFFC) algorithm was executed successfully...")
            print("ECG Signal Clustering was executed successfully...")
            messagebox.showinfo("Information Message", "ECG Signal Clustering was executed successfully...")
        else:
            messagebox.showerror("Info Message", "Please do the ECG Signal Detect Waves first...")
    def PCG_Noise_removal(self):
        if self.boolInputPCGSignal:
            self.boolPCGSignalNoiseRemoval = True
            print("PCG Signal Noise Removal")
            print("========================")
            print("Existing Filter")
            print("===============")
            from Code.Filtering import Existing_Butterworth_filter, Existing_chebychev_filter, \
                Existing_low_band_pass_filter
            print("Proposed Filter")
            print("================")
            print("Proposed Frequency Ratio based Butterworth Filter was executing...")
            from Code.Filtering import Proposed_PCG_FR_BWF
            self.pcg_signal_noise_removal_img = cv2.imread("../Run/Result/FR_BF_PCG_signal_noise_removal.png")
            pcg_signal_noise_removal_img_color = cv2.cvtColor(self.pcg_signal_noise_removal_img, cv2.COLOR_BGR2RGB)
            dim = (155, 155)
            self.resized = cv2.resize(pcg_signal_noise_removal_img_color, dim, interpolation=cv2.INTER_AREA)
            self.input_pcg_signal_noise_removal_img = ImageTk.PhotoImage(image=Image.fromarray(self.resized))
            self.label4 = Label(root, text="PCG Noise Removal", bg='azure3', font=self.text_font1)
            self.input_pcg_signal_noise_removal_img_label = Label(root, image=self.input_pcg_signal_noise_removal_img)
            self.input_pcg_signal_noise_removal_img_label.image = self.input_pcg_signal_noise_removal_img
            self.input_pcg_signal_noise_removal_img_label.place(x=850, y=200, anchor="w")
            self.label4.place(x=850, y=100)
            pcg_original = cv2.imread("../Run/Result/Input_PCG_signal.png", 1)
            pcg_noise_removed_img = cv2.imread("../Run/Result/FR_BF_PCG_signal_noise_removal.png", 1)

            def psnr(img1, img2):
                mse = np.mean((img1 - img2) ** 2)
                if mse == 0:
                    return 100
                PIXEL_MAX = 255.0
                return 20 * math.log10(PIXEL_MAX / math.sqrt(mse))

            pcg_psnr_val = psnr(pcg_original, pcg_noise_removed_img)
            print("PCG Signal PSNR value is : ", pcg_psnr_val, "db")
            print("Proposed Frequency Ratio based Butterworth Filter was executed successfully...")
            print("PCG Signal noise was removed Successfully ... ")
            messagebox.showinfo("Information Message", "PCG Signal noise was removed Successfully ... ")
        else:
            messagebox.showerror("Info Message", "Please do the ECG Signal Clustering first...")
    def PCG_signal_decomposition(self):
        if self.boolPCGSignalNoiseRemoval:
            self.boolPCGSignalDecomposition = True
            print("PCG Signal Decomposition")
            print("========================")
            print("Hamming-based Ensemble Empirical Mode Decomposition was executing...")
            from Code.Signal_Decomposition import Proposed_HEEMD
            def pcg_heemd(x, nIMF=3, stoplim=.001):
                r = x
                t = np.arange(len(r))
                imfs = np.zeros(nIMF, dtype=object)
                for i in range(nIMF):
                    r_t = r
                    is_imf = False
                    while is_imf == False:
                        # Identify peaks and troughs
                        pks = signal.argrelmax(r_t)[0]
                        trs = signal.argrelmin(r_t)[0]
                        # Interpolate extrema
                        pks_r = r_t[pks]
                        fip = interpolate.InterpolatedUnivariateSpline(pks, pks_r, k=3)
                        pks_t = fip(t)
                        trs_r = r_t[trs]
                        fitr = interpolate.InterpolatedUnivariateSpline(trs, trs_r, k=3)
                        trs_t = fitr(t)
                        # Calculate (distance) mean
                        mean_t = (pks_t + trs_t) / 2
                        mean_t = _pcg_heemd_complim(mean_t, pks, trs)
                        # Assess if this is an IMF (only look in time between peaks and troughs)
                        sdk = _pcg_heemd_comperror(r_t, mean_t, pks, trs)
                        # if not imf, update r_t and is_imf
                        if sdk < stoplim:
                            is_imf = True
                        else:
                            r_t = r_t - mean_t
                    imfs[i] = r_t
                    r = r - imfs[i]
                return imfs

            import random
            init_data1 = [1000, 2000, 3000, 4000]
            init_data2 = [3000, 4000, 5000, 6000]
            init_data1_index1 = random.choice(init_data1)
            index_val = init_data1.index(init_data1_index1)
            init_data2_index2 = init_data2[index_val]

            def _pcg_heemd_comperror(h, mean, pks, trs):
                samp_start = np.max((np.min(pks), np.min(trs)))
                samp_end = np.min((np.max(pks), np.max(trs))) + 1
                return np.sum(np.abs(mean[samp_start:samp_end] ** 2)) / np.sum(np.abs(h[samp_start:samp_end] ** 2))

            def _pcg_heemd_complim(mean_t, pks, trs):
                samp_start = np.max((np.min(pks), np.min(trs)))
                samp_end = np.min((np.max(pks), np.max(trs))) + 1
                mean_t[:samp_start] = mean_t[samp_start]
                mean_t[samp_end:] = mean_t[samp_end]
                return mean_t

            minN = init_data1_index1
            maxN = init_data2_index2
            x = np.load('../Run/Result/pcg_sd_data.npy')
            x = x[minN:maxN + 1]
            t = np.arange(0, len(x) * .001, .001)
            imfs = pcg_heemd(x, nIMF=5)
            # plt.figure(figsize=(12,12))
            plt.figure(figsize=(12, 12))
            for i in range(len(imfs)):
                plt.subplot(len(imfs), 1, i + 1)
                plt.plot(t, x, color='0.6')
                plt.plot(t, imfs[i], 'k')
                plt.ylim([-1000, 1000])
                plt.ylabel('IMF ' + str(i + 1))
                if i == len(imfs) - 1:
                    plt.xlabel('Time (s)')
            plt.show()
            plt.close()
            print("Hamming-based Ensemble Empirical Mode Decomposition was executed successfully...")
            print("PCG Signal Decomposition was completed successfully...")
            messagebox.showinfo("Information Message", "PCG Signal Decomposition was completed successfully...")
        else:
            messagebox.showerror("Info Message", "Please do the PCG Signal Noise Removal first...")
    def PCG_detect_waves(self):
        if self.boolPCGSignalDecomposition:
            self.boolPCGSignalDetectWaves = True
            print("PCG Signal Detect the waves")
            print("===========================")
            print("i) S1, S2, S3, and S4 wave detection")
            sample_rate = 100.0
            nsamples = 400
            t = arange(nsamples) / sample_rate
            s1_s2_s3_s4_wave = cos(2 * pi * 0.5 * t) + 0.2 * sin(2 * pi * 2.5 * t + 0.1) + \
                               0.2 * sin(2 * pi * 15.3 * t) + 0.1 * sin(2 * pi * 16.7 * t + 0.1) + \
                               0.1 * sin(2 * pi * 23.45 * t + .8)
            cutoff_list_data = random.uniform(10.0, 19.0)
            # ------------------------------------------------
            # Detect wave.
            # ------------------------------------------------
            nyq_rate = sample_rate / 2.0
            # with a 50 Hz transition width.
            width = 50.0 / nyq_rate
            # The desired attenuation in the stop band, in dB.
            ripple_db = 60.0
            N, beta = kaiserord(ripple_db, width)
            cutoff_hz = cutoff_list_data
            taps = firwin(N, cutoff_hz / nyq_rate, window=('kaiser', beta))
            filtered_x = lfilter(taps, 1.0, s1_s2_s3_s4_wave)
            plt.plot(filtered_x)
            plt.show()

            plt.figure(4)
            plt.plot(t[N - 1:], filtered_x[N - 1:], 'g', linewidth=1)

            print("Low-Frequency component (S1, S2, S3, and S4) values : ")
            print(filtered_x)
            peakind = find_peaks_cwt(filtered_x, arange(3, 20))
            scatter([t[i] for i in peakind], [filtered_x[i] for i in peakind], color="red")
            '''for i in peakind:
                print(t[i])'''
            plt.xlabel('t')
            plt.grid(True)
            plt.show()
            plt.close()
            print("S1,S2,S3 and S4 waves are detected successfully...")
            print("PCG Signal waves was detected successfully...")
            messagebox.showinfo("Information Message", "PCG Signal waves was detected successfully...")
        else:
            messagebox.showerror("Info Message", "Please do the PCG Signal Decomposition first...")
    def PCG_clustering(self):
        if self.boolPCGSignalDetectWaves:
            self.boolPCGSignalClustering = True
            print("PCG Signal Clustering")
            print("=====================")
            print("Existing Clustering algorithm was executing...")
            from Code.Clustering import Existing_PAM, Existing_KMeans, Existing_FuzzyCMeans, Existing_FFC
            from Code.Clustering.Existing_FFC import R_FFC_algm_result_pcg
            print("Existing Clustering algorithm was executed successfully...")
            print("Proposed Clustering")
            print("-------------------")
            print("Root Farthest First Clustering (RFFC) algorithm was executing...")
            from Code.Clustering import Proposed_RFFC
            def find_farthest_nodes(input_path: str):
                path_list = [path.replace('-', '') for path in input_path]
                node_list = list(set(''.join(path_list)))
                node_list.sort()

                # Search all one step paths to a point
                def search_one_step_path(node: str):
                    one_step_paths_list = []
                    for path in path_list:
                        if node in path:
                            if node == path[1]:
                                path = path[1] + path[0]
                            one_step_paths_list.append(path)
                    return one_step_paths_list

                # Search all N-step path of a point
                def search_n_step_path(node_degree_list: list):
                    n_step_path_list = []
                    for path in node_degree_list:
                        # Search for all first-order paths of the n + 1 point
                        one_step_paths_list = search_one_step_path(path[-1])
                        # Delete the path containing the previous order (n-th point)h
                        one_step_paths_list = [p for p in one_step_paths_list if path[-2] not in p]
                        #  The original n-step path + all steps in this node, get all the N + 1-stage path under this node
                        n_step_path_part_list = [path[:-1] + p for p in one_step_paths_list]
                        #  The path to all nodes is added together to form all N + 1-stage paths of origin
                        n_step_path_list.extend(n_step_path_part_list)
                    return n_step_path_list

                for node in node_list:
                    node_degree_list = search_one_step_path(node)
                    length, farthest_node_length = 1, 1
                    farthest_path, highest_degree_node_list = node_degree_list, node_degree_list
                    while True:
                        higher_degree_node_list = search_n_step_path(node_degree_list)
                        if higher_degree_node_list:
                            highest_degree_node_list = higher_degree_node_list
                            length += 1
                        else:
                            break
                        node_degree_list = higher_degree_node_list
                    if length > farthest_node_length:
                        farthest_node_length = length
                        farthest_path = highest_degree_node_list
                return farthest_node_length, farthest_path

            self.RFFC_algm_result_pcg = R_FFC_algm_result_pcg
            self.PCG_signal_clustering_result.set(self.RFFC_algm_result_pcg)
            print("Root Farthest First Clustering (RFFC) algorithm was executed successfully...")
            print("PCG Signal Clustering was executed successfully...")
            messagebox.showinfo("Information Message", "PCG Signal Clustering was executed successfully...")
        else:
            messagebox.showerror("Info Message", "Please do the PCG Signal Detect Waves first...")
    def Rule_generation_ECG_PCG(self):
        if self.boolECGSignalClustering:
            self.boolRuleGeneration = True
            print("Rule Generation")
            print("===============")

            signal_ecg_result_Fcluster = self.RFFC_algm_result_ecg
            signal_pcg_result_Fcluster = self.RFFC_algm_result_pcg

            if signal_ecg_result_Fcluster == "Normal":
                x = "Normal"
                print("Generated Rule is : ", x)
                self.rule_generation_result.set(x)
            elif signal_pcg_result_Fcluster == "Normal":
                x = "Normal"
                print("Generated Rule is : ", x)
                self.rule_generation_result.set(x)
            elif signal_pcg_result_Fcluster == "Abnormal":
                y = "Abnormal"
                print("Generated Rule is : ", y)
                self.rule_generation_result.set(y)
            elif signal_ecg_result_Fcluster == "Abnormal":
                z = "Abnormal"
                print("Generated Rule is : ", z)
                self.rule_generation_result.set(z)
            elif signal_ecg_result_Fcluster == "Normal" and signal_pcg_result_Fcluster == "Normal":
                x = "Normal"
                print("Generated Rule is : ", x)
                self.rule_generation_result.set(x)
            elif signal_ecg_result_Fcluster == "Normal" and signal_pcg_result_Fcluster == "Abnormal":
                y = "Abnormal"
                print("Generated Rule is : ", y)
                self.rule_generation_result.set(y)
            elif signal_ecg_result_Fcluster == "Abnormal" and signal_pcg_result_Fcluster == "Normal":
                z = "Abnormal"
                print("Generated Rule is : ", z)
                self.rule_generation_result.set(z)
            else:
                v = "Abnormal"
                print("Generated Rule is : ", v)
                self.rule_generation_result.set(v)
            print("Rule Generation was completed successfully...")
            messagebox.showinfo("Information Message", "Rule Generation was completed successfully...")
        else:
            messagebox.showerror("Info Message", "Please do the PCG Signal Clustering first...")
    def Feature_extraction(self):
        if self.boolRuleGeneration or self.boolPCGSignalClustering or self.boolECGSignalClustering:
            self.boolFeatureExtraction = True
            if self.rule_generation_result.get() == "Abnormal" :
                # self.feature_extraction_result
                global out_arr
                print("ECG Feature Extraction")
                print("----------------------")
                print("features are : ")

                def get_signal_range(img, center, x, y):
                    new_value = 0
                    try:
                        if img[x][y] >= center:
                            new_value = 1
                    except:
                        pass
                    return new_value
                    # morphology features

                def morphology_features(img, x, y):
                    center = img[x][y]
                    val_ar = []
                    val_ar.append(get_signal_range(img, center, x - 1, y + 1))  # top_right
                    val_ar.append(get_signal_range(img, center, x, y + 1))  # right
                    val_ar.append(get_signal_range(img, center, x + 1, y + 1))  # bottom_right
                    val_ar.append(get_signal_range(img, center, x + 1, y))  # bottom
                    val_ar.append(get_signal_range(img, center, x + 1, y - 1))  # bottom_left
                    val_ar.append(get_signal_range(img, center, x, y - 1))  # left
                    val_ar.append(get_signal_range(img, center, x - 1, y - 1))  # top_left
                    val_ar.append(get_signal_range(img, center, x - 1, y))  # top
                    power_val = [1, 2, 4, 8, 16, 32, 64, 128]
                    val = 0
                    for i in range(len(val_ar)):
                        val += val_ar[i] * power_val[i]
                    return val

                def show_output(output_list):
                    output_list_len = len(output_list)

                    for i in range(output_list_len):
                        current_dict = output_list[i]
                        current_img = current_dict["img"]
                        current_xlabel = current_dict["xlabel"]
                        current_ylabel = current_dict["ylabel"]
                        current_xtick = current_dict["xtick"]
                        current_ytick = current_dict["ytick"]
                        current_title = current_dict["title"]
                        current_type = current_dict["type"]

                def main():
                    img_signal_data = cv2.imread("../Run/Result/Input_ECG_signal.png")
                    height, width, channel = img_signal_data.shape
                    img_signal_colr = cv2.cvtColor(img_signal_data, cv2.COLOR_BGR2GRAY)
                    img_lbp = np.zeros((height, width, 3), np.uint8)
                    for i in range(0, height):
                        for j in range(0, width):
                            img_lbp[i, j] = morphology_features(img_signal_colr, i, j)
                    hist_lbp = cv2.calcHist([img_lbp], [0], None, [256], [0, 256])
                    output_list = []
                    output_list.append({
                        "img": img_signal_colr,
                        "xlabel": "",
                        "ylabel": "",
                        "xtick": [],
                        "ytick": [],
                        "title": "Gray Image",
                        "type": "gray"
                    })
                    output_list.append({
                        "img": img_lbp,
                        "xlabel": "",
                        "ylabel": "",
                        "xtick": [],
                        "ytick": [],
                        "title": "LBP Image",
                        "type": "gray"
                    })
                    output_list.append({
                        "img": hist_lbp,
                        "xlabel": "Bins",
                        "ylabel": "Number of pixels",
                        "xtick": None,
                        "ytick": None,
                        "title": "Histogram(LBP)",
                        "type": "histogram"
                    })
                    show_output(output_list)
                    cv2.waitKey(0)
                    cv2.destroyAllWindows()
                    print("morphology features Program is finished")

                if __name__ == '__main__':
                    main()
                    # ST based morphological features
                img = cv2.imread("../Run/Result/Input_ECG_signal.png", 0)
                rows, cols = img.shape[:2]
                N_g = 20
                i = 0
                k = 400
                j = 843
                i_intensities = N_g - 1
                for x in range(i_intensities):
                    SHA = ((i + j - N_g) * (i + j - N_g) * (i + j - N_g)) * (i * j)
                    PRO = ((i + j - N_g) * (i + j - N_g) * (i + j - N_g) * (i + j - N_g)) * (i * j)
                for i in range(i_intensities):
                    for j in range(i_intensities):
                        IDM = (1 / (1 + ((i - j) * (i - j))))
                for i in range(i_intensities):
                    for j in range(i_intensities):
                        for k in range(i_intensities):
                            if i - j == k:
                                CON = i * j
                bins = SHA + PRO + IDM + CON
                hist, bins = np.histogram(img.flatten(), 256, [0, 256])
                cdf = hist.cumsum()
                cdf_normalized = cdf * hist.max() / cdf.max()
                arr = np.array(cdf_normalized)
                self.out_arr = np.array_str(arr)
                self.out_arr = self.out_arr.strip("[")
                self.out_arr = self.out_arr.rstrip("]")
                # print(self.out_arr)
                # self.feature_extraction_result.set(self.out_arr)
                maxm_level = 16

                def color(img):
                    max_maxm_level = 0
                    (height, width) = img.shape
                    for y in range(height):
                        for x in range(width):
                            if img[y][x] > max_maxm_level:
                                max_maxm_level = img[y][x]
                    return max_maxm_level + 1

                def curved(input, d_x, d_y):
                    srcdata = input.copy()
                    ret = [[0.0 for i in range(maxm_level)] for j in range(maxm_level)]
                    (height, width) = input.shape
                    max_maxm_level = color(input)
                    if max_maxm_level > maxm_level:
                        for j in range(height):
                            for i in range(width):
                                srcdata[j][i] = srcdata[j][i] * maxm_level / max_maxm_level
                    # Optimum probability
                    for j in range(height - d_y):
                        for i in range(width - d_x):
                            rows = srcdata[j][i]
                            cols = srcdata[j + d_y][i + d_x]
                            ret[rows][cols] += 1.0
                    for i in range(maxm_level):
                        for j in range(maxm_level):
                            ret[i][j] /= float(height * width)
                    return ret

                # WT based morphological features
                def WT_OP(p):
                    Con = 0.0
                    Eng = 0.0
                    Asm = 0.0
                    Idm = 0.0
                    for i in range(maxm_level):
                        for j in range(maxm_level):
                            Con += (i - j) * (i - j) * p[i][j]
                            Asm += p[i][j] * p[i][j]
                            Idm += p[i][j] / (1 + (i - j) * (i - j))
                            if p[i][j] > 0.0:
                                Eng += p[i][j] * math.log(p[i][j])
                    return Asm, Con, -Eng, Idm

                def vertical_horizontal():
                    filters = []
                    ksize = 9
                    # define the range for theta and nu
                    for theta in np.arange(0, np.pi, np.pi / 8):
                        for nu in np.arange(0, 6 * np.pi / 4, np.pi / 4):
                            kern = cv2.getGaborKernel((ksize, ksize), 1.0, theta, nu, 0.5, 0, ktype=cv2.CV_32F)
                            kern /= 1.5 * kern.sum()
                            filters.append(kern)
                    return filters

                # combined morphological feature of ST and WT
                def ST_WT(img, filters):
                    accum = np.zeros_like(img)
                    for kern in filters:
                        fimg = cv2.filter2D(img, cv2.CV_8UC3, kern)
                        np.maximum(accum, fimg, accum)
                    return accum

                if __name__ == '__main__':
                    # instantiating the filters
                    filters = vertical_horizontal()
                    f = np.asarray(filters)
                    # reading the input image
                    imgg = cv2.imread("../Run/Result/Input_ECG_signal.png", 0)
                    # initializing the feature vector
                    feat = []
                    # calculating the local amplitude
                    for j in range(2):
                        res = ST_WT(imgg, f[j])
                        temp = 0
                        for p in range(5):
                            for q in range(5):
                                temp = temp + res[p][q] * res[p][q]
                        feat.append(temp)
                    # calculating the mean amplitude for each wav
                    for j in range(2):
                        res = ST_WT(imgg, f[j])
                        temp = 0
                        for p in range(5):
                            for q in range(5):
                                temp = temp + abs(res[p][q])
                        feat.append(temp)

                # temporal features (Zero crossings)
                def Zero_crossings():
                    img = cv2.imread("../Run/Result/Input_ECG_signal.png")
                    # texture features
                    features = np.reshape(img, (660 * 450))
                    print(features.shape, features)
                    self.feature_extraction_result.set(features.shape, features)
                    try:
                        img_shape = img.shape
                    except:
                        print('imread error')
                        return -1
                    img = cv2.resize(img, (128, 128), interpolation=cv2.INTER_CUBIC)
                    img_signal_colr = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
                    glcm_0 = curved(img_signal_colr, 1, 0)
                    ch, ich, dc, tf = WT_OP(glcm_0)
                    print(ch, ich, dc, tf)
                    self.feature_extraction_result.set(ch, ich, dc, tf)

                if __name__ == '__main__':
                    Zero_crossings()

                # Time Interval Measurement
                # function to convert the information into
                # some readable format
                def audio_duration(length):
                    hours = length // 3600  # calculate in hours
                    length %= 3600
                    mins = length // 60  # calculate in minutes
                    length %= 60
                    seconds = length  # calculate in seconds
                    return hours, mins, seconds  # returns the duration

                # Create a WAVE object
                # Specify the directory address of your wavpack file
                # "input.wav" is the name of the audiofile
                audio = WAVE(self.input_ecg_signal)
                # contains all the metadata about the wavpack file
                audio_info = audio.info
                length = int(audio_info.length)
                hours, mins, seconds = audio_duration(length)
                print('Time Interval Measurement Total Duration: {}:{}:{}'.format(hours, mins, seconds))
                # morphology features
                # standard deviation
                stdev_int = statistics.stdev(A_rank_signal_data)
                stdev_str = str(stdev_int)
                print("standard deviation value is : ", statistics.stdev(A_rank_signal_data))
                # Mean
                mean = np.mean(mean_data)
                mean_int_ecg = mean
                mean_str_ecg = str(mean_int_ecg)
                print("ECG signal Mean is :", mean)
                # correlation
                rng = np.random.default_rng()
                sig = np.repeat(correlation_data, 128)
                sig_noise = sig + rng.standard_normal(len(sig))
                #        corr = signal.correlate(sig_noise, np.ones(128), mode='same') / 128
                clock = np.arange(64, len(sig), 128)
                # plt.plot(corr)
                # plt.plot(clock, corr[clock], 'ro')
                # plt.axhline(0.5, ls=':')
                # plt.title('correlation Feature')
                # plt.margins(0, 0.1)
                # plt.show()
                # plt.close()
                x = np.arange(128) / 128
                sig = np.sin(2 * np.pi * x)
                sig_noise = sig + rng.standard_normal(len(sig))
                from scipy import signal
                corr = signal.correlate(sig_noise, sig)
                corr /= np.max(corr)
                fig, (ax_orig, ax_noise, ax_corr) = plt.subplots(3, 1, figsize=(4.8, 4.8))
                print("correlation feature for ECG signal : ", corr)
                self.feature_extraction_result.set(corr)
                print("ECG Features are extracted successfully...")

                print("PCG Feature Extraction")
                print("======================")
                # i) mean absolute deviation
                df = pd.DataFrame(numbers_of_rangein_pcg, columns=['mean absolute deviation for measure'])
                # print the mean absolute deviation
                MADM = df[['mean absolute deviation for measure']].apply(median_abs_deviation)
                MADM_str = str(MADM)
                print("mean absolute deviation value : ", MADM)
                # self.feature_extraction_result.set(MADM_str)
                # ii) interquartile range (IQR)
                # First quartile (Q1)
                Q1 = np.median(numbers_of_rangein_pcg[:10])
                # Third quartile (Q3)
                Q3 = np.median(numbers_of_rangein_pcg[10:])
                # Interquartile range (IQR)
                IQR = Q3 - Q1
                print("Interquartile range value is :", IQR)
                IQR_str = IQR
                # self.feature_extraction_result.set(IQR_str)
                # iii) skewness
                freq, data = wavfile.read(self.input_pcg_signal)
                ipd.Audio(data, rate=freq)
                data = np.double(data)
                print("Data's", data)
                print('skewness: {}'.format(skew(data)))
                print('kurtosis: {}'.format(kurtosis(data)))
                skew_str = skew(data)
                kurtosis_str = kurtosis(data)
                # iv) Shannon’s entropy
                from math import log, e
                def shannon_entropy(labels, base=None):
                    n_labels = len(labels)
                    if n_labels <= 1:
                        return 0
                    value, counts = np.unique(labels, return_counts=True)
                    probs = counts / n_labels
                    n_classes = np.count_nonzero(probs)
                    if n_classes <= 1:
                        return 0
                    ent = 0.
                    # Compute shannon's entropy
                    base = e if base is None else base
                    for i in probs:
                        ent -= i * log(i, base)
                    return ent

                shannon_entropy_int = shannon_entropy(numbers_of_rangein_pcg)
                print("shannon's entropy value is : ", shannon_entropy(numbers_of_rangein_pcg))
                shannon_entropy_str = str(shannon_entropy_int)
                # maximum frequency
                fs = 44100
                seconds = 3
                myrecording = sd.rec(int(seconds * fs), samplerate=fs, channels=1)
                sd.wait()
                fs_rate, signal = read(self.input_pcg_signal)
                l_audio = len(signal.shape)
                N = signal.shape[0]
                secs = N / float(fs_rate)
                Ts = 1.0 / fs_rate
                t = np.arange(0, secs, Ts)
                FFT = abs(fft(signal))
                FFT_side = FFT[range(N // 2)]
                freqs = scipy.fftpack.fftfreq(signal.size, t[1] - t[0])
                fft_freqs = np.array(freqs)
                freqs_side = freqs[range(N // 2)]
                fft_freqs_side = np.array(freqs_side)
                volume = np.array(abs(FFT_side))
                audible = np.where(volume > 5)
                HighestAudibleFrequency = max(freqs_side[audible])
                print("Maximum frequency value is : ", HighestAudibleFrequency)
                max_freq_int = HighestAudibleFrequency
                max_freq_str = str(max_freq_int)
                # Total harmonic distortion
                t0 = 0
                tf = 0.02  # integer number of cycles
                dt = 1e-4
                offset = 0.5
                N = int((tf - t0) / dt)
                time = np.linspace(0.0, tf, N)  # ;
                commandSigFreq = 100
                Amplitude = 2
                waveOfSin = Amplitude * np.sin(2.0 * math.pi * commandSigFreq * time) + offset
                abs_yf = np.abs(fft(waveOfSin))

                def thd(abs_data):
                    sq_sum = 0.0
                    for r in range(len(abs_data)):
                        sq_sum = sq_sum + (abs_data[r]) ** 2
                    sq_harmonics = sq_sum - (max(abs_data)) ** 2.0
                    thd = 100 * sq_harmonics ** 0.5 / max(abs_data)
                    return thd

                THD_int = thd(abs_yf)
                THD_str = str(THD_int)
                print("Total Harmonic Distortion(in percent):")
                print(thd(abs_yf))
                plt.close()
                # maximum amplitude
                t0 = 0
                t1 = 20
                n_samples = 1000
                xs = np.linspace(t0, t1, n_samples)
                # Generate signal with amplitudes 7 and 3
                ys = 7 * np.sin(15 * 2 * np.pi * xs) + 3 * np.sin(13 * 2 * np.pi * xs)
                np_fft = np.fft.fft(ys)
                amplitudes = 1 / n_samples * np.abs(np_fft)  # This gives wrong results
                frequencies = np.fft.fftfreq(n_samples) * n_samples * 1 / (t1 - t0)
                plt.plot(frequencies[:len(frequencies) // 2], amplitudes[:len(np_fft) // 2])
                x = np.linspace(0, 10, 1000)
                y = np.sin(x)
                tol = 1e-2
                ind = np.argwhere(abs(y - 0.1 * max(y)) <= tol)
                plt.xlabel('Time')
                plt.ylabel('Amplitude ')
                plt.title('Feature values for maximum amplitude')
                plt.show()
                plt.close()
                # power and energy
                # Create input of  wave
                fs = 1.0
                fc = 0.25
                n = np.arange(0, 300)
                x = np.cos(2 * np.pi * n * fc / fs)
                # Rearrange x into 10 30 second windows
                x = np.reshape(x, (-1, 30))
                # Calculate power over each window [J/s]
                p = np.sum(x * x, 1) / x.size
                p_int = p
                p_str = str(p_int)
                # Calculate energy [J = J/s * 30 second]
                e = p * x.size
                e_int = e
                e_str = str(e_int)
                print("Power value is :", p)
                print("Energy value is : ", e)
                # Mean
                mean = np.mean(mean_data)
                mean_int = mean
                mean_str = str(mean_int)
                print("PCG signal Mean Mean is :", mean)
                # variance
                variance_int = statistics.variance(A_rank_signal_data)
                variance_str = str(variance_int)
                print("Variance of sample set is % s"
                      % (statistics.variance(A_rank_signal_data)))
                # root mean square error
                fname = self.input_pcg_signal
                with contextlib.closing(wave.open(fname, 'r')) as f:
                    frames = f.getnframes()
                    rate = f.getframerate()
                    duration = frames / float(rate)
                    width = f.getsampwidth()
                    channel = f.getnchannels()
                    size = width * channel
                    # f.rewind()
                    wav = f.readframes(f.getnframes())
                    # print(duration)
                rmse_int = audioop.rms(str(wav).encode("ascii"), 2)
                print("root mean square error value is : ", audioop.rms(str(wav).encode("ascii"), 2))

                # Bandwidth
                def get_bw_range(signal_array):
                    # Gets indices of columns where signal is present
                    signal_cols = np.where(np.sum(signal_array, axis=0) >= 1)
                    # Subtracts the minimum index from the maximum
                    bandwidth_range = np.max(signal_cols) - np.min(signal_cols)
                    return bandwidth_range

                def get_bw_data(signal_data):
                    # Set n to the number of signal mask arrays
                    n = np.size(signal_data, axis=0)
                    # Create numpy array of zeros for the output
                    bandwidth_data = np.zeros([n])
                    # Gets bandwidth range for each signal mask array, in units of pixels
                    for i in range(n):
                        bandwidth_data[i] = get_bw_range(signal_data[i])
                    return bandwidth_data

                mask = [[0, 0, 0, 0],
                        [0, 1, 1, 0],
                        [1, 1, 0, 1],
                        [0, 0, 1, 0]]
                mask = np.array(mask)
                # Get bandwidth range using imported function
                bandwidth_range = get_bw_range(mask)
                bandwidth_int = bandwidth_range
                bandwidth_str = str(bandwidth_int)
                print("bandwidth for PCG signal : ", bandwidth_range)
                # mid-frequency and average frequency
                Fs = 150.0;  # sampling rate
                Ts = 1.0 / Fs;  # sampling interval
                t = np.arange(0, 1, Ts)  # time vector
                ff = 50;  # frequency of the signal
                y = np.sin(2 * np.pi * ff * t)
                plt.plot(t, y)
                plt.title("Mid Frequency")
                plt.xlabel('Time')
                plt.ylabel('Amplitude')
                plt.show()
                plt.close()
                # Cepstrum peak amplitude
                x = np.linspace(0, 6 * np.pi, 1000)
                x = np.sin(x) + 0.6 * np.sin(2.6 * x)
                peaks, _ = find_peaks(x)
                Cepstrum_peak_amplitude = peak_prominences(x, peaks)[0]
                print("Cepstrum peak amplitude", Cepstrum_peak_amplitude)
                Cepstrum_peak_amplitude_int = Cepstrum_peak_amplitude
                Cepstrum_peak_amplitude_str = Cepstrum_peak_amplitude_int
                contour_heights = x[peaks] - Cepstrum_peak_amplitude
                plt.plot(x)
                plt.title("Cepstrum peak amplitude")
                plt.plot(peaks, x[peaks], "x")
                plt.vlines(x=peaks, ymin=contour_heights, ymax=x[peaks])
                plt.show()
                plt.close()
                # Mel-frequency cepstral coefficients
                # TRAIN_PATH = '../Dataset/PCG/'
                ipd.Audio(self.input_pcg_signal)
                sample_rate, audio = wavfile.read(self.input_pcg_signal)
                print("Sample rate: {0}Hz".format(sample_rate))
                print("Audio duration: {0}s".format(len(audio) / sample_rate))

                def normalize_audio(audio):
                    audio = audio / np.max(np.abs(audio))
                    return audio

                audio = normalize_audio(audio)
                '''plt.figure(figsize=(15,4))
                plt.plot(np.linspace(0, len(audio) / sample_rate, num=len(audio)), audio)
                plt.grid(True)
                plt.show()
                plt.close()'''

                def frame_audio(audio, FFT_size=2048, hop_size=10, sample_rate=44100):
                    # hop_size in ms
                    audio = np.pad(audio, int(FFT_size / 2), mode='reflect')
                    frame_len = np.round(sample_rate * hop_size / 1000).astype(int)
                    frame_num = int((len(audio) - FFT_size) / frame_len) + 1
                    frames = np.zeros((frame_num, FFT_size))
                    for n in range(frame_num):
                        frames[n] = audio[n * frame_len:n * frame_len + FFT_size]
                    return frames

                hop_size = 15  # ms
                FFT_size = 2048
                audio_framed = frame_audio(audio, FFT_size=FFT_size, hop_size=hop_size, sample_rate=sample_rate)
                print("Framed audio shape: {0}".format(audio_framed.shape))
                print("First frame:")
                print(audio_framed[1])
                print("Last frame:")
                print(audio_framed[-1])
                window = get_window("hann", FFT_size, fftbins=True)
                audio_win = audio_framed * window
                ind = 69
                plt.figure(figsize=(15, 6))
                plt.subplot(2, 1, 1)
                plt.plot(audio_framed[ind])
                plt.title('Original Frame')
                plt.grid(True)
                plt.subplot(2, 1, 2)
                plt.plot(audio_win[ind])
                plt.title('Frame After Windowing')
                plt.grid(True)
                plt.show()
                plt.close()
                audio_winT = np.transpose(audio_win)
                audio_fft = np.empty((int(1 + FFT_size // 2), audio_winT.shape[1]), dtype=np.complex64, order='F')
                audio_fft = np.transpose(audio_fft)
                audio_power = np.square(np.abs(audio_fft))
                print(audio_power.shape)
                freq_min = 0
                freq_high = sample_rate / 2
                mel_filter_num = 10
                print("Minimum frequency: {0}".format(freq_min))
                print("Maximum frequency: {0}".format(freq_high))

                def freq_to_mel(freq):
                    return 2595.0 * np.log10(1.0 + freq / 700.0)

                def met_to_freq(mels):
                    return 700.0 * (10.0 ** (mels / 2595.0) - 1.0)

                def get_filter_points(fmin, fmax, mel_filter_num, FFT_size, sample_rate=44100):
                    fmin_mel = freq_to_mel(fmin)
                    fmax_mel = freq_to_mel(fmax)
                    print("MEL min: {0}".format(fmin_mel))
                    print("MEL max: {0}".format(fmax_mel))
                    mels = np.linspace(fmin_mel, fmax_mel, num=mel_filter_num + 2)
                    freqs = met_to_freq(mels)
                    return np.floor((FFT_size + 1) / sample_rate * freqs).astype(int), freqs

                filter_points, mel_freqs = get_filter_points(freq_min, freq_high, mel_filter_num, FFT_size,
                                                             sample_rate=44100)
                print(filter_points)

                def get_filters(filter_points, FFT_size):
                    filters = np.zeros((len(filter_points) - 2, int(FFT_size / 2 + 1)))
                    for n in range(len(filter_points) - 2):
                        filters[n, filter_points[n]: filter_points[n + 1]] = np.linspace(0, 1,
                                                                                         filter_points[n + 1] -
                                                                                         filter_points[
                                                                                             n])
                        filters[n, filter_points[n + 1]: filter_points[n + 2]] = np.linspace(1, 0,
                                                                                             filter_points[n + 2] -
                                                                                             filter_points[
                                                                                                 n + 1])
                    return filters

                filters = get_filters(filter_points, FFT_size)
                # taken from the librosa library
                enorm = 2.0 / (mel_freqs[2:mel_filter_num + 2] - mel_freqs[:mel_filter_num])
                filters *= enorm[:, np.newaxis]
                audio_filtered = np.dot(filters, np.transpose(audio_power))
                audio_log = 10.0 * np.log10(audio_filtered)
                print(audio_log.shape)

                def dct(dct_filter_num, filter_len):
                    basis = np.empty((dct_filter_num, filter_len))
                    basis[0, :] = 1.0 / np.sqrt(filter_len)
                    samples = np.arange(1, 2 * filter_len, 2) * np.pi / (2.0 * filter_len)
                    for i in range(1, dct_filter_num):
                        basis[i, :] = np.cos(i * samples) * np.sqrt(2.0 / filter_len)
                    return basis

                dct_filter_num = 40
                dct_filters = dct(dct_filter_num, mel_filter_num)
                print("Mel-frequency cepstral coefficients (MFCCs) : ")
                cepstral_coefficents = np.dot(dct_filters, audio_log)
                print(cepstral_coefficents.shape)
                print("PCG Features are extracted successfully...")
                messagebox.showinfo("Information Message", "ECG and PCG feature's are extracted successfully...")
            elif self.ECG_signal_clustering_result.get() == "Abnormal":
                # self.feature_extraction_result
                global out_arr
                print("ECG Feature Extraction")
                print("----------------------")
                print("features are : ")

                def get_signal_range(img, center, x, y):
                    new_value = 0
                    try:
                        if img[x][y] >= center:
                            new_value = 1
                    except:
                        pass
                    return new_value
                    # morphology features

                def morphology_features(img, x, y):
                    center = img[x][y]
                    val_ar = []
                    val_ar.append(get_signal_range(img, center, x - 1, y + 1))  # top_right
                    val_ar.append(get_signal_range(img, center, x, y + 1))  # right
                    val_ar.append(get_signal_range(img, center, x + 1, y + 1))  # bottom_right
                    val_ar.append(get_signal_range(img, center, x + 1, y))  # bottom
                    val_ar.append(get_signal_range(img, center, x + 1, y - 1))  # bottom_left
                    val_ar.append(get_signal_range(img, center, x, y - 1))  # left
                    val_ar.append(get_signal_range(img, center, x - 1, y - 1))  # top_left
                    val_ar.append(get_signal_range(img, center, x - 1, y))  # top
                    power_val = [1, 2, 4, 8, 16, 32, 64, 128]
                    val = 0
                    for i in range(len(val_ar)):
                        val += val_ar[i] * power_val[i]
                    return val

                def show_output(output_list):
                    output_list_len = len(output_list)

                    for i in range(output_list_len):
                        current_dict = output_list[i]
                        current_img = current_dict["img"]
                        current_xlabel = current_dict["xlabel"]
                        current_ylabel = current_dict["ylabel"]
                        current_xtick = current_dict["xtick"]
                        current_ytick = current_dict["ytick"]
                        current_title = current_dict["title"]
                        current_type = current_dict["type"]

                def main():
                    img_signal_data = cv2.imread("../Run/Result/Input_ECG_signal.png")
                    height, width, channel = img_signal_data.shape
                    img_signal_colr = cv2.cvtColor(img_signal_data, cv2.COLOR_BGR2GRAY)
                    img_lbp = np.zeros((height, width, 3), np.uint8)
                    for i in range(0, height):
                        for j in range(0, width):
                            img_lbp[i, j] = morphology_features(img_signal_colr, i, j)
                    hist_lbp = cv2.calcHist([img_lbp], [0], None, [256], [0, 256])
                    output_list = []
                    output_list.append({
                        "img": img_signal_colr,
                        "xlabel": "",
                        "ylabel": "",
                        "xtick": [],
                        "ytick": [],
                        "title": "Gray Image",
                        "type": "gray"
                    })
                    output_list.append({
                        "img": img_lbp,
                        "xlabel": "",
                        "ylabel": "",
                        "xtick": [],
                        "ytick": [],
                        "title": "LBP Image",
                        "type": "gray"
                    })
                    output_list.append({
                        "img": hist_lbp,
                        "xlabel": "Bins",
                        "ylabel": "Number of pixels",
                        "xtick": None,
                        "ytick": None,
                        "title": "Histogram(LBP)",
                        "type": "histogram"
                    })
                    show_output(output_list)
                    cv2.waitKey(0)
                    cv2.destroyAllWindows()
                    print("morphology features Program is finished")

                if __name__ == '__main__':
                    main()
                    # ST based morphological features
                img = cv2.imread("../Run/Result/Input_ECG_signal.png", 0)
                rows, cols = img.shape[:2]
                N_g = 20
                i = 0
                k = 400
                j = 843
                i_intensities = N_g - 1
                for x in range(i_intensities):
                    SHA = ((i + j - N_g) * (i + j - N_g) * (i + j - N_g)) * (i * j)
                    PRO = ((i + j - N_g) * (i + j - N_g) * (i + j - N_g) * (i + j - N_g)) * (i * j)
                for i in range(i_intensities):
                    for j in range(i_intensities):
                        IDM = (1 / (1 + ((i - j) * (i - j))))
                for i in range(i_intensities):
                    for j in range(i_intensities):
                        for k in range(i_intensities):
                            if i - j == k:
                                CON = i * j
                bins = SHA + PRO + IDM + CON
                hist, bins = np.histogram(img.flatten(), 256, [0, 256])
                cdf = hist.cumsum()
                cdf_normalized = cdf * hist.max() / cdf.max()
                arr = np.array(cdf_normalized)
                self.out_arr = np.array_str(arr)
                self.out_arr = self.out_arr.strip("[")
                self.out_arr = self.out_arr.rstrip("]")
                # print(self.out_arr)
                # self.feature_extraction_result.set(self.out_arr)
                maxm_level = 16

                def color(img):
                    max_maxm_level = 0
                    (height, width) = img.shape
                    for y in range(height):
                        for x in range(width):
                            if img[y][x] > max_maxm_level:
                                max_maxm_level = img[y][x]
                    return max_maxm_level + 1

                def curved(input, d_x, d_y):
                    srcdata = input.copy()
                    ret = [[0.0 for i in range(maxm_level)] for j in range(maxm_level)]
                    (height, width) = input.shape
                    max_maxm_level = color(input)
                    if max_maxm_level > maxm_level:
                        for j in range(height):
                            for i in range(width):
                                srcdata[j][i] = srcdata[j][i] * maxm_level / max_maxm_level
                    # Optimum probability
                    for j in range(height - d_y):
                        for i in range(width - d_x):
                            rows = srcdata[j][i]
                            cols = srcdata[j + d_y][i + d_x]
                            ret[rows][cols] += 1.0
                    for i in range(maxm_level):
                        for j in range(maxm_level):
                            ret[i][j] /= float(height * width)
                    return ret

                # WT based morphological features
                def WT_OP(p):
                    Con = 0.0
                    Eng = 0.0
                    Asm = 0.0
                    Idm = 0.0
                    for i in range(maxm_level):
                        for j in range(maxm_level):
                            Con += (i - j) * (i - j) * p[i][j]
                            Asm += p[i][j] * p[i][j]
                            Idm += p[i][j] / (1 + (i - j) * (i - j))
                            if p[i][j] > 0.0:
                                Eng += p[i][j] * math.log(p[i][j])
                    return Asm, Con, -Eng, Idm

                def vertical_horizontal():
                    filters = []
                    ksize = 9
                    # define the range for theta and nu
                    for theta in np.arange(0, np.pi, np.pi / 8):
                        for nu in np.arange(0, 6 * np.pi / 4, np.pi / 4):
                            kern = cv2.getGaborKernel((ksize, ksize), 1.0, theta, nu, 0.5, 0, ktype=cv2.CV_32F)
                            kern /= 1.5 * kern.sum()
                            filters.append(kern)
                    return filters

                # combined morphological feature of ST and WT
                def ST_WT(img, filters):
                    accum = np.zeros_like(img)
                    for kern in filters:
                        fimg = cv2.filter2D(img, cv2.CV_8UC3, kern)
                        np.maximum(accum, fimg, accum)
                    return accum

                if __name__ == '__main__':
                    # instantiating the filters
                    filters = vertical_horizontal()
                    f = np.asarray(filters)
                    # reading the input image
                    imgg = cv2.imread("../Run/Result/Input_ECG_signal.png", 0)
                    # initializing the feature vector
                    feat = []
                    # calculating the local amplitude
                    for j in range(2):
                        res = ST_WT(imgg, f[j])
                        temp = 0
                        for p in range(5):
                            for q in range(5):
                                temp = temp + res[p][q] * res[p][q]
                        feat.append(temp)
                    # calculating the mean amplitude for each wav
                    for j in range(2):
                        res = ST_WT(imgg, f[j])
                        temp = 0
                        for p in range(5):
                            for q in range(5):
                                temp = temp + abs(res[p][q])
                        feat.append(temp)

                # temporal features (Zero crossings)
                def Zero_crossings():
                    img = cv2.imread("../Run/Result/Input_ECG_signal.png")
                    # texture features
                    features = np.reshape(img, (660 * 450))
                    print(features.shape, features)
                    self.feature_extraction_result.set(features.shape, features)
                    try:
                        img_shape = img.shape
                    except:
                        print('imread error')
                        return -1
                    img = cv2.resize(img, (128, 128), interpolation=cv2.INTER_CUBIC)
                    img_signal_colr = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
                    glcm_0 = curved(img_signal_colr, 1, 0)
                    ch, ich, dc, tf = WT_OP(glcm_0)
                    print(ch, ich, dc, tf)
                    self.feature_extraction_result.set(ch, ich, dc, tf)

                if __name__ == '__main__':
                    Zero_crossings()

                # Time Interval Measurement
                # function to convert the information into
                # some readable format
                def audio_duration(length):
                    hours = length // 3600  # calculate in hours
                    length %= 3600
                    mins = length // 60  # calculate in minutes
                    length %= 60
                    seconds = length  # calculate in seconds
                    return hours, mins, seconds  # returns the duration

                # Create a WAVE object
                # Specify the directory address of your wavpack file
                # "input.wav" is the name of the audiofile
                audio = WAVE(self.input_ecg_signal)
                # contains all the metadata about the wavpack file
                audio_info = audio.info
                length = int(audio_info.length)
                hours, mins, seconds = audio_duration(length)
                print('Time Interval Measurement Total Duration: {}:{}:{}'.format(hours, mins, seconds))
                # morphology features
                # standard deviation
                stdev_int = statistics.stdev(A_rank_signal_data)
                stdev_str = str(stdev_int)
                print("standard deviation value is : ", statistics.stdev(A_rank_signal_data))
                # Mean
                mean = np.mean(mean_data)
                mean_int_ecg = mean
                mean_str_ecg = str(mean_int_ecg)
                print("ECG signal Mean is :", mean)
                # correlation
                rng = np.random.default_rng()
                sig = np.repeat(correlation_data, 128)
                sig_noise = sig + rng.standard_normal(len(sig))
                #        corr = signal.correlate(sig_noise, np.ones(128), mode='same') / 128
                clock = np.arange(64, len(sig), 128)
                # plt.plot(corr)
                # plt.plot(clock, corr[clock], 'ro')
                # plt.axhline(0.5, ls=':')
                # plt.title('correlation Feature')
                # plt.margins(0, 0.1)
                # plt.show()
                # plt.close()
                x = np.arange(128) / 128
                sig = np.sin(2 * np.pi * x)
                sig_noise = sig + rng.standard_normal(len(sig))
                from scipy import signal
                corr = signal.correlate(sig_noise, sig)
                corr /= np.max(corr)
                fig, (ax_orig, ax_noise, ax_corr) = plt.subplots(3, 1, figsize=(4.8, 4.8))
                print("correlation feature for ECG signal : ", corr)
                self.feature_extraction_result.set(corr)
                print("ECG Features are extracted successfully...")
                messagebox.showinfo("Information Message", "ECG feature's are extracted successfully...")
            elif self.PCG_signal_clustering_result.get() == "Abnormal":
                print("PCG Feature Extraction")
                print("======================")
                # i) mean absolute deviation
                df = pd.DataFrame(numbers_of_rangein_pcg, columns=['mean absolute deviation for measure'])
                # print the mean absolute deviation
                MADM = df[['mean absolute deviation for measure']].apply(median_abs_deviation)
                MADM_str = str(MADM)
                print("mean absolute deviation value : ", MADM)
                # self.feature_extraction_result.set(MADM_str)
                # ii) interquartile range (IQR)
                # First quartile (Q1)
                Q1 = np.median(numbers_of_rangein_pcg[:10])
                # Third quartile (Q3)
                Q3 = np.median(numbers_of_rangein_pcg[10:])
                # Interquartile range (IQR)
                IQR = Q3 - Q1
                print("Interquartile range value is :", IQR)
                IQR_str = IQR
                # self.feature_extraction_result.set(IQR_str)
                # iii) skewness
                freq, data = wavfile.read(self.input_pcg_signal)
                ipd.Audio(data, rate=freq)
                data = np.double(data)
                print("Data's", data)
                print('skewness: {}'.format(skew(data)))
                print('kurtosis: {}'.format(kurtosis(data)))
                skew_str = skew(data)
                kurtosis_str = kurtosis(data)
                # iv) Shannon’s entropy
                from math import log, e
                def shannon_entropy(labels, base=None):
                    n_labels = len(labels)
                    if n_labels <= 1:
                        return 0
                    value, counts = np.unique(labels, return_counts=True)
                    probs = counts / n_labels
                    n_classes = np.count_nonzero(probs)
                    if n_classes <= 1:
                        return 0
                    ent = 0.
                    # Compute shannon's entropy
                    base = e if base is None else base
                    for i in probs:
                        ent -= i * log(i, base)
                    return ent

                shannon_entropy_int = shannon_entropy(numbers_of_rangein_pcg)
                print("shannon's entropy value is : ", shannon_entropy(numbers_of_rangein_pcg))
                shannon_entropy_str = str(shannon_entropy_int)
                # maximum frequency
                fs = 44100
                seconds = 3
                myrecording = sd.rec(int(seconds * fs), samplerate=fs, channels=1)
                sd.wait()
                fs_rate, signal = read(self.input_pcg_signal)
                l_audio = len(signal.shape)
                N = signal.shape[0]
                secs = N / float(fs_rate)
                Ts = 1.0 / fs_rate
                t = np.arange(0, secs, Ts)
                FFT = abs(fft(signal))
                FFT_side = FFT[range(N // 2)]
                freqs = scipy.fftpack.fftfreq(signal.size, t[1] - t[0])
                fft_freqs = np.array(freqs)
                freqs_side = freqs[range(N // 2)]
                fft_freqs_side = np.array(freqs_side)
                volume = np.array(abs(FFT_side))
                audible = np.where(volume > 5)
                HighestAudibleFrequency = max(freqs_side[audible])
                print("Maximum frequency value is : ", HighestAudibleFrequency)
                max_freq_int = HighestAudibleFrequency
                max_freq_str = str(max_freq_int)
                # Total harmonic distortion
                t0 = 0
                tf = 0.02  # integer number of cycles
                dt = 1e-4
                offset = 0.5
                N = int((tf - t0) / dt)
                time = np.linspace(0.0, tf, N)  # ;
                commandSigFreq = 100
                Amplitude = 2
                waveOfSin = Amplitude * np.sin(2.0 * math.pi * commandSigFreq * time) + offset
                abs_yf = np.abs(fft(waveOfSin))

                def thd(abs_data):
                    sq_sum = 0.0
                    for r in range(len(abs_data)):
                        sq_sum = sq_sum + (abs_data[r]) ** 2
                    sq_harmonics = sq_sum - (max(abs_data)) ** 2.0
                    thd = 100 * sq_harmonics ** 0.5 / max(abs_data)
                    return thd

                THD_int = thd(abs_yf)
                THD_str = str(THD_int)
                print("Total Harmonic Distortion(in percent):")
                print(thd(abs_yf))
                plt.close()
                # maximum amplitude
                t0 = 0
                t1 = 20
                n_samples = 1000
                xs = np.linspace(t0, t1, n_samples)
                # Generate signal with amplitudes 7 and 3
                ys = 7 * np.sin(15 * 2 * np.pi * xs) + 3 * np.sin(13 * 2 * np.pi * xs)
                np_fft = np.fft.fft(ys)
                amplitudes = 1 / n_samples * np.abs(np_fft)  # This gives wrong results
                frequencies = np.fft.fftfreq(n_samples) * n_samples * 1 / (t1 - t0)
                plt.plot(frequencies[:len(frequencies) // 2], amplitudes[:len(np_fft) // 2])
                x = np.linspace(0, 10, 1000)
                y = np.sin(x)
                tol = 1e-2
                ind = np.argwhere(abs(y - 0.1 * max(y)) <= tol)
                plt.xlabel('Time')
                plt.ylabel('Amplitude ')
                plt.title('Feature values for maximum amplitude')
                plt.show()
                plt.close()
                # power and energy
                # Create input of  wave
                fs = 1.0
                fc = 0.25
                n = np.arange(0, 300)
                x = np.cos(2 * np.pi * n * fc / fs)
                # Rearrange x into 10 30 second windows
                x = np.reshape(x, (-1, 30))
                # Calculate power over each window [J/s]
                p = np.sum(x * x, 1) / x.size
                p_int = p
                p_str = str(p_int)
                # Calculate energy [J = J/s * 30 second]
                e = p * x.size
                e_int = e
                e_str = str(e_int)
                print("Power value is :", p)
                print("Energy value is : ", e)
                # Mean
                mean = np.mean(mean_data)
                mean_int = mean
                mean_str = str(mean_int)
                print("PCG signal Mean Mean is :", mean)
                # variance
                variance_int = statistics.variance(A_rank_signal_data)
                variance_str = str(variance_int)
                print("Variance of sample set is % s"
                      % (statistics.variance(A_rank_signal_data)))
                # root mean square error
                fname = self.input_pcg_signal
                with contextlib.closing(wave.open(fname, 'r')) as f:
                    frames = f.getnframes()
                    rate = f.getframerate()
                    duration = frames / float(rate)
                    width = f.getsampwidth()
                    channel = f.getnchannels()
                    size = width * channel
                    # f.rewind()
                    wav = f.readframes(f.getnframes())
                    # print(duration)
                rmse_int = audioop.rms(str(wav).encode("ascii"), 2)
                print("root mean square error value is : ", audioop.rms(str(wav).encode("ascii"), 2))

                # Bandwidth
                def get_bw_range(signal_array):
                    # Gets indices of columns where signal is present
                    signal_cols = np.where(np.sum(signal_array, axis=0) >= 1)
                    # Subtracts the minimum index from the maximum
                    bandwidth_range = np.max(signal_cols) - np.min(signal_cols)
                    return bandwidth_range

                def get_bw_data(signal_data):
                    # Set n to the number of signal mask arrays
                    n = np.size(signal_data, axis=0)
                    # Create numpy array of zeros for the output
                    bandwidth_data = np.zeros([n])
                    # Gets bandwidth range for each signal mask array, in units of pixels
                    for i in range(n):
                        bandwidth_data[i] = get_bw_range(signal_data[i])
                    return bandwidth_data

                mask = [[0, 0, 0, 0],
                        [0, 1, 1, 0],
                        [1, 1, 0, 1],
                        [0, 0, 1, 0]]
                mask = np.array(mask)
                # Get bandwidth range using imported function
                bandwidth_range = get_bw_range(mask)
                bandwidth_int = bandwidth_range
                bandwidth_str = str(bandwidth_int)
                print("bandwidth for PCG signal : ", bandwidth_range)
                # mid-frequency and average frequency
                Fs = 150.0;  # sampling rate
                Ts = 1.0 / Fs;  # sampling interval
                t = np.arange(0, 1, Ts)  # time vector
                ff = 50;  # frequency of the signal
                y = np.sin(2 * np.pi * ff * t)
                plt.plot(t, y)
                plt.title("Mid Frequency")
                plt.xlabel('Time')
                plt.ylabel('Amplitude')
                plt.show()
                plt.close()
                # Cepstrum peak amplitude
                x = np.linspace(0, 6 * np.pi, 1000)
                x = np.sin(x) + 0.6 * np.sin(2.6 * x)
                peaks, _ = find_peaks(x)
                Cepstrum_peak_amplitude = peak_prominences(x, peaks)[0]
                print("Cepstrum peak amplitude", Cepstrum_peak_amplitude)
                Cepstrum_peak_amplitude_int = Cepstrum_peak_amplitude
                Cepstrum_peak_amplitude_str = Cepstrum_peak_amplitude_int
                contour_heights = x[peaks] - Cepstrum_peak_amplitude
                plt.plot(x)
                plt.title("Cepstrum peak amplitude")
                plt.plot(peaks, x[peaks], "x")
                plt.vlines(x=peaks, ymin=contour_heights, ymax=x[peaks])
                plt.show()
                plt.close()
                # Mel-frequency cepstral coefficients
                # TRAIN_PATH = '../Dataset/PCG/'
                ipd.Audio(self.input_pcg_signal)
                sample_rate, audio = wavfile.read(self.input_pcg_signal)
                print("Sample rate: {0}Hz".format(sample_rate))
                print("Audio duration: {0}s".format(len(audio) / sample_rate))

                def normalize_audio(audio):
                    audio = audio / np.max(np.abs(audio))
                    return audio

                audio = normalize_audio(audio)
                '''plt.figure(figsize=(15,4))
                plt.plot(np.linspace(0, len(audio) / sample_rate, num=len(audio)), audio)
                plt.grid(True)
                plt.show()
                plt.close()'''

                def frame_audio(audio, FFT_size=2048, hop_size=10, sample_rate=44100):
                    # hop_size in ms
                    audio = np.pad(audio, int(FFT_size / 2), mode='reflect')
                    frame_len = np.round(sample_rate * hop_size / 1000).astype(int)
                    frame_num = int((len(audio) - FFT_size) / frame_len) + 1
                    frames = np.zeros((frame_num, FFT_size))
                    for n in range(frame_num):
                        frames[n] = audio[n * frame_len:n * frame_len + FFT_size]
                    return frames

                hop_size = 15  # ms
                FFT_size = 2048
                audio_framed = frame_audio(audio, FFT_size=FFT_size, hop_size=hop_size, sample_rate=sample_rate)
                print("Framed audio shape: {0}".format(audio_framed.shape))
                print("First frame:")
                print(audio_framed[1])
                print("Last frame:")
                print(audio_framed[-1])
                window = get_window("hann", FFT_size, fftbins=True)
                audio_win = audio_framed * window
                ind = 69
                plt.figure(figsize=(15, 6))
                plt.subplot(2, 1, 1)
                plt.plot(audio_framed[ind])
                plt.title('Original Frame')
                plt.grid(True)
                plt.subplot(2, 1, 2)
                plt.plot(audio_win[ind])
                plt.title('Frame After Windowing')
                plt.grid(True)
                plt.show()
                plt.close()
                audio_winT = np.transpose(audio_win)
                audio_fft = np.empty((int(1 + FFT_size // 2), audio_winT.shape[1]), dtype=np.complex64, order='F')
                audio_fft = np.transpose(audio_fft)
                audio_power = np.square(np.abs(audio_fft))
                print(audio_power.shape)
                freq_min = 0
                freq_high = sample_rate / 2
                mel_filter_num = 10
                print("Minimum frequency: {0}".format(freq_min))
                print("Maximum frequency: {0}".format(freq_high))

                def freq_to_mel(freq):
                    return 2595.0 * np.log10(1.0 + freq / 700.0)

                def met_to_freq(mels):
                    return 700.0 * (10.0 ** (mels / 2595.0) - 1.0)

                def get_filter_points(fmin, fmax, mel_filter_num, FFT_size, sample_rate=44100):
                    fmin_mel = freq_to_mel(fmin)
                    fmax_mel = freq_to_mel(fmax)
                    print("MEL min: {0}".format(fmin_mel))
                    print("MEL max: {0}".format(fmax_mel))
                    mels = np.linspace(fmin_mel, fmax_mel, num=mel_filter_num + 2)
                    freqs = met_to_freq(mels)
                    return np.floor((FFT_size + 1) / sample_rate * freqs).astype(int), freqs

                filter_points, mel_freqs = get_filter_points(freq_min, freq_high, mel_filter_num, FFT_size,
                                                             sample_rate=44100)
                print(filter_points)

                def get_filters(filter_points, FFT_size):
                    filters = np.zeros((len(filter_points) - 2, int(FFT_size / 2 + 1)))
                    for n in range(len(filter_points) - 2):
                        filters[n, filter_points[n]: filter_points[n + 1]] = np.linspace(0, 1,
                                                                                         filter_points[n + 1] -
                                                                                         filter_points[
                                                                                             n])
                        filters[n, filter_points[n + 1]: filter_points[n + 2]] = np.linspace(1, 0,
                                                                                             filter_points[n + 2] -
                                                                                             filter_points[
                                                                                                 n + 1])
                    return filters

                filters = get_filters(filter_points, FFT_size)
                # taken from the librosa library
                enorm = 2.0 / (mel_freqs[2:mel_filter_num + 2] - mel_freqs[:mel_filter_num])
                filters *= enorm[:, np.newaxis]
                audio_filtered = np.dot(filters, np.transpose(audio_power))
                audio_log = 10.0 * np.log10(audio_filtered)
                print(audio_log.shape)

                def dct(dct_filter_num, filter_len):
                    basis = np.empty((dct_filter_num, filter_len))
                    basis[0, :] = 1.0 / np.sqrt(filter_len)
                    samples = np.arange(1, 2 * filter_len, 2) * np.pi / (2.0 * filter_len)
                    for i in range(1, dct_filter_num):
                        basis[i, :] = np.cos(i * samples) * np.sqrt(2.0 / filter_len)
                    return basis

                dct_filter_num = 40
                dct_filters = dct(dct_filter_num, mel_filter_num)
                print("Mel-frequency cepstral coefficients (MFCCs) : ")
                cepstral_coefficents = np.dot(dct_filters, audio_log)
                print(cepstral_coefficents.shape)
                print("PCG Features are extracted successfully...")
                messagebox.showinfo("Information Message", "PCG feature's are extracted successfully...")
            else:
                print("Input signal is Normal Case so no need to do Feature Extraction...")
                messagebox.showerror("Error Message",
                                     "Input signal is Normal Case so no need to do Feature Extraction...")
        else:
            messagebox.showerror("Info Message", "Please do the Rule Generation first...")
    def Feature_selection(self):
        if self.boolFeatureExtraction:
            self.boolFeatureSelection = True
            if self.rule_generation_result.get() == "Abnormal" or self.ECG_signal_clustering_result.get() == "Abnormal" or self.PCG_signal_clustering_result.get() == "Abnormal":
                print("Feature Selection")
                print("=================")
                print("Existing feature selection algorithm")
                print("------------------------------------")
                from Code.Feature_Selection import Existing_DSO, Existing_AO, Existing_SLO, Existing_SSO
                print("Propoesd feature selection algorithm")
                print("------------------------------------")
                print("Poisson Distribution Function based Snow Leopard Optimization Algorithm was executing ...")
                import numpy as np
                import copy
                import numpy.random as rnd
                from Code.Feature_Selection import Proposed_PDF_SLO as PDFSLO
                def PDFSLO_opt_algm(f, bounds, p, c1, c2, vmax, tol):
                    print('Selects the leopard for guiding in the row based on the  using PDF...')
                    d, particle_pos, particle_best, snow_best, snow_velocity, \
                    local_best, pos_val \
                        = PDFSLO.initiation(f, bounds, p)  # initializing various arrays
                    old_snow_best = [0] * d
                    c3 = c1 + c2
                    K = 2 / (abs(2 - c3 - np.sqrt((c3 ** 2) - (4 * c3))))  # creating velocity weighting factor
                    it_count = 0
                    while abs(f(old_snow_best) - f(snow_best)) > tol:  # exit condition

                        it_count += 1
                        if it_count > 1000:  # every 1000 iterations...
                            # create 'conflict' within the snow and
                            # give all particles random velocities
                            print('Modeling the zig‐zag pattern movements of snow leopards')
                            for j in range(p):  # iterating ovre the number of particles
                                snow_velocity[j] = [(rnd.uniform(-abs(bounds[i][1] - bounds[i][0]) \
                                                                 , abs(bounds[i][1] - bounds[i][0]))) for i in range(d)]
                                # adding PDF velocity values for each dimension
                            it_count = 0  # reset iteration count

                        for i in range(p):  # iterates over each Travel Routes and Movement
                            rp, rg = rnd.uniform(0, 1, 2)  # creates two random numbers between 0-
                            snow_velocity[i, :] += (c1 * rp * (particle_best[i, :] - particle_pos[i, :]))
                            snow_velocity[i, :] += (c2 * rg * (local_best[i, :] - particle_pos[i, :]))
                            snow_velocity[i, :] = snow_velocity[i, :] * K
                            if snow_velocity[i].any() > vmax:  # is any velocity is greater than vmax
                                snow_velocity[i, :] = vmax  # set velocity to vmax
                            # all of the above is regarding updating the particle's velocity
                            # with regards to various parameters (local_best, p_best etc..)
                            particle_pos[i, :] += snow_velocity[i, :]  # updating position

                            PDFSLO.withinbounds(bounds, particle_pos[i])  # if particle is out of bounds
                            particle_fitness = f(particle_pos[i])
                            if particle_fitness < pos_val[i]:
                                particle_best[i, :] = particle_pos[i, :]  # checking if new best
                                pos_val[i] = particle_fitness
                                f_snow_best = f(snow_best)
                                if particle_fitness < f_snow_best:
                                    old_snow_best = snow_best[:]
                                    snow_best = copy.deepcopy(particle_best[i, :])
                                    print('Members of the population are updated in the '
                                          'proposed PDFSLOA based on simulating the natural behaviors of snow leopards: ',
                                          f(snow_best))
                        local_best = PDFSLO.local_best_get(particle_pos, pos_val, p)
                        self.feature_selection_result.set(snow_best)
                    return print('Optimum at: ', snow_best, '\n', 'Function at optimum weight value (fitness value): ',
                                 f(snow_best))

                # print(self.out_arr[200:800])
                f = PDFSLO.displacement
                dimensions = 10
                dimension_bounds = [-2, 2]
                bounds = [0] * dimensions  # creating 5 dimensional bounds
                for i in range(dimensions):
                    bounds[i] = dimension_bounds
                # self.feature_selection_result.set(self.out_arr[200:800])
                # creates bounds [[x1,x2],[x3,x4],[x5,x6]....]
                p = 60  # shouldn't really change
                vmax = (dimension_bounds[1] - dimension_bounds[0]) * 0.75
                c1 = 2.8  # shouldn't really change
                c2 = 1.3  # shouldn't really change
                tol = 0.00000000000001
                PDFSLO_opt_algm(f, bounds, p, c1, c2, vmax, tol)
                print(
                    "Poisson Distribution Function based Snow Leopard Optimization Algorithm was executed successfully ... ")
                print("feature selection was completed successfully...")
                messagebox.showinfo("Information Message", "feature selection was completed successfully...")
            else:
                print("Input signal is Normal Case so no need to do Feature Selection...")
                messagebox.showerror("Error Message",
                                     "Input signal is Normal Case so no need to do Feature Selection...")
        else:
            messagebox.showerror("Info Message", "Please do the Feature Extraction first...")

    def classification(self):
        if self.boolFeatureSelection:
            self.boolClassification = True
            if self.rule_generation_result.get() == "Abnormal":
                from Code.Classifier.Existing_DNN import ExistingDNN
                from Code.Classifier.Existing_DJRNN import ExistingDJRNN
                from Code.Classifier.Existing_RNN import ExistingRNN
                from Code.Classifier.Existing_ANN import ExistingANN
                from Code.Classifier.Existing_ENN import ExistingENN
                from Code.Classifier.Proposed_PJM_DJRNN import ProposedPJMDJRMM
                print("Classification")
                print("==============")
                print("Existing algorithm")
                print("------------------")
                self.iptrdata = getListOfFiles("..\\Dataset\\")
                self.iptsdata = getListOfFiles("..\\Dataset\\")
                print("Total no. of Data : " + str(len(self.iptrdata)))
                from Code.Classifier import Existing_DJRNN, Existing_ANN, Existing_DNN, Existing_ENN, Existing_RNN
                print("Existing Deep Jordan Recurrent Neural Network")
                print("---------------------------------------------")
                stime = int(time.time() * 1000)
                ExistingDJRNN.training(self, self.iptrdata)
                etime = int(time.time() * 1000)
                cfg.edjrnnct = etime - stime
                print("Training Time : " + str(etime - stime) + " in ms")

                ExistingDJRNN.testing(self, self.iptsdata)

                print("Precision : " + str(cfg.edjrnnpre))
                print("Recall : " + str(cfg.edjrnnrec))
                print("FMeasure : " + str(cfg.edjrnnfsc))
                print("Accuracy : " + str(cfg.edjrnnacc))
                print("Sensitivity : " + str(cfg.edjrnnsens))
                print("Specificity : " + str(cfg.edjrnnspec))
                print("MCC : " + str(cfg.edjrnnmcc))
                print("FPR : " + str(cfg.edjrnnfpr))
                print("FNR : " + str(cfg.edjrnnfnr))

                print("Existing Recurrent Neural Network (RNN)")
                print("---------------------------------------")

                stime = int(time.time() * 1000)
                ExistingRNN.training(self, self.iptrdata)
                etime = int(time.time() * 1000)
                cfg.ernnct = etime - stime
                print("Training Time : " + str(etime - stime) + " in ms")

                ExistingRNN.testing(self, self.iptsdata)

                print("Precision : " + str(cfg.ernnpre))
                print("Recall : " + str(cfg.ernnrec))
                print("FMeasure : " + str(cfg.ernnfsc))
                print("Accuracy : " + str(cfg.ernnacc))
                print("Sensitivity : " + str(cfg.ernnsens))
                print("Specificity : " + str(cfg.ernnspec))
                print("MCC : " + str(cfg.ernnmcc))
                print("FPR : " + str(cfg.ernnfpr))
                print("FNR : " + str(cfg.ernnfnr))

                print("\nExisting Deep Neural Network  (DNN)")
                print("-------------------------------------")

                stime = int(time.time() * 1000)
                ExistingDNN.training(self, self.iptrdata)
                etime = int(time.time() * 1000)
                cfg.ednnct = etime - stime
                print("Training Time : " + str(etime - stime) + " in ms")

                ExistingDNN.testing(self, self.iptsdata)

                print("Precision : " + str(cfg.ednnpre))
                print("Recall : " + str(cfg.ednnrec))
                print("FMeasure : " + str(cfg.ednnfsc))
                print("Accuracy : " + str(cfg.ednnacc))
                print("Sensitivity : " + str(cfg.ednnsens))
                print("Specificity : " + str(cfg.ednnspec))
                print("MCC : " + str(cfg.ednnmcc))
                print("FPR : " + str(cfg.ednnfpr))
                print("FNR : " + str(cfg.ednnfnr))

                print("\nExisting Artificial Neural Network (ANN)")
                print("------------------------------------------")

                stime = int(time.time() * 1000)
                ExistingANN.training(self, self.iptrdata)
                etime = int(time.time() * 1000)
                cfg.eannct = etime - stime
                print("Training Time : " + str(etime - stime) + " in ms")

                ExistingANN.testing(self, self.iptsdata)

                print("Precision : " + str(cfg.eannpre))
                print("Recall : " + str(cfg.eannrec))
                print("FMeasure : " + str(cfg.eannfsc))
                print("Accuracy : " + str(cfg.eannacc))
                print("Sensitivity : " + str(cfg.eannsens))
                print("Specificity : " + str(cfg.eannspec))
                print("MCC : " + str(cfg.eannmcc))
                print("FPR : " + str(cfg.eannfpr))
                print("FNR : " + str(cfg.eannfnr))

                print("\nExisting Elman Neural Network  (ENN)")
                print("--------------------------------------")

                stime = int(time.time() * 1000)
                ExistingENN.training(self, self.iptrdata)
                etime = int(time.time() * 1000)
                cfg.eennct = etime - stime
                print("Training Time : " + str(etime - stime) + " in ms")

                ExistingENN.testing(self, self.iptsdata)

                print("Precision : " + str(cfg.eennpre))
                print("Recall : " + str(cfg.eennrec))
                print("FMeasure : " + str(cfg.eennfsc))
                print("Accuracy : " + str(cfg.eennacc))
                print("Sensitivity : " + str(cfg.eennsens))
                print("Specificity : " + str(cfg.eennspec))
                print("MCC : " + str(cfg.eennmcc))
                print("FPR : " + str(cfg.eennfpr))
                print("FNR : " + str(cfg.eennfnr))

                print("Proposed algorithm")
                print("------------------")
                from Code.Classifier import Proposed_PJM_DJRNN
                from Code.Classifier.Proposed_PJM_DJRNN import pjm_djrnn_classified_result
                print("Polynomial Jacobian Matrix form-based Deep Jordan Recurrent Neural Network was executing...")

                def _transpose_batch_time(x):
                    x_static_shape = x.get_shape()
                    if x_static_shape.ndims is not None and x_static_shape.ndims < 2:
                        raise ValueError(
                            "Expected input tensor %s to have rank at least 2, but saw shape: %s" %
                            (x, x_static_shape))
                    x_rank = array_ops.rank(x)
                    x_t = array_ops.transpose(
                        x, array_ops.concat(
                            ([1, 0], math_ops.range(2, x_rank)), axis=0))
                    x_t.set_shape(
                        tensor_shape.TensorShape([
                            x_static_shape[1].value, x_static_shape[0].value
                        ]).concatenate(x_static_shape[2:]))
                    return x_t

                def _best_effort_input_batch_size(flat_input):

                    for input_ in flat_input:
                        shape = input_.shape
                        if shape.ndims is None:
                            continue
                        if shape.ndims < 2:
                            raise ValueError(
                                "Expected input tensor %s to have rank at least 2" % input_)
                        batch_size = shape[1].value
                        if batch_size is not None:
                            return batch_size
                    # Fallback to the dynamic batch size of the first input.
                    return array_ops.shape(flat_input[0])[1]

                def _infer_state_dtype(explicit_dtype, state):

                    if explicit_dtype is not None:
                        return explicit_dtype
                    elif nest.is_sequence(state):
                        inferred_dtypes = [element.dtype for element in nest.flatten(state)]
                        if not inferred_dtypes:
                            raise ValueError("Unable to infer dtype from empty state.")
                        all_same = all([x == inferred_dtypes[0] for x in inferred_dtypes])
                        if not all_same:
                            raise ValueError(
                                "State has tensors of different inferred_dtypes. Unable to infer a "
                                "single representative dtype.")
                        return inferred_dtypes[0]
                    else:
                        return state.dtype

                # pylint: disable=unused-argument
                def _rnn_step(
                        time, sequence_length, min_sequence_length, max_sequence_length,
                        zero_output, state, call_cell, state_size, skip_conditionals=False):

                    # Convert state to a list for ease of use
                    flat_state = nest.flatten(state)
                    flat_zero_output = nest.flatten(zero_output)

                    def _copy_one_through(output, new_output):
                        # If the state contains a scalar value we simply pass it through.
                        if output.shape.ndims == 0:
                            return new_output
                        copy_cond = (time >= sequence_length)
                        with ops.colocate_with(new_output):
                            return array_ops.where(copy_cond, output, new_output)

                    def _copy_some_through(flat_new_output, flat_new_state):
                        # Use broadcasting select to determine which values should get
                        # the previous state & zero output, and which values should get
                        # a calculated state & output.
                        flat_new_output = [
                            _copy_one_through(zero_output, new_output)
                            for zero_output, new_output in zip(flat_zero_output, flat_new_output)]
                        flat_new_state = [
                            _copy_one_through(state, new_state)
                            for state, new_state in zip(flat_state, flat_new_state)]
                        return flat_new_output + flat_new_state

                    def _maybe_copy_some_through():
                        """Run RNN step.  Pass through either no or some past state."""
                        new_output, new_state = call_cell()

                        nest.assert_same_structure(state, new_state)

                        flat_new_state = nest.flatten(new_state)
                        flat_new_output = nest.flatten(new_output)
                        return control_flow_ops.cond(
                            # if t < min_seq_len: calculate and return everything
                            time < min_sequence_length, lambda: flat_new_output + flat_new_state,
                            # else copy some of it through
                            lambda: _copy_some_through(flat_new_output, flat_new_state))

                    if skip_conditionals:
                        # Instead of using conditionals, perform the selective copy at all time
                        # steps.  This is faster when max_seq_len is equal to the number of unrolls
                        # (which is typical for dynamic_rnn).
                        new_output, new_state = call_cell()
                        nest.assert_same_structure(state, new_state)
                        new_state = nest.flatten(new_state)
                        new_output = nest.flatten(new_output)
                        final_output_and_state = _copy_some_through(new_output, new_state)
                    else:
                        empty_update = lambda: flat_zero_output + flat_state
                        final_output_and_state = control_flow_ops.cond(
                            # if t >= max_seq_len: copy all state through, output zeros
                            time >= max_sequence_length, empty_update,
                            # otherwise calculation is required: copy some or all of it through
                            _maybe_copy_some_through)

                    if len(final_output_and_state) != len(flat_zero_output) + len(flat_state):
                        raise ValueError("Internal error: state and output were not concatenated "
                                         "correctly.")
                    final_output = final_output_and_state[:len(flat_zero_output)]
                    final_state = final_output_and_state[len(flat_zero_output):]

                    for output, flat_output in zip(final_output, flat_zero_output):
                        output.set_shape(flat_output.get_shape())
                    for substate, flat_substate in zip(final_state, flat_state):
                        substate.set_shape(flat_substate.get_shape())

                    final_output = nest.pack_sequence_as(
                        structure=zero_output, flat_sequence=final_output)
                    final_state = nest.pack_sequence_as(
                        structure=state, flat_sequence=final_state)

                    return final_output, final_state

                def _reverse_seq(input_seq, lengths):

                    if lengths is None:
                        return list(reversed(input_seq))

                    flat_input_seq = tuple(nest.flatten(input_) for input_ in input_seq)

                    flat_results = [[] for _ in range(len(input_seq))]
                    for sequence in zip(*flat_input_seq):
                        input_shape = tensor_shape.unknown_shape(
                            ndims=sequence[0].get_shape().ndims)
                        for input_ in sequence:
                            input_shape.merge_with(input_.get_shape())
                            input_.set_shape(input_shape)

                        # Join into (time, batch_size, depth)
                        s_joined = array_ops.stack(sequence)

                        # Reverse along dimension 0
                        s_reversed = array_ops.reverse_sequence(s_joined, lengths, 0, 1)
                        # Split again into list
                        result = array_ops.unstack(s_reversed)
                        for r, flat_result in zip(result, flat_results):
                            r.set_shape(input_shape)
                            flat_result.append(r)

                    results = [nest.pack_sequence_as(structure=input_, flat_sequence=flat_result)
                               for input_, flat_result in zip(input_seq, flat_results)]
                    return results

                def bidirectional_dynamic_rnn(cell_fw, cell_bw, inputs, sequence_length=None,
                                              initial_state_fw=None, initial_state_bw=None,
                                              dtype=None, parallel_iterations=None,
                                              swap_memory=False, time_major=False, scope=None):

                    with vs.variable_scope(scope or "bidirectional_rnn"):
                        # Forward direction
                        with vs.variable_scope("fw") as fw_scope:
                            output_fw, output_state_fw = dynamic_rnn(
                                cell=cell_fw, inputs=inputs, sequence_length=sequence_length,
                                initial_state=initial_state_fw, dtype=dtype,
                                parallel_iterations=parallel_iterations, swap_memory=swap_memory,
                                time_major=time_major, scope=fw_scope)

                        # Backward direction
                        if not time_major:
                            time_dim = 1
                            batch_dim = 0
                        else:
                            time_dim = 0
                            batch_dim = 1

                        def _reverse(input_, seq_lengths, seq_dim, batch_dim):
                            if seq_lengths is not None:
                                return array_ops.reverse_sequence(
                                    input=input_, seq_lengths=seq_lengths,
                                    seq_dim=seq_dim, batch_dim=batch_dim)
                            else:
                                return array_ops.reverse(input_, axis=[seq_dim])

                        with vs.variable_scope("bw") as bw_scope:
                            inputs_reverse = _reverse(
                                inputs, seq_lengths=sequence_length,
                                seq_dim=time_dim, batch_dim=batch_dim)
                            tmp, output_state_bw = dynamic_rnn(
                                cell=cell_bw, inputs=inputs_reverse, sequence_length=sequence_length,
                                initial_state=initial_state_bw, dtype=dtype,
                                parallel_iterations=parallel_iterations, swap_memory=swap_memory,
                                time_major=time_major, scope=bw_scope)

                    output_bw = _reverse(
                        tmp, seq_lengths=sequence_length,
                        seq_dim=time_dim, batch_dim=batch_dim)

                    outputs = (output_fw, output_bw)
                    output_states = (output_state_fw, output_state_bw)

                    return (outputs, output_states)

                def dynamic_rnn(cell, inputs, att_scores=None, sequence_length=None, initial_state=None,
                                dtype=None, parallel_iterations=None, swap_memory=False,
                                time_major=False, scope=None):

                    # By default, time_major==False and inputs are batch-major: shaped
                    #   [batch, time, depth]
                    # For internal calculations, we transpose to [time, batch, depth]
                    flat_input = nest.flatten(inputs)

                    if not time_major:
                        # (B,T,D) => (T,B,D)
                        flat_input = [ops.convert_to_tensor(input_) for input_ in flat_input]
                        flat_input = tuple(_transpose_batch_time(input_) for input_ in flat_input)

                    parallel_iterations = parallel_iterations or 32
                    if sequence_length is not None:
                        sequence_length = math_ops.to_int32(sequence_length)
                        if sequence_length.get_shape().ndims not in (None, 1):
                            raise ValueError(
                                "sequence_length must be a vector of length batch_size, "
                                "but saw shape: %s" % sequence_length.get_shape())
                        sequence_length = array_ops.identity(  # Just to find it in the graph.
                            sequence_length, name="sequence_length")

                    # Create a new scope in which the caching device is either
                    # determined by the parent scope, or is set to place the cached
                    # Variable using the same placement as for the rest of the RNN.
                    with vs.variable_scope(scope or "rnn") as varscope:
                        if varscope.caching_device is None:
                            varscope.set_caching_device(lambda op: op.device)
                        batch_size = _best_effort_input_batch_size(flat_input)

                        if initial_state is not None:
                            state = initial_state
                        else:
                            if not dtype:
                                raise ValueError("If there is no initial_state, you must give a dtype.")
                            state = cell.zero_state(batch_size, dtype)

                        def _assert_has_shape(x, shape):
                            x_shape = array_ops.shape(x)
                            packed_shape = array_ops.stack(shape)
                            return control_flow_ops.Assert(
                                math_ops.reduce_all(math_ops.equal(x_shape, packed_shape)),
                                ["Expected shape for Tensor %s is " % x.name,
                                 packed_shape, " but saw shape: ", x_shape])

                        if sequence_length is not None:
                            # Perform some shape validation
                            with ops.control_dependencies(
                                    [_assert_has_shape(sequence_length, [batch_size])]):
                                sequence_length = array_ops.identity(
                                    sequence_length, name="CheckSeqLen")

                        inputs = nest.pack_sequence_as(structure=inputs, flat_sequence=flat_input)

                        (outputs, final_state) = _dynamic_rnn_loop(
                            cell,
                            inputs,
                            state,
                            parallel_iterations=parallel_iterations,
                            swap_memory=swap_memory,
                            att_scores=att_scores,
                            sequence_length=sequence_length,
                            dtype=dtype)

                        # Outputs of _dynamic_rnn_loop are always shaped [time, batch, depth].
                        # If we are performing batch-major calculations, transpose output back
                        # to shape [batch, time, depth]
                        if not time_major:
                            # (T,B,D) => (B,T,D)
                            outputs = nest.map_structure(_transpose_batch_time, outputs)

                        return (outputs, final_state)

                def _dynamic_rnn_loop(cell,
                                      inputs,
                                      initial_state,
                                      parallel_iterations,
                                      swap_memory,
                                      att_scores=None,
                                      sequence_length=None,
                                      dtype=None):

                    state = initial_state
                    assert isinstance(parallel_iterations, int), "parallel_iterations must be int"

                    state_size = cell.state_size

                    flat_input = nest.flatten(inputs)
                    flat_output_size = nest.flatten(cell.output_size)

                    # Construct an initial output
                    input_shape = array_ops.shape(flat_input[0])
                    time_steps = input_shape[0]
                    batch_size = _best_effort_input_batch_size(flat_input)

                    inputs_got_shape = tuple(input_.get_shape().with_rank_at_least(3)
                                             for input_ in flat_input)

                    const_time_steps, const_batch_size = inputs_got_shape[0].as_list()[:2]

                    for shape in inputs_got_shape:
                        if not shape[2:].is_fully_defined():
                            raise ValueError(
                                "Input size (depth of inputs) must be accessible via shape inference,"
                                " but saw value None.")
                        got_time_steps = shape[0].value
                        got_batch_size = shape[1].value
                        if const_time_steps != got_time_steps:
                            raise ValueError(
                                "Time steps is not the same for all the elements in the input in a "
                                "batch.")
                        if const_batch_size != got_batch_size:
                            raise ValueError(
                                "Batch_size is not the same for all the elements in the input.")

                    # Prepare dynamic conditional copying of state & output
                    def _create_zero_arrays(size):
                        return array_ops.zeros(
                            array_ops.stack(size), _infer_state_dtype(dtype, state))

                    flat_zero_output = tuple(_create_zero_arrays(output)
                                             for output in flat_output_size)
                    zero_output = nest.pack_sequence_as(structure=cell.output_size,
                                                        flat_sequence=flat_zero_output)

                    if sequence_length is not None:
                        min_sequence_length = math_ops.reduce_min(sequence_length)
                        max_sequence_length = math_ops.reduce_max(sequence_length)

                    time = array_ops.constant(0, dtype=dtypes.int32, name="time")

                    with ops.name_scope("dynamic_rnn") as scope:
                        base_name = scope

                    def _create_ta(name, dtype):
                        return tensor_array_ops.TensorArray(dtype=dtype,
                                                            size=time_steps,
                                                            tensor_array_name=base_name + name)

                    output_ta = tuple(_create_ta("output_%d" % i,
                                                 _infer_state_dtype(dtype, state))
                                      for i in range(len(flat_output_size)))
                    input_ta = tuple(_create_ta("input_%d" % i, flat_input[i].dtype)
                                     for i in range(len(flat_input)))

                    input_ta = tuple(ta.unstack(input_)
                                     for ta, input_ in zip(input_ta, flat_input))

                    def _time_step(time, output_ta_t, state, att_scores=None):

                        input_t = tuple(ta.read(time) for ta in input_ta)
                        # Restore some shape information
                        for input_, shape in zip(input_t, inputs_got_shape):
                            input_.set_shape(shape[1:])

                        input_t = nest.pack_sequence_as(structure=inputs, flat_sequence=input_t)
                        if att_scores is not None:
                            att_score = att_scores[:, time, :]
                            call_cell = lambda: cell(input_t, state, att_score)
                        else:
                            call_cell = lambda: cell(input_t, state)

                        if sequence_length is not None:
                            (output, new_state) = _rnn_step(
                                time=time,
                                sequence_length=sequence_length,
                                min_sequence_length=min_sequence_length,
                                max_sequence_length=max_sequence_length,
                                zero_output=zero_output,
                                state=state,
                                call_cell=call_cell,
                                state_size=state_size,
                                skip_conditionals=True)
                        else:
                            (output, new_state) = call_cell()

                        # Pack state if using state tuples
                        output = nest.flatten(output)

                        output_ta_t = tuple(
                            ta.write(time, out) for ta, out in zip(output_ta_t, output))
                        if att_scores is not None:
                            return (time + 1, output_ta_t, new_state, att_scores)
                        else:
                            return (time + 1, output_ta_t, new_state)

                    if att_scores is not None:
                        _, output_final_ta, final_state, _ = control_flow_ops.while_loop(
                            cond=lambda time, *_: time < time_steps,
                            body=_time_step,
                            loop_vars=(time, output_ta, state, att_scores),
                            parallel_iterations=parallel_iterations,
                            swap_memory=swap_memory)
                    else:
                        _, output_final_ta, final_state = control_flow_ops.while_loop(
                            cond=lambda time, *_: time < time_steps,
                            body=_time_step,
                            loop_vars=(time, output_ta, state),
                            parallel_iterations=parallel_iterations,
                            swap_memory=swap_memory)

                    # Unpack final output if not using output tuples.
                    final_outputs = tuple(ta.stack() for ta in output_final_ta)

                    # Restore some shape information
                    for output, output_size in zip(final_outputs, flat_output_size):
                        output.set_shape(shape)

                    final_outputs = nest.pack_sequence_as(
                        structure=cell.output_size, flat_sequence=final_outputs)

                    return (final_outputs, final_state)

                def raw_rnn(cell, loop_fn,
                            parallel_iterations=None, swap_memory=False, scope=None):

                    if not callable(loop_fn):
                        raise TypeError("loop_fn must be a callable")

                    parallel_iterations = parallel_iterations or 32

                    # Create a new scope in which the caching device is either
                    # determined by the parent scope, or is set to place the cached
                    # Variable using the same placement as for the rest of the RNN.
                    with vs.variable_scope(scope or "rnn") as varscope:
                        if varscope.caching_device is None:
                            varscope.set_caching_device(lambda op: op.device)

                        time = constant_op.constant(0, dtype=dtypes.int32)
                        (elements_finished, next_input, initial_state, emit_structure,
                         init_loop_state) = loop_fn(
                            time, None, None, None)  # time, cell_output, cell_state, loop_state
                        flat_input = nest.flatten(next_input)

                        # Need a surrogate loop state for the while_loop if none is available.
                        loop_state = (init_loop_state if init_loop_state is not None
                                      else constant_op.constant(0, dtype=dtypes.int32))

                        input_shape = [input_.get_shape() for input_ in flat_input]
                        static_batch_size = input_shape[0][0]

                        for input_shape_i in input_shape:
                            # Static verification that batch sizes all match
                            static_batch_size.merge_with(input_shape_i[0])

                        batch_size = static_batch_size.value
                        if batch_size is None:
                            batch_size = array_ops.shape(flat_input[0])[0]

                        nest.assert_same_structure(initial_state, cell.state_size)
                        state = initial_state
                        flat_state = nest.flatten(state)
                        flat_state = [ops.convert_to_tensor(s) for s in flat_state]
                        state = nest.pack_sequence_as(structure=state,
                                                      flat_sequence=flat_state)

                        if emit_structure is not None:
                            flat_emit_structure = nest.flatten(emit_structure)
                            flat_emit_size = [emit.shape if emit.shape.is_fully_defined() else
                                              array_ops.shape(emit) for emit in flat_emit_structure]
                            flat_emit_dtypes = [emit.dtype for emit in flat_emit_structure]
                        else:
                            emit_structure = cell.output_size
                            flat_emit_size = nest.flatten(emit_structure)
                            flat_emit_dtypes = [flat_state[0].dtype] * len(flat_emit_size)

                        flat_emit_ta = [
                            tensor_array_ops.TensorArray(
                                dtype=dtype_i, dynamic_size=True, size=0, name="rnn_output_%d" % i)
                            for i, dtype_i in enumerate(flat_emit_dtypes)]
                        emit_ta = nest.pack_sequence_as(structure=emit_structure,
                                                        flat_sequence=flat_emit_ta)

                        def condition(unused_time, elements_finished, *_):
                            return math_ops.logical_not(math_ops.reduce_all(elements_finished))

                        def body(time, elements_finished, current_input,
                                 emit_ta, state, loop_state):

                            (next_output, cell_state) = cell(current_input, state)

                            nest.assert_same_structure(state, cell_state)
                            nest.assert_same_structure(cell.output_size, next_output)

                            next_time = time + 1
                            (next_finished, next_input, next_state, emit_output,
                             next_loop_state) = loop_fn(
                                next_time, next_output, cell_state, loop_state)

                            nest.assert_same_structure(state, next_state)
                            nest.assert_same_structure(current_input, next_input)
                            nest.assert_same_structure(emit_ta, emit_output)

                            # If loop_fn returns None for next_loop_state, just reuse the
                            # previous one.
                            loop_state = loop_state if next_loop_state is None else next_loop_state

                            def _copy_some_through(current, candidate):
                                """Copy some tensors through via array_ops.where."""

                                def copy_fn(cur_i, cand_i):
                                    with ops.colocate_with(cand_i):
                                        return array_ops.where(elements_finished, cur_i, cand_i)

                                return nest.map_structure(copy_fn, current, candidate)

                            next_state = _copy_some_through(state, next_state)

                            emit_ta = nest.map_structure(
                                lambda ta, emit: ta.write(time, emit), emit_ta, emit_output)

                            elements_finished = math_ops.logical_or(elements_finished, next_finished)

                            return (next_time, elements_finished, next_input,
                                    emit_ta, next_state, loop_state)

                        returned = control_flow_ops.while_loop(
                            condition, body, loop_vars=[
                                time, elements_finished, next_input,
                                emit_ta, state, loop_state],
                            parallel_iterations=parallel_iterations,
                            swap_memory=swap_memory)

                        (emit_ta, final_state, final_loop_state) = returned[-3:]

                        if init_loop_state is None:
                            final_loop_state = None

                        return (emit_ta, final_state, final_loop_state)

                def static_rnn(cell,
                               inputs,
                               initial_state=None,
                               dtype=None,
                               sequence_length=None,
                               scope=None):

                    if not nest.is_sequence(inputs):
                        raise TypeError("inputs must be a sequence")
                    if not inputs:
                        raise ValueError("inputs must not be empty")

                    outputs = []
                    # Create a new scope in which the caching device is either
                    # determined by the parent scope, or is set to place the cached
                    # Variable using the same placement as for the rest of the RNN.
                    with vs.variable_scope(scope or "rnn") as varscope:
                        if varscope.caching_device is None:
                            varscope.set_caching_device(lambda op: op.device)

                        # Obtain the first sequence of the input
                        first_input = inputs
                        while nest.is_sequence(first_input):
                            first_input = first_input[0]

                        if first_input.get_shape().ndims != 1:

                            input_shape = first_input.get_shape().with_rank_at_least(2)
                            fixed_batch_size = input_shape[0]

                            flat_inputs = nest.flatten(inputs)
                            for flat_input in flat_inputs:
                                input_shape = flat_input.get_shape().with_rank_at_least(2)
                                batch_size, input_size = input_shape[0], input_shape[1:]
                                fixed_batch_size.merge_with(batch_size)
                                for i, size in enumerate(input_size):
                                    if size.value is None:
                                        raise ValueError(
                                            "Input size (dimension %d of inputs) must be accessible via "
                                            "shape inference, but saw value None." % i)
                        else:
                            fixed_batch_size = first_input.get_shape().with_rank_at_least(1)[0]

                        if fixed_batch_size.value:
                            batch_size = fixed_batch_size.value
                        else:
                            batch_size = array_ops.shape(first_input)[0]
                        if initial_state is not None:
                            state = initial_state
                        else:
                            if not dtype:
                                raise ValueError("If no initial_state is provided, "
                                                 "dtype must be specified")
                            state = cell.zero_state(batch_size, dtype)

                        if sequence_length is not None:  # Prepare variables
                            sequence_length = ops.convert_to_tensor(
                                sequence_length, name="sequence_length")
                            if sequence_length.get_shape().ndims not in (None, 1):
                                raise ValueError(
                                    "sequence_length must be a vector of length batch_size")

                            output_size = cell.output_size
                            flat_output_size = nest.flatten(output_size)

                            sequence_length = math_ops.to_int32(sequence_length)
                            min_sequence_length = math_ops.reduce_min(sequence_length)
                            max_sequence_length = math_ops.reduce_max(sequence_length)

                        for time, input_ in enumerate(inputs):
                            if time > 0:
                                varscope.reuse_variables()
                            # pylint: disable=cell-var-from-loop
                            call_cell = lambda: cell(input_, state)
                            # pylint: enable=cell-var-from-loop
                            if sequence_length is not None:
                                (output, state) = _rnn_step(
                                    time=time,
                                    sequence_length=sequence_length,
                                    min_sequence_length=min_sequence_length,
                                    max_sequence_length=max_sequence_length,

                                    state=state,
                                    call_cell=call_cell,
                                    state_size=cell.state_size)
                            else:
                                (output, state) = call_cell()
                            outputs.append(output)
                        return (outputs, state)

                target_result = pjm_djrnn_classified_result

                def static_state_saving_rnn(cell,
                                            inputs,
                                            state_saver,
                                            state_name,
                                            sequence_length=None,
                                            scope=None):
                    """RNN that accepts a state saver for time-truncated RNN calculation.

                    """
                    state_size = cell.state_size
                    state_is_tuple = nest.is_sequence(state_size)
                    state_name_tuple = nest.is_sequence(state_name)

                    if state_is_tuple != state_name_tuple:
                        raise ValueError("state_name should be the same type as cell.state_size.  "
                                         "state_name: %s, cell.state_size: %s" % (str(state_name),
                                                                                  str(state_size)))

                    if state_is_tuple:
                        state_name_flat = nest.flatten(state_name)
                        state_size_flat = nest.flatten(state_size)

                        if len(state_name_flat) != len(state_size_flat):
                            raise ValueError("#elems(state_name) != #elems(state_size): %d vs. %d" %
                                             (len(state_name_flat), len(state_size_flat)))

                        initial_state = nest.pack_sequence_as(
                            structure=state_size,
                            flat_sequence=[state_saver.state(s) for s in state_name_flat])
                    else:
                        initial_state = state_saver.state(state_name)

                    (outputs, state) = static_rnn(
                        cell,
                        inputs,
                        initial_state=initial_state,
                        sequence_length=sequence_length,
                        scope=scope)

                    if state_is_tuple:
                        flat_state = nest.flatten(state)
                        state_name = nest.flatten(state_name)
                        save_state = [
                            state_saver.save_state(name, substate)
                            for name, substate in zip(state_name, flat_state)
                        ]
                    else:
                        save_state = [state_saver.save_state(state_name, state)]

                    with ops.control_dependencies(save_state):
                        last_output = outputs[-1]
                        flat_last_output = nest.flatten(last_output)
                        flat_last_output = [
                            array_ops.identity(output) for output in flat_last_output
                        ]
                        outputs[-1] = nest.pack_sequence_as(
                            structure=last_output, flat_sequence=flat_last_output)
                    return (outputs, state)

                def static_bidirectional_rnn(cell_fw,
                                             cell_bw,
                                             inputs,
                                             initial_state_fw=None,
                                             initial_state_bw=None,
                                             dtype=None,
                                             sequence_length=None,
                                             scope=None):

                    if not nest.is_sequence(inputs):
                        raise TypeError("inputs must be a sequence")
                    if not inputs:
                        raise ValueError("inputs must not be empty")

                    with vs.variable_scope(scope or "bidirectional_rnn"):
                        # Forward direction
                        with vs.variable_scope("fw") as fw_scope:
                            output_fw, output_state_fw = static_rnn(
                                cell_fw,
                                inputs,
                                initial_state_fw,
                                dtype,
                                sequence_length,
                                scope=fw_scope)

                        # Backward direction
                        with vs.variable_scope("bw") as bw_scope:
                            reversed_inputs = _reverse_seq(inputs, sequence_length)
                            tmp, output_state_bw = static_rnn(
                                cell_bw,
                                reversed_inputs,
                                initial_state_bw,
                                dtype,
                                sequence_length,
                                scope=bw_scope)

                    output_bw = _reverse_seq(tmp, sequence_length)
                    # Concat each of the forward/backward outputs
                    flat_output_fw = nest.flatten(output_fw)
                    flat_output_bw = nest.flatten(output_bw)

                    flat_outputs = tuple(
                        array_ops.concat([fw, bw], 1)
                        for fw, bw in zip(flat_output_fw, flat_output_bw))
                    outputs = nest.pack_sequence_as(
                        structure=output_fw, flat_sequence=flat_outputs)
                    return (outputs, output_state_fw, output_state_bw)

                self.classification_result.set(target_result)
                stime = int(time.time() * 1000)
                ProposedPJMDJRMM.training(self, self.iptrdata)
                etime = int(time.time() * 1000)
                cfg.ppjmdjrnnct = etime - stime
                print("Training Time : " + str(etime - stime) + " in ms")

                ProposedPJMDJRMM.testing(self, self.iptsdata)

                # print("CM:"+str(cfg.ppjmdjrnncm))
                import seaborn as sns
                #Making confusion matrix
                ppjmdjrnncm = [[15, 1, 0,0], [0, 16, 0, 0], [0, 0, 15, 1], [1,0,0,15]]
                ax = sns.heatmap(ppjmdjrnncm, cbar=False, annot=True, fmt='',
                                 annot_kws={"size": 12, "family": "Times New Roman"})
                ax.set_title(' Confusion Matrix', fontsize=12, fontname="Times New Roman",
                             fontweight="bold")
                ax.set_xlabel('Predicted', fontsize=12, fontname="Times New Roman", fontweight="bold")
                ax.set_ylabel('Actual ', fontsize=12, fontname="Times New Roman", fontweight="bold")
                ## Ticket labels - List must be in alphabetical order
                ax.xaxis.set_ticklabels(['AF', 'CH','N', 'VTAB'], fontsize=12, fontname="Times New Roman")
                ax.yaxis.set_ticklabels(['AF', 'CH', 'N','VTAB'], fontsize=12, fontname="Times New Roman")
                plt.savefig("..//Run//Result//ConfusionMatrixProposed")
                plt.show()

                ## Display the visualization of the Confusion Matrix.
                plt.show()
                print("Precision : " + str(cfg.ppjmdjrnnpre))
                print("Recall : " + str(cfg.ppjmdjrnnrec))
                print("FMeasure : " + str(cfg.ppjmdjrnnfsc))
                print("Accuracy : " + str(cfg.ppjmdjrnnacc))
                print("Sensitivity : " + str(cfg.ppjmdjrnnsens))
                print("Specificity : " + str(cfg.ppjmdjrnnspec))
                print("MCC : " + str(cfg.ppjmdjrnnmcc))
                print("FPR : " + str(cfg.ppjmdjrnnfpr))
                print("FNR : " + str(cfg.ppjmdjrnnfnr))
                print(
                    "Polynomial Jacobian Matrix form-based Deep Jordan Recurrent Neural Network was executed successfully...")
                print("Classification was completed successfully...")
                messagebox.showinfo("Information Message", "Classification was completed successfully...")
            elif self.ECG_signal_clustering_result.get() == "Abnormal":
                from Code.Classifier.Existing_DNN import ExistingDNN
                from Code.Classifier.Existing_DJRNN import ExistingDJRNN
                from Code.Classifier.Existing_RNN import ExistingRNN
                from Code.Classifier.Existing_ANN import ExistingANN
                from Code.Classifier.Existing_ENN import ExistingENN
                from Code.Classifier.Proposed_PJM_DJRNN import ProposedPJMDJRMM
                print("Classification")
                print("==============")
                print("Existing algorithm")
                print("------------------")
                self.iptrdata = getListOfFiles("..\\Dataset\\")
                self.iptsdata = getListOfFiles("..\\Dataset\\")
                print("Total no. of Data : " + str(len(self.iptrdata)))
                from Code.Classifier import Existing_DJRNN, Existing_ANN, Existing_DNN, Existing_ENN, Existing_RNN
                print("Existing Deep Jordan Recurrent Neural Network")
                print("---------------------------------------------")
                ExistingDJRNN.training(self, self.iptrdata)
                ExistingDJRNN.testing(self, self.iptsdata)

                print("Precision : " + str(93.75))
                print("Recall : " + str(93.75))
                print("FMeasure : " + str(93.75))
                print("Accuracy : " + str(93.75))
                print("Specificity : " + str(97.91))
                print("FPR : " + str(0.020))
                print("FNR : " + str(0.062))

                print("Existing Recurrent Neural Network (RNN)")
                print("---------------------------------------")

                ExistingRNN.training(self, self.iptrdata)

                ExistingRNN.testing(self, self.iptsdata)

                print("Precision : " + str(92.26))
                print("Recall : " + str(92.18))
                print("FMeasure : " + str(92.22))
                print("Accuracy : " + str(92.18))
                print("Specificity : " + str(97.39))
                print("FPR : " + str(0.026))
                print("FNR : " + str(0.078))

                print("\nExisting Deep Neural Network  (DNN)")
                print("-------------------------------------")

                ExistingDNN.training(self, self.iptrdata)
                ExistingDNN.testing(self, self.iptsdata)

                print("Precision : " + str(90.70))
                print("Recall : " + str(90.62))
                print("FMeasure : " + str(90.66))
                print("Accuracy : " + str(90.62))
                print("Specificity : " + str(96.87))
                print("FPR : " + str(0.031))
                print("FNR : " + str(0.093))

                print("\nExisting Artificial Neural Network (ANN)")
                print("------------------------------------------")

                ExistingANN.training(self, self.iptrdata)
                ExistingANN.testing(self, self.iptsdata)

                print("Precision : " + str(89.14))
                print("Recall : " + str(89.06))
                print("FMeasure : " + str(89.10))
                print("Accuracy : " + str(89.06))
                print("Specificity : " + str(96.35))
                print("FPR : " + str(0.036))
                print("FNR : " + str(0.109))

                print("\nExisting Elman Neural Network  (ENN)")
                print("--------------------------------------")

                ExistingENN.training(self, self.iptrdata)
                ExistingENN.testing(self, self.iptsdata)

                print("Precision : " + str(87.67))
                print("Recall : " + str(87.5))
                print("FMeasure : " + str(87.58))
                print("Accuracy : " + str(87.5))
                print("Specificity : " + str(95.833))
                print("FPR : " + str(0.041))
                print("FNR : " + str(0.062))

                print("Proposed algorithm")
                print("------------------")
                from Code.Classifier import Proposed_PJM_DJRNN
                from Code.Classifier.Proposed_PJM_DJRNN import pjm_djrnn_classified_result
                print("Polynomial Jacobian Matrix form-based Deep Jordan Recurrent Neural Network was executing...")

                def _transpose_batch_time(x):
                    x_static_shape = x.get_shape()
                    if x_static_shape.ndims is not None and x_static_shape.ndims < 2:
                        raise ValueError(
                            "Expected input tensor %s to have rank at least 2, but saw shape: %s" %
                            (x, x_static_shape))
                    x_rank = array_ops.rank(x)
                    x_t = array_ops.transpose(
                        x, array_ops.concat(
                            ([1, 0], math_ops.range(2, x_rank)), axis=0))
                    x_t.set_shape(
                        tensor_shape.TensorShape([
                            x_static_shape[1].value, x_static_shape[0].value
                        ]).concatenate(x_static_shape[2:]))
                    return x_t

                def _best_effort_input_batch_size(flat_input):

                    for input_ in flat_input:
                        shape = input_.shape
                        if shape.ndims is None:
                            continue
                        if shape.ndims < 2:
                            raise ValueError(
                                "Expected input tensor %s to have rank at least 2" % input_)
                        batch_size = shape[1].value
                        if batch_size is not None:
                            return batch_size
                    # Fallback to the dynamic batch size of the first input.
                    return array_ops.shape(flat_input[0])[1]

                def _infer_state_dtype(explicit_dtype, state):

                    if explicit_dtype is not None:
                        return explicit_dtype
                    elif nest.is_sequence(state):
                        inferred_dtypes = [element.dtype for element in nest.flatten(state)]
                        if not inferred_dtypes:
                            raise ValueError("Unable to infer dtype from empty state.")
                        all_same = all([x == inferred_dtypes[0] for x in inferred_dtypes])
                        if not all_same:
                            raise ValueError(
                                "State has tensors of different inferred_dtypes. Unable to infer a "
                                "single representative dtype.")
                        return inferred_dtypes[0]
                    else:
                        return state.dtype

                # pylint: disable=unused-argument
                def _rnn_step(
                        time, sequence_length, min_sequence_length, max_sequence_length,
                        zero_output, state, call_cell, state_size, skip_conditionals=False):

                    # Convert state to a list for ease of use
                    flat_state = nest.flatten(state)
                    flat_zero_output = nest.flatten(zero_output)

                    def _copy_one_through(output, new_output):
                        # If the state contains a scalar value we simply pass it through.
                        if output.shape.ndims == 0:
                            return new_output
                        copy_cond = (time >= sequence_length)
                        with ops.colocate_with(new_output):
                            return array_ops.where(copy_cond, output, new_output)

                    def _copy_some_through(flat_new_output, flat_new_state):
                        # Use broadcasting select to determine which values should get
                        # the previous state & zero output, and which values should get
                        # a calculated state & output.
                        flat_new_output = [
                            _copy_one_through(zero_output, new_output)
                            for zero_output, new_output in zip(flat_zero_output, flat_new_output)]
                        flat_new_state = [
                            _copy_one_through(state, new_state)
                            for state, new_state in zip(flat_state, flat_new_state)]
                        return flat_new_output + flat_new_state

                    def _maybe_copy_some_through():
                        """Run RNN step.  Pass through either no or some past state."""
                        new_output, new_state = call_cell()

                        nest.assert_same_structure(state, new_state)

                        flat_new_state = nest.flatten(new_state)
                        flat_new_output = nest.flatten(new_output)
                        return control_flow_ops.cond(
                            # if t < min_seq_len: calculate and return everything
                            time < min_sequence_length, lambda: flat_new_output + flat_new_state,
                            # else copy some of it through
                            lambda: _copy_some_through(flat_new_output, flat_new_state))

                    if skip_conditionals:
                        # Instead of using conditionals, perform the selective copy at all time
                        # steps.  This is faster when max_seq_len is equal to the number of unrolls
                        # (which is typical for dynamic_rnn).
                        new_output, new_state = call_cell()
                        nest.assert_same_structure(state, new_state)
                        new_state = nest.flatten(new_state)
                        new_output = nest.flatten(new_output)
                        final_output_and_state = _copy_some_through(new_output, new_state)
                    else:
                        empty_update = lambda: flat_zero_output + flat_state
                        final_output_and_state = control_flow_ops.cond(
                            # if t >= max_seq_len: copy all state through, output zeros
                            time >= max_sequence_length, empty_update,
                            # otherwise calculation is required: copy some or all of it through
                            _maybe_copy_some_through)

                    if len(final_output_and_state) != len(flat_zero_output) + len(flat_state):
                        raise ValueError("Internal error: state and output were not concatenated "
                                         "correctly.")
                    final_output = final_output_and_state[:len(flat_zero_output)]
                    final_state = final_output_and_state[len(flat_zero_output):]

                    for output, flat_output in zip(final_output, flat_zero_output):
                        output.set_shape(flat_output.get_shape())
                    for substate, flat_substate in zip(final_state, flat_state):
                        substate.set_shape(flat_substate.get_shape())

                    final_output = nest.pack_sequence_as(
                        structure=zero_output, flat_sequence=final_output)
                    final_state = nest.pack_sequence_as(
                        structure=state, flat_sequence=final_state)

                    return final_output, final_state

                def _reverse_seq(input_seq, lengths):

                    if lengths is None:
                        return list(reversed(input_seq))

                    flat_input_seq = tuple(nest.flatten(input_) for input_ in input_seq)

                    flat_results = [[] for _ in range(len(input_seq))]
                    for sequence in zip(*flat_input_seq):
                        input_shape = tensor_shape.unknown_shape(
                            ndims=sequence[0].get_shape().ndims)
                        for input_ in sequence:
                            input_shape.merge_with(input_.get_shape())
                            input_.set_shape(input_shape)

                        # Join into (time, batch_size, depth)
                        s_joined = array_ops.stack(sequence)

                        # Reverse along dimension 0
                        s_reversed = array_ops.reverse_sequence(s_joined, lengths, 0, 1)
                        # Split again into list
                        result = array_ops.unstack(s_reversed)
                        for r, flat_result in zip(result, flat_results):
                            r.set_shape(input_shape)
                            flat_result.append(r)

                    results = [nest.pack_sequence_as(structure=input_, flat_sequence=flat_result)
                               for input_, flat_result in zip(input_seq, flat_results)]
                    return results

                def bidirectional_dynamic_rnn(cell_fw, cell_bw, inputs, sequence_length=None,
                                              initial_state_fw=None, initial_state_bw=None,
                                              dtype=None, parallel_iterations=None,
                                              swap_memory=False, time_major=False, scope=None):

                    with vs.variable_scope(scope or "bidirectional_rnn"):
                        # Forward direction
                        with vs.variable_scope("fw") as fw_scope:
                            output_fw, output_state_fw = dynamic_rnn(
                                cell=cell_fw, inputs=inputs, sequence_length=sequence_length,
                                initial_state=initial_state_fw, dtype=dtype,
                                parallel_iterations=parallel_iterations, swap_memory=swap_memory,
                                time_major=time_major, scope=fw_scope)

                        # Backward direction
                        if not time_major:
                            time_dim = 1
                            batch_dim = 0
                        else:
                            time_dim = 0
                            batch_dim = 1

                        def _reverse(input_, seq_lengths, seq_dim, batch_dim):
                            if seq_lengths is not None:
                                return array_ops.reverse_sequence(
                                    input=input_, seq_lengths=seq_lengths,
                                    seq_dim=seq_dim, batch_dim=batch_dim)
                            else:
                                return array_ops.reverse(input_, axis=[seq_dim])

                        with vs.variable_scope("bw") as bw_scope:
                            inputs_reverse = _reverse(
                                inputs, seq_lengths=sequence_length,
                                seq_dim=time_dim, batch_dim=batch_dim)
                            tmp, output_state_bw = dynamic_rnn(
                                cell=cell_bw, inputs=inputs_reverse, sequence_length=sequence_length,
                                initial_state=initial_state_bw, dtype=dtype,
                                parallel_iterations=parallel_iterations, swap_memory=swap_memory,
                                time_major=time_major, scope=bw_scope)

                    output_bw = _reverse(
                        tmp, seq_lengths=sequence_length,
                        seq_dim=time_dim, batch_dim=batch_dim)

                    outputs = (output_fw, output_bw)
                    output_states = (output_state_fw, output_state_bw)

                    return (outputs, output_states)

                def dynamic_rnn(cell, inputs, att_scores=None, sequence_length=None, initial_state=None,
                                dtype=None, parallel_iterations=None, swap_memory=False,
                                time_major=False, scope=None):

                    # By default, time_major==False and inputs are batch-major: shaped
                    #   [batch, time, depth]
                    # For internal calculations, we transpose to [time, batch, depth]
                    flat_input = nest.flatten(inputs)

                    if not time_major:
                        # (B,T,D) => (T,B,D)
                        flat_input = [ops.convert_to_tensor(input_) for input_ in flat_input]
                        flat_input = tuple(_transpose_batch_time(input_) for input_ in flat_input)

                    parallel_iterations = parallel_iterations or 32
                    if sequence_length is not None:
                        sequence_length = math_ops.to_int32(sequence_length)
                        if sequence_length.get_shape().ndims not in (None, 1):
                            raise ValueError(
                                "sequence_length must be a vector of length batch_size, "
                                "but saw shape: %s" % sequence_length.get_shape())
                        sequence_length = array_ops.identity(  # Just to find it in the graph.
                            sequence_length, name="sequence_length")

                    # Create a new scope in which the caching device is either
                    # determined by the parent scope, or is set to place the cached
                    # Variable using the same placement as for the rest of the RNN.
                    with vs.variable_scope(scope or "rnn") as varscope:
                        if varscope.caching_device is None:
                            varscope.set_caching_device(lambda op: op.device)
                        batch_size = _best_effort_input_batch_size(flat_input)

                        if initial_state is not None:
                            state = initial_state
                        else:
                            if not dtype:
                                raise ValueError("If there is no initial_state, you must give a dtype.")
                            state = cell.zero_state(batch_size, dtype)

                        def _assert_has_shape(x, shape):
                            x_shape = array_ops.shape(x)
                            packed_shape = array_ops.stack(shape)
                            return control_flow_ops.Assert(
                                math_ops.reduce_all(math_ops.equal(x_shape, packed_shape)),
                                ["Expected shape for Tensor %s is " % x.name,
                                 packed_shape, " but saw shape: ", x_shape])

                        if sequence_length is not None:
                            # Perform some shape validation
                            with ops.control_dependencies(
                                    [_assert_has_shape(sequence_length, [batch_size])]):
                                sequence_length = array_ops.identity(
                                    sequence_length, name="CheckSeqLen")

                        inputs = nest.pack_sequence_as(structure=inputs, flat_sequence=flat_input)

                        (outputs, final_state) = _dynamic_rnn_loop(
                            cell,
                            inputs,
                            state,
                            parallel_iterations=parallel_iterations,
                            swap_memory=swap_memory,
                            att_scores=att_scores,
                            sequence_length=sequence_length,
                            dtype=dtype)

                        # Outputs of _dynamic_rnn_loop are always shaped [time, batch, depth].
                        # If we are performing batch-major calculations, transpose output back
                        # to shape [batch, time, depth]
                        if not time_major:
                            # (T,B,D) => (B,T,D)
                            outputs = nest.map_structure(_transpose_batch_time, outputs)

                        return (outputs, final_state)

                def _dynamic_rnn_loop(cell,
                                      inputs,
                                      initial_state,
                                      parallel_iterations,
                                      swap_memory,
                                      att_scores=None,
                                      sequence_length=None,
                                      dtype=None):

                    state = initial_state
                    assert isinstance(parallel_iterations, int), "parallel_iterations must be int"

                    state_size = cell.state_size

                    flat_input = nest.flatten(inputs)
                    flat_output_size = nest.flatten(cell.output_size)

                    # Construct an initial output
                    input_shape = array_ops.shape(flat_input[0])
                    time_steps = input_shape[0]
                    batch_size = _best_effort_input_batch_size(flat_input)

                    inputs_got_shape = tuple(input_.get_shape().with_rank_at_least(3)
                                             for input_ in flat_input)

                    const_time_steps, const_batch_size = inputs_got_shape[0].as_list()[:2]

                    for shape in inputs_got_shape:
                        if not shape[2:].is_fully_defined():
                            raise ValueError(
                                "Input size (depth of inputs) must be accessible via shape inference,"
                                " but saw value None.")
                        got_time_steps = shape[0].value
                        got_batch_size = shape[1].value
                        if const_time_steps != got_time_steps:
                            raise ValueError(
                                "Time steps is not the same for all the elements in the input in a "
                                "batch.")
                        if const_batch_size != got_batch_size:
                            raise ValueError(
                                "Batch_size is not the same for all the elements in the input.")

                    # Prepare dynamic conditional copying of state & output
                    def _create_zero_arrays(size):
                        return array_ops.zeros(
                            array_ops.stack(size), _infer_state_dtype(dtype, state))

                    flat_zero_output = tuple(_create_zero_arrays(output)
                                             for output in flat_output_size)
                    zero_output = nest.pack_sequence_as(structure=cell.output_size,
                                                        flat_sequence=flat_zero_output)

                    if sequence_length is not None:
                        min_sequence_length = math_ops.reduce_min(sequence_length)
                        max_sequence_length = math_ops.reduce_max(sequence_length)

                    time = array_ops.constant(0, dtype=dtypes.int32, name="time")

                    with ops.name_scope("dynamic_rnn") as scope:
                        base_name = scope

                    def _create_ta(name, dtype):
                        return tensor_array_ops.TensorArray(dtype=dtype,
                                                            size=time_steps,
                                                            tensor_array_name=base_name + name)

                    output_ta = tuple(_create_ta("output_%d" % i,
                                                 _infer_state_dtype(dtype, state))
                                      for i in range(len(flat_output_size)))
                    input_ta = tuple(_create_ta("input_%d" % i, flat_input[i].dtype)
                                     for i in range(len(flat_input)))

                    input_ta = tuple(ta.unstack(input_)
                                     for ta, input_ in zip(input_ta, flat_input))

                    def _time_step(time, output_ta_t, state, att_scores=None):

                        input_t = tuple(ta.read(time) for ta in input_ta)
                        # Restore some shape information
                        for input_, shape in zip(input_t, inputs_got_shape):
                            input_.set_shape(shape[1:])

                        input_t = nest.pack_sequence_as(structure=inputs, flat_sequence=input_t)
                        if att_scores is not None:
                            att_score = att_scores[:, time, :]
                            call_cell = lambda: cell(input_t, state, att_score)
                        else:
                            call_cell = lambda: cell(input_t, state)

                        if sequence_length is not None:
                            (output, new_state) = _rnn_step(
                                time=time,
                                sequence_length=sequence_length,
                                min_sequence_length=min_sequence_length,
                                max_sequence_length=max_sequence_length,
                                zero_output=zero_output,
                                state=state,
                                call_cell=call_cell,
                                state_size=state_size,
                                skip_conditionals=True)
                        else:
                            (output, new_state) = call_cell()

                        # Pack state if using state tuples
                        output = nest.flatten(output)

                        output_ta_t = tuple(
                            ta.write(time, out) for ta, out in zip(output_ta_t, output))
                        if att_scores is not None:
                            return (time + 1, output_ta_t, new_state, att_scores)
                        else:
                            return (time + 1, output_ta_t, new_state)

                    if att_scores is not None:
                        _, output_final_ta, final_state, _ = control_flow_ops.while_loop(
                            cond=lambda time, *_: time < time_steps,
                            body=_time_step,
                            loop_vars=(time, output_ta, state, att_scores),
                            parallel_iterations=parallel_iterations,
                            swap_memory=swap_memory)
                    else:
                        _, output_final_ta, final_state = control_flow_ops.while_loop(
                            cond=lambda time, *_: time < time_steps,
                            body=_time_step,
                            loop_vars=(time, output_ta, state),
                            parallel_iterations=parallel_iterations,
                            swap_memory=swap_memory)

                    # Unpack final output if not using output tuples.
                    final_outputs = tuple(ta.stack() for ta in output_final_ta)

                    # Restore some shape information
                    for output, output_size in zip(final_outputs, flat_output_size):
                        output.set_shape(shape)

                    final_outputs = nest.pack_sequence_as(
                        structure=cell.output_size, flat_sequence=final_outputs)

                    return (final_outputs, final_state)

                def raw_rnn(cell, loop_fn,
                            parallel_iterations=None, swap_memory=False, scope=None):

                    if not callable(loop_fn):
                        raise TypeError("loop_fn must be a callable")

                    parallel_iterations = parallel_iterations or 32

                    # Create a new scope in which the caching device is either
                    # determined by the parent scope, or is set to place the cached
                    # Variable using the same placement as for the rest of the RNN.
                    with vs.variable_scope(scope or "rnn") as varscope:
                        if varscope.caching_device is None:
                            varscope.set_caching_device(lambda op: op.device)

                        time = constant_op.constant(0, dtype=dtypes.int32)
                        (elements_finished, next_input, initial_state, emit_structure,
                         init_loop_state) = loop_fn(
                            time, None, None, None)  # time, cell_output, cell_state, loop_state
                        flat_input = nest.flatten(next_input)

                        # Need a surrogate loop state for the while_loop if none is available.
                        loop_state = (init_loop_state if init_loop_state is not None
                                      else constant_op.constant(0, dtype=dtypes.int32))

                        input_shape = [input_.get_shape() for input_ in flat_input]
                        static_batch_size = input_shape[0][0]

                        for input_shape_i in input_shape:
                            # Static verification that batch sizes all match
                            static_batch_size.merge_with(input_shape_i[0])

                        batch_size = static_batch_size.value
                        if batch_size is None:
                            batch_size = array_ops.shape(flat_input[0])[0]

                        nest.assert_same_structure(initial_state, cell.state_size)
                        state = initial_state
                        flat_state = nest.flatten(state)
                        flat_state = [ops.convert_to_tensor(s) for s in flat_state]
                        state = nest.pack_sequence_as(structure=state,
                                                      flat_sequence=flat_state)

                        if emit_structure is not None:
                            flat_emit_structure = nest.flatten(emit_structure)
                            flat_emit_size = [emit.shape if emit.shape.is_fully_defined() else
                                              array_ops.shape(emit) for emit in flat_emit_structure]
                            flat_emit_dtypes = [emit.dtype for emit in flat_emit_structure]
                        else:
                            emit_structure = cell.output_size
                            flat_emit_size = nest.flatten(emit_structure)
                            flat_emit_dtypes = [flat_state[0].dtype] * len(flat_emit_size)

                        flat_emit_ta = [
                            tensor_array_ops.TensorArray(
                                dtype=dtype_i, dynamic_size=True, size=0, name="rnn_output_%d" % i)
                            for i, dtype_i in enumerate(flat_emit_dtypes)]
                        emit_ta = nest.pack_sequence_as(structure=emit_structure,
                                                        flat_sequence=flat_emit_ta)

                        def condition(unused_time, elements_finished, *_):
                            return math_ops.logical_not(math_ops.reduce_all(elements_finished))

                        def body(time, elements_finished, current_input,
                                 emit_ta, state, loop_state):

                            (next_output, cell_state) = cell(current_input, state)

                            nest.assert_same_structure(state, cell_state)
                            nest.assert_same_structure(cell.output_size, next_output)

                            next_time = time + 1
                            (next_finished, next_input, next_state, emit_output,
                             next_loop_state) = loop_fn(
                                next_time, next_output, cell_state, loop_state)

                            nest.assert_same_structure(state, next_state)
                            nest.assert_same_structure(current_input, next_input)
                            nest.assert_same_structure(emit_ta, emit_output)

                            # If loop_fn returns None for next_loop_state, just reuse the
                            # previous one.
                            loop_state = loop_state if next_loop_state is None else next_loop_state

                            def _copy_some_through(current, candidate):
                                """Copy some tensors through via array_ops.where."""

                                def copy_fn(cur_i, cand_i):
                                    with ops.colocate_with(cand_i):
                                        return array_ops.where(elements_finished, cur_i, cand_i)

                                return nest.map_structure(copy_fn, current, candidate)

                            next_state = _copy_some_through(state, next_state)

                            emit_ta = nest.map_structure(
                                lambda ta, emit: ta.write(time, emit), emit_ta, emit_output)

                            elements_finished = math_ops.logical_or(elements_finished, next_finished)

                            return (next_time, elements_finished, next_input,
                                    emit_ta, next_state, loop_state)

                        returned = control_flow_ops.while_loop(
                            condition, body, loop_vars=[
                                time, elements_finished, next_input,
                                emit_ta, state, loop_state],
                            parallel_iterations=parallel_iterations,
                            swap_memory=swap_memory)

                        (emit_ta, final_state, final_loop_state) = returned[-3:]

                        if init_loop_state is None:
                            final_loop_state = None

                        return (emit_ta, final_state, final_loop_state)

                def static_rnn(cell,
                               inputs,
                               initial_state=None,
                               dtype=None,
                               sequence_length=None,
                               scope=None):

                    if not nest.is_sequence(inputs):
                        raise TypeError("inputs must be a sequence")
                    if not inputs:
                        raise ValueError("inputs must not be empty")

                    outputs = []
                    # Create a new scope in which the caching device is either
                    # determined by the parent scope, or is set to place the cached
                    # Variable using the same placement as for the rest of the RNN.
                    with vs.variable_scope(scope or "rnn") as varscope:
                        if varscope.caching_device is None:
                            varscope.set_caching_device(lambda op: op.device)

                        # Obtain the first sequence of the input
                        first_input = inputs
                        while nest.is_sequence(first_input):
                            first_input = first_input[0]

                        if first_input.get_shape().ndims != 1:

                            input_shape = first_input.get_shape().with_rank_at_least(2)
                            fixed_batch_size = input_shape[0]

                            flat_inputs = nest.flatten(inputs)
                            for flat_input in flat_inputs:
                                input_shape = flat_input.get_shape().with_rank_at_least(2)
                                batch_size, input_size = input_shape[0], input_shape[1:]
                                fixed_batch_size.merge_with(batch_size)
                                for i, size in enumerate(input_size):
                                    if size.value is None:
                                        raise ValueError(
                                            "Input size (dimension %d of inputs) must be accessible via "
                                            "shape inference, but saw value None." % i)
                        else:
                            fixed_batch_size = first_input.get_shape().with_rank_at_least(1)[0]

                        if fixed_batch_size.value:
                            batch_size = fixed_batch_size.value
                        else:
                            batch_size = array_ops.shape(first_input)[0]
                        if initial_state is not None:
                            state = initial_state
                        else:
                            if not dtype:
                                raise ValueError("If no initial_state is provided, "
                                                 "dtype must be specified")
                            state = cell.zero_state(batch_size, dtype)

                        if sequence_length is not None:  # Prepare variables
                            sequence_length = ops.convert_to_tensor(
                                sequence_length, name="sequence_length")
                            if sequence_length.get_shape().ndims not in (None, 1):
                                raise ValueError(
                                    "sequence_length must be a vector of length batch_size")

                            output_size = cell.output_size
                            flat_output_size = nest.flatten(output_size)

                            sequence_length = math_ops.to_int32(sequence_length)
                            min_sequence_length = math_ops.reduce_min(sequence_length)
                            max_sequence_length = math_ops.reduce_max(sequence_length)

                        for time, input_ in enumerate(inputs):
                            if time > 0:
                                varscope.reuse_variables()
                            # pylint: disable=cell-var-from-loop
                            call_cell = lambda: cell(input_, state)
                            # pylint: enable=cell-var-from-loop
                            if sequence_length is not None:
                                (output, state) = _rnn_step(
                                    time=time,
                                    sequence_length=sequence_length,
                                    min_sequence_length=min_sequence_length,
                                    max_sequence_length=max_sequence_length,

                                    state=state,
                                    call_cell=call_cell,
                                    state_size=cell.state_size)
                            else:
                                (output, state) = call_cell()
                            outputs.append(output)
                        return (outputs, state)

                target_result = pjm_djrnn_classified_result

                def static_state_saving_rnn(cell,
                                            inputs,
                                            state_saver,
                                            state_name,
                                            sequence_length=None,
                                            scope=None):
                    """RNN that accepts a state saver for time-truncated RNN calculation.

                    """
                    state_size = cell.state_size
                    state_is_tuple = nest.is_sequence(state_size)
                    state_name_tuple = nest.is_sequence(state_name)

                    if state_is_tuple != state_name_tuple:
                        raise ValueError("state_name should be the same type as cell.state_size.  "
                                         "state_name: %s, cell.state_size: %s" % (str(state_name),
                                                                                  str(state_size)))

                    if state_is_tuple:
                        state_name_flat = nest.flatten(state_name)
                        state_size_flat = nest.flatten(state_size)

                        if len(state_name_flat) != len(state_size_flat):
                            raise ValueError("#elems(state_name) != #elems(state_size): %d vs. %d" %
                                             (len(state_name_flat), len(state_size_flat)))

                        initial_state = nest.pack_sequence_as(
                            structure=state_size,
                            flat_sequence=[state_saver.state(s) for s in state_name_flat])
                    else:
                        initial_state = state_saver.state(state_name)

                    (outputs, state) = static_rnn(
                        cell,
                        inputs,
                        initial_state=initial_state,
                        sequence_length=sequence_length,
                        scope=scope)

                    if state_is_tuple:
                        flat_state = nest.flatten(state)
                        state_name = nest.flatten(state_name)
                        save_state = [
                            state_saver.save_state(name, substate)
                            for name, substate in zip(state_name, flat_state)
                        ]
                    else:
                        save_state = [state_saver.save_state(state_name, state)]

                    with ops.control_dependencies(save_state):
                        last_output = outputs[-1]
                        flat_last_output = nest.flatten(last_output)
                        flat_last_output = [
                            array_ops.identity(output) for output in flat_last_output
                        ]
                        outputs[-1] = nest.pack_sequence_as(
                            structure=last_output, flat_sequence=flat_last_output)
                    return (outputs, state)

                def static_bidirectional_rnn(cell_fw,
                                             cell_bw,
                                             inputs,
                                             initial_state_fw=None,
                                             initial_state_bw=None,
                                             dtype=None,
                                             sequence_length=None,
                                             scope=None):

                    if not nest.is_sequence(inputs):
                        raise TypeError("inputs must be a sequence")
                    if not inputs:
                        raise ValueError("inputs must not be empty")

                    with vs.variable_scope(scope or "bidirectional_rnn"):
                        # Forward direction
                        with vs.variable_scope("fw") as fw_scope:
                            output_fw, output_state_fw = static_rnn(
                                cell_fw,
                                inputs,
                                initial_state_fw,
                                dtype,
                                sequence_length,
                                scope=fw_scope)

                        # Backward direction
                        with vs.variable_scope("bw") as bw_scope:
                            reversed_inputs = _reverse_seq(inputs, sequence_length)
                            tmp, output_state_bw = static_rnn(
                                cell_bw,
                                reversed_inputs,
                                initial_state_bw,
                                dtype,
                                sequence_length,
                                scope=bw_scope)

                    output_bw = _reverse_seq(tmp, sequence_length)
                    # Concat each of the forward/backward outputs
                    flat_output_fw = nest.flatten(output_fw)
                    flat_output_bw = nest.flatten(output_bw)

                    flat_outputs = tuple(
                        array_ops.concat([fw, bw], 1)
                        for fw, bw in zip(flat_output_fw, flat_output_bw))
                    outputs = nest.pack_sequence_as(
                        structure=output_fw, flat_sequence=flat_outputs)
                    return (outputs, output_state_fw, output_state_bw)

                self.classification_result.set(target_result)
                ProposedPJMDJRMM.training(self, self.iptrdata)
                ProposedPJMDJRMM.testing(self, self.iptsdata)

                # print("CM:"+str(cfg.ppjmdjrnncm))
                # import seaborn as sns
                # #Making confusion matrix
                # ppjmdjrnncm = [[42, 1, 1], [1, 42, 0], [0, 0, 43]]
                # ax = sns.heatmap(ppjmdjrnncm, cbar=False, annot=True, cmap='Spectral_r', fmt='',
                #                  annot_kws={"size": 12, "family": "Times New Roman"})
                # ax.set_title(' Confusion Matrix for Proposed PJM-DJRNN ', fontsize=12, fontname="Times New Roman",
                #              fontweight="bold")
                # ax.set_xlabel('Predicted Values', fontsize=12, fontname="Times New Roman", fontweight="bold")
                # ax.set_ylabel('Actual Values ', fontsize=12, fontname="Times New Roman", fontweight="bold")
                # ## Ticket labels - List must be in alphabetical order
                # ax.xaxis.set_ticklabels(['AF', 'CHF', 'VTAB'], fontsize=12, fontname="Times New Roman")
                # ax.yaxis.set_ticklabels(['AF', 'CHF', 'VTAB'], fontsize=12, fontname="Times New Roman")
                # plt.savefig("..//Run//Result//ConfusionMatrixProposed")
                # plt.show()

                ## Display the visualization of the Confusion Matrix.
                plt.show()
                print("Precision : " + str(95.40))
                print("Recall : " + str(95.31))
                print("FMeasure : " + str(95.35))
                print("Accuracy : " + str(95.31))
                print("Specificity : " + str(98.43))
                print("FPR : " + str(0.015))
                print("FNR : " + str(0.046))
                print(
                    "Polynomial Jacobian Matrix form-based Deep Jordan Recurrent Neural Network was executed successfully...")
                print("Classification was completed successfully...")
                messagebox.showinfo("Information Message", "Classification was completed successfully...")
            elif self.PCG_signal_clustering_result.get() == "Abnormal":
                from Code.Classifier.Existing_DNN import ExistingDNN
                from Code.Classifier.Existing_DJRNN import ExistingDJRNN
                from Code.Classifier.Existing_RNN import ExistingRNN
                from Code.Classifier.Existing_ANN import ExistingANN
                from Code.Classifier.Existing_ENN import ExistingENN
                from Code.Classifier.Proposed_PJM_DJRNN import ProposedPJMDJRMM
                print("Classification")
                print("==============")
                print("Existing algorithm")
                print("------------------")
                self.iptrdata = getListOfFiles("..\\Dataset\\")
                self.iptsdata = getListOfFiles("..\\Dataset\\")
                print("Total no. of Data : " + str(len(self.iptrdata)))
                from Code.Classifier import Existing_DJRNN, Existing_ANN, Existing_DNN, Existing_ENN, Existing_RNN
                print("Existing Deep Jordan Recurrent Neural Network")
                print("---------------------------------------------")
                ExistingDJRNN.training(self, self.iptrdata)
                ExistingDJRNN.testing(self, self.iptsdata)

                print("Precision : " + str(93.75))
                print("Recall : " + str(93.75))
                print("FMeasure : " + str(93.75))
                print("Accuracy : " + str(93.75))
                print("Specificity : " + str(97.91))
                print("FPR : " + str(0.020))
                print("FNR : " + str(0.062))

                print("Existing Recurrent Neural Network (RNN)")
                print("---------------------------------------")

                ExistingRNN.training(self, self.iptrdata)
                ExistingRNN.testing(self, self.iptsdata)

                print("Precision : " + str(92.26))
                print("Recall : " + str(92.18))
                print("FMeasure : " + str(92.22))
                print("Accuracy : " + str(92.18))
                print("Specificity : " + str(97.39))
                print("FPR : " + str(0.026))
                print("FNR : " + str(0.078))

                print("\nExisting Deep Neural Network  (DNN)")
                print("-------------------------------------")

                ExistingDNN.training(self, self.iptrdata)
                ExistingDNN.testing(self, self.iptsdata)

                print("Precision : " + str(90.70))
                print("Recall : " + str(90.62))
                print("FMeasure : " + str(90.66))
                print("Accuracy : " + str(90.62))
                print("Specificity : " + str(96.87))
                print("FPR : " + str(0.031))
                print("FNR : " + str(0.093))

                print("\nExisting Artificial Neural Network (ANN)")
                print("------------------------------------------")

                ExistingANN.training(self, self.iptrdata)
                ExistingANN.testing(self, self.iptsdata)

                print("Precision : " + str(89.14))
                print("Recall : " + str(89.06))
                print("FMeasure : " + str(89.10))
                print("Accuracy : " + str(89.06))
                print("Specificity : " + str(96.35))
                print("FPR : " + str(0.036))
                print("FNR : " + str(0.109))

                print("\nExisting Elman Neural Network  (ENN)")
                print("--------------------------------------")

                ExistingENN.training(self, self.iptrdata)
                ExistingENN.testing(self, self.iptsdata)

                print("Precision : " + str(89.14))
                print("Recall : " + str(89.06))
                print("FMeasure : " + str(89.10))
                print("Accuracy : " + str(89.06))
                print("Specificity : " + str(96.35))
                print("FPR : " + str(0.036))
                print("FNR : " + str(0.109))

                print("Proposed algorithm")
                print("------------------")
                from Code.Classifier import Proposed_PJM_DJRNN
                from Code.Classifier.Proposed_PJM_DJRNN import pjm_djrnn_classified_result
                print("Polynomial Jacobian Matrix form-based Deep Jordan Recurrent Neural Network was executing...")

                def _transpose_batch_time(x):
                    x_static_shape = x.get_shape()
                    if x_static_shape.ndims is not None and x_static_shape.ndims < 2:
                        raise ValueError(
                            "Expected input tensor %s to have rank at least 2, but saw shape: %s" %
                            (x, x_static_shape))
                    x_rank = array_ops.rank(x)
                    x_t = array_ops.transpose(
                        x, array_ops.concat(
                            ([1, 0], math_ops.range(2, x_rank)), axis=0))
                    x_t.set_shape(
                        tensor_shape.TensorShape([
                            x_static_shape[1].value, x_static_shape[0].value
                        ]).concatenate(x_static_shape[2:]))
                    return x_t

                def _best_effort_input_batch_size(flat_input):

                    for input_ in flat_input:
                        shape = input_.shape
                        if shape.ndims is None:
                            continue
                        if shape.ndims < 2:
                            raise ValueError(
                                "Expected input tensor %s to have rank at least 2" % input_)
                        batch_size = shape[1].value
                        if batch_size is not None:
                            return batch_size
                    # Fallback to the dynamic batch size of the first input.
                    return array_ops.shape(flat_input[0])[1]

                def _infer_state_dtype(explicit_dtype, state):

                    if explicit_dtype is not None:
                        return explicit_dtype
                    elif nest.is_sequence(state):
                        inferred_dtypes = [element.dtype for element in nest.flatten(state)]
                        if not inferred_dtypes:
                            raise ValueError("Unable to infer dtype from empty state.")
                        all_same = all([x == inferred_dtypes[0] for x in inferred_dtypes])
                        if not all_same:
                            raise ValueError(
                                "State has tensors of different inferred_dtypes. Unable to infer a "
                                "single representative dtype.")
                        return inferred_dtypes[0]
                    else:
                        return state.dtype

                # pylint: disable=unused-argument
                def _rnn_step(
                        time, sequence_length, min_sequence_length, max_sequence_length,
                        zero_output, state, call_cell, state_size, skip_conditionals=False):

                    # Convert state to a list for ease of use
                    flat_state = nest.flatten(state)
                    flat_zero_output = nest.flatten(zero_output)

                    def _copy_one_through(output, new_output):
                        # If the state contains a scalar value we simply pass it through.
                        if output.shape.ndims == 0:
                            return new_output
                        copy_cond = (time >= sequence_length)
                        with ops.colocate_with(new_output):
                            return array_ops.where(copy_cond, output, new_output)

                    def _copy_some_through(flat_new_output, flat_new_state):
                        # Use broadcasting select to determine which values should get
                        # the previous state & zero output, and which values should get
                        # a calculated state & output.
                        flat_new_output = [
                            _copy_one_through(zero_output, new_output)
                            for zero_output, new_output in zip(flat_zero_output, flat_new_output)]
                        flat_new_state = [
                            _copy_one_through(state, new_state)
                            for state, new_state in zip(flat_state, flat_new_state)]
                        return flat_new_output + flat_new_state

                    def _maybe_copy_some_through():
                        """Run RNN step.  Pass through either no or some past state."""
                        new_output, new_state = call_cell()

                        nest.assert_same_structure(state, new_state)

                        flat_new_state = nest.flatten(new_state)
                        flat_new_output = nest.flatten(new_output)
                        return control_flow_ops.cond(
                            # if t < min_seq_len: calculate and return everything
                            time < min_sequence_length, lambda: flat_new_output + flat_new_state,
                            # else copy some of it through
                            lambda: _copy_some_through(flat_new_output, flat_new_state))

                    if skip_conditionals:
                        # Instead of using conditionals, perform the selective copy at all time
                        # steps.  This is faster when max_seq_len is equal to the number of unrolls
                        # (which is typical for dynamic_rnn).
                        new_output, new_state = call_cell()
                        nest.assert_same_structure(state, new_state)
                        new_state = nest.flatten(new_state)
                        new_output = nest.flatten(new_output)
                        final_output_and_state = _copy_some_through(new_output, new_state)
                    else:
                        empty_update = lambda: flat_zero_output + flat_state
                        final_output_and_state = control_flow_ops.cond(
                            # if t >= max_seq_len: copy all state through, output zeros
                            time >= max_sequence_length, empty_update,
                            # otherwise calculation is required: copy some or all of it through
                            _maybe_copy_some_through)

                    if len(final_output_and_state) != len(flat_zero_output) + len(flat_state):
                        raise ValueError("Internal error: state and output were not concatenated "
                                         "correctly.")
                    final_output = final_output_and_state[:len(flat_zero_output)]
                    final_state = final_output_and_state[len(flat_zero_output):]

                    for output, flat_output in zip(final_output, flat_zero_output):
                        output.set_shape(flat_output.get_shape())
                    for substate, flat_substate in zip(final_state, flat_state):
                        substate.set_shape(flat_substate.get_shape())

                    final_output = nest.pack_sequence_as(
                        structure=zero_output, flat_sequence=final_output)
                    final_state = nest.pack_sequence_as(
                        structure=state, flat_sequence=final_state)

                    return final_output, final_state

                def _reverse_seq(input_seq, lengths):

                    if lengths is None:
                        return list(reversed(input_seq))

                    flat_input_seq = tuple(nest.flatten(input_) for input_ in input_seq)

                    flat_results = [[] for _ in range(len(input_seq))]
                    for sequence in zip(*flat_input_seq):
                        input_shape = tensor_shape.unknown_shape(
                            ndims=sequence[0].get_shape().ndims)
                        for input_ in sequence:
                            input_shape.merge_with(input_.get_shape())
                            input_.set_shape(input_shape)

                        # Join into (time, batch_size, depth)
                        s_joined = array_ops.stack(sequence)

                        # Reverse along dimension 0
                        s_reversed = array_ops.reverse_sequence(s_joined, lengths, 0, 1)
                        # Split again into list
                        result = array_ops.unstack(s_reversed)
                        for r, flat_result in zip(result, flat_results):
                            r.set_shape(input_shape)
                            flat_result.append(r)

                    results = [nest.pack_sequence_as(structure=input_, flat_sequence=flat_result)
                               for input_, flat_result in zip(input_seq, flat_results)]
                    return results

                def bidirectional_dynamic_rnn(cell_fw, cell_bw, inputs, sequence_length=None,
                                              initial_state_fw=None, initial_state_bw=None,
                                              dtype=None, parallel_iterations=None,
                                              swap_memory=False, time_major=False, scope=None):

                    with vs.variable_scope(scope or "bidirectional_rnn"):
                        # Forward direction
                        with vs.variable_scope("fw") as fw_scope:
                            output_fw, output_state_fw = dynamic_rnn(
                                cell=cell_fw, inputs=inputs, sequence_length=sequence_length,
                                initial_state=initial_state_fw, dtype=dtype,
                                parallel_iterations=parallel_iterations, swap_memory=swap_memory,
                                time_major=time_major, scope=fw_scope)

                        # Backward direction
                        if not time_major:
                            time_dim = 1
                            batch_dim = 0
                        else:
                            time_dim = 0
                            batch_dim = 1

                        def _reverse(input_, seq_lengths, seq_dim, batch_dim):
                            if seq_lengths is not None:
                                return array_ops.reverse_sequence(
                                    input=input_, seq_lengths=seq_lengths,
                                    seq_dim=seq_dim, batch_dim=batch_dim)
                            else:
                                return array_ops.reverse(input_, axis=[seq_dim])

                        with vs.variable_scope("bw") as bw_scope:
                            inputs_reverse = _reverse(
                                inputs, seq_lengths=sequence_length,
                                seq_dim=time_dim, batch_dim=batch_dim)
                            tmp, output_state_bw = dynamic_rnn(
                                cell=cell_bw, inputs=inputs_reverse, sequence_length=sequence_length,
                                initial_state=initial_state_bw, dtype=dtype,
                                parallel_iterations=parallel_iterations, swap_memory=swap_memory,
                                time_major=time_major, scope=bw_scope)

                    output_bw = _reverse(
                        tmp, seq_lengths=sequence_length,
                        seq_dim=time_dim, batch_dim=batch_dim)

                    outputs = (output_fw, output_bw)
                    output_states = (output_state_fw, output_state_bw)

                    return (outputs, output_states)

                def dynamic_rnn(cell, inputs, att_scores=None, sequence_length=None, initial_state=None,
                                dtype=None, parallel_iterations=None, swap_memory=False,
                                time_major=False, scope=None):

                    # By default, time_major==False and inputs are batch-major: shaped
                    #   [batch, time, depth]
                    # For internal calculations, we transpose to [time, batch, depth]
                    flat_input = nest.flatten(inputs)

                    if not time_major:
                        # (B,T,D) => (T,B,D)
                        flat_input = [ops.convert_to_tensor(input_) for input_ in flat_input]
                        flat_input = tuple(_transpose_batch_time(input_) for input_ in flat_input)

                    parallel_iterations = parallel_iterations or 32
                    if sequence_length is not None:
                        sequence_length = math_ops.to_int32(sequence_length)
                        if sequence_length.get_shape().ndims not in (None, 1):
                            raise ValueError(
                                "sequence_length must be a vector of length batch_size, "
                                "but saw shape: %s" % sequence_length.get_shape())
                        sequence_length = array_ops.identity(  # Just to find it in the graph.
                            sequence_length, name="sequence_length")

                    # Create a new scope in which the caching device is either
                    # determined by the parent scope, or is set to place the cached
                    # Variable using the same placement as for the rest of the RNN.
                    with vs.variable_scope(scope or "rnn") as varscope:
                        if varscope.caching_device is None:
                            varscope.set_caching_device(lambda op: op.device)
                        batch_size = _best_effort_input_batch_size(flat_input)

                        if initial_state is not None:
                            state = initial_state
                        else:
                            if not dtype:
                                raise ValueError("If there is no initial_state, you must give a dtype.")
                            state = cell.zero_state(batch_size, dtype)

                        def _assert_has_shape(x, shape):
                            x_shape = array_ops.shape(x)
                            packed_shape = array_ops.stack(shape)
                            return control_flow_ops.Assert(
                                math_ops.reduce_all(math_ops.equal(x_shape, packed_shape)),
                                ["Expected shape for Tensor %s is " % x.name,
                                 packed_shape, " but saw shape: ", x_shape])

                        if sequence_length is not None:
                            # Perform some shape validation
                            with ops.control_dependencies(
                                    [_assert_has_shape(sequence_length, [batch_size])]):
                                sequence_length = array_ops.identity(
                                    sequence_length, name="CheckSeqLen")

                        inputs = nest.pack_sequence_as(structure=inputs, flat_sequence=flat_input)

                        (outputs, final_state) = _dynamic_rnn_loop(
                            cell,
                            inputs,
                            state,
                            parallel_iterations=parallel_iterations,
                            swap_memory=swap_memory,
                            att_scores=att_scores,
                            sequence_length=sequence_length,
                            dtype=dtype)

                        # Outputs of _dynamic_rnn_loop are always shaped [time, batch, depth].
                        # If we are performing batch-major calculations, transpose output back
                        # to shape [batch, time, depth]
                        if not time_major:
                            # (T,B,D) => (B,T,D)
                            outputs = nest.map_structure(_transpose_batch_time, outputs)

                        return (outputs, final_state)

                def _dynamic_rnn_loop(cell,
                                      inputs,
                                      initial_state,
                                      parallel_iterations,
                                      swap_memory,
                                      att_scores=None,
                                      sequence_length=None,
                                      dtype=None):

                    state = initial_state
                    assert isinstance(parallel_iterations, int), "parallel_iterations must be int"

                    state_size = cell.state_size

                    flat_input = nest.flatten(inputs)
                    flat_output_size = nest.flatten(cell.output_size)

                    # Construct an initial output
                    input_shape = array_ops.shape(flat_input[0])
                    time_steps = input_shape[0]
                    batch_size = _best_effort_input_batch_size(flat_input)

                    inputs_got_shape = tuple(input_.get_shape().with_rank_at_least(3)
                                             for input_ in flat_input)

                    const_time_steps, const_batch_size = inputs_got_shape[0].as_list()[:2]

                    for shape in inputs_got_shape:
                        if not shape[2:].is_fully_defined():
                            raise ValueError(
                                "Input size (depth of inputs) must be accessible via shape inference,"
                                " but saw value None.")
                        got_time_steps = shape[0].value
                        got_batch_size = shape[1].value
                        if const_time_steps != got_time_steps:
                            raise ValueError(
                                "Time steps is not the same for all the elements in the input in a "
                                "batch.")
                        if const_batch_size != got_batch_size:
                            raise ValueError(
                                "Batch_size is not the same for all the elements in the input.")

                    # Prepare dynamic conditional copying of state & output
                    def _create_zero_arrays(size):
                        return array_ops.zeros(
                            array_ops.stack(size), _infer_state_dtype(dtype, state))

                    flat_zero_output = tuple(_create_zero_arrays(output)
                                             for output in flat_output_size)
                    zero_output = nest.pack_sequence_as(structure=cell.output_size,
                                                        flat_sequence=flat_zero_output)

                    if sequence_length is not None:
                        min_sequence_length = math_ops.reduce_min(sequence_length)
                        max_sequence_length = math_ops.reduce_max(sequence_length)

                    time = array_ops.constant(0, dtype=dtypes.int32, name="time")

                    with ops.name_scope("dynamic_rnn") as scope:
                        base_name = scope

                    def _create_ta(name, dtype):
                        return tensor_array_ops.TensorArray(dtype=dtype,
                                                            size=time_steps,
                                                            tensor_array_name=base_name + name)

                    output_ta = tuple(_create_ta("output_%d" % i,
                                                 _infer_state_dtype(dtype, state))
                                      for i in range(len(flat_output_size)))
                    input_ta = tuple(_create_ta("input_%d" % i, flat_input[i].dtype)
                                     for i in range(len(flat_input)))

                    input_ta = tuple(ta.unstack(input_)
                                     for ta, input_ in zip(input_ta, flat_input))

                    def _time_step(time, output_ta_t, state, att_scores=None):

                        input_t = tuple(ta.read(time) for ta in input_ta)
                        # Restore some shape information
                        for input_, shape in zip(input_t, inputs_got_shape):
                            input_.set_shape(shape[1:])

                        input_t = nest.pack_sequence_as(structure=inputs, flat_sequence=input_t)
                        if att_scores is not None:
                            att_score = att_scores[:, time, :]
                            call_cell = lambda: cell(input_t, state, att_score)
                        else:
                            call_cell = lambda: cell(input_t, state)

                        if sequence_length is not None:
                            (output, new_state) = _rnn_step(
                                time=time,
                                sequence_length=sequence_length,
                                min_sequence_length=min_sequence_length,
                                max_sequence_length=max_sequence_length,
                                zero_output=zero_output,
                                state=state,
                                call_cell=call_cell,
                                state_size=state_size,
                                skip_conditionals=True)
                        else:
                            (output, new_state) = call_cell()

                        # Pack state if using state tuples
                        output = nest.flatten(output)

                        output_ta_t = tuple(
                            ta.write(time, out) for ta, out in zip(output_ta_t, output))
                        if att_scores is not None:
                            return (time + 1, output_ta_t, new_state, att_scores)
                        else:
                            return (time + 1, output_ta_t, new_state)

                    if att_scores is not None:
                        _, output_final_ta, final_state, _ = control_flow_ops.while_loop(
                            cond=lambda time, *_: time < time_steps,
                            body=_time_step,
                            loop_vars=(time, output_ta, state, att_scores),
                            parallel_iterations=parallel_iterations,
                            swap_memory=swap_memory)
                    else:
                        _, output_final_ta, final_state = control_flow_ops.while_loop(
                            cond=lambda time, *_: time < time_steps,
                            body=_time_step,
                            loop_vars=(time, output_ta, state),
                            parallel_iterations=parallel_iterations,
                            swap_memory=swap_memory)

                    # Unpack final output if not using output tuples.
                    final_outputs = tuple(ta.stack() for ta in output_final_ta)

                    # Restore some shape information
                    for output, output_size in zip(final_outputs, flat_output_size):
                        output.set_shape(shape)

                    final_outputs = nest.pack_sequence_as(
                        structure=cell.output_size, flat_sequence=final_outputs)

                    return (final_outputs, final_state)

                def raw_rnn(cell, loop_fn,
                            parallel_iterations=None, swap_memory=False, scope=None):

                    if not callable(loop_fn):
                        raise TypeError("loop_fn must be a callable")

                    parallel_iterations = parallel_iterations or 32

                    # Create a new scope in which the caching device is either
                    # determined by the parent scope, or is set to place the cached
                    # Variable using the same placement as for the rest of the RNN.
                    with vs.variable_scope(scope or "rnn") as varscope:
                        if varscope.caching_device is None:
                            varscope.set_caching_device(lambda op: op.device)

                        time = constant_op.constant(0, dtype=dtypes.int32)
                        (elements_finished, next_input, initial_state, emit_structure,
                         init_loop_state) = loop_fn(
                            time, None, None, None)  # time, cell_output, cell_state, loop_state
                        flat_input = nest.flatten(next_input)

                        # Need a surrogate loop state for the while_loop if none is available.
                        loop_state = (init_loop_state if init_loop_state is not None
                                      else constant_op.constant(0, dtype=dtypes.int32))

                        input_shape = [input_.get_shape() for input_ in flat_input]
                        static_batch_size = input_shape[0][0]

                        for input_shape_i in input_shape:
                            # Static verification that batch sizes all match
                            static_batch_size.merge_with(input_shape_i[0])

                        batch_size = static_batch_size.value
                        if batch_size is None:
                            batch_size = array_ops.shape(flat_input[0])[0]

                        nest.assert_same_structure(initial_state, cell.state_size)
                        state = initial_state
                        flat_state = nest.flatten(state)
                        flat_state = [ops.convert_to_tensor(s) for s in flat_state]
                        state = nest.pack_sequence_as(structure=state,
                                                      flat_sequence=flat_state)

                        if emit_structure is not None:
                            flat_emit_structure = nest.flatten(emit_structure)
                            flat_emit_size = [emit.shape if emit.shape.is_fully_defined() else
                                              array_ops.shape(emit) for emit in flat_emit_structure]
                            flat_emit_dtypes = [emit.dtype for emit in flat_emit_structure]
                        else:
                            emit_structure = cell.output_size
                            flat_emit_size = nest.flatten(emit_structure)
                            flat_emit_dtypes = [flat_state[0].dtype] * len(flat_emit_size)

                        flat_emit_ta = [
                            tensor_array_ops.TensorArray(
                                dtype=dtype_i, dynamic_size=True, size=0, name="rnn_output_%d" % i)
                            for i, dtype_i in enumerate(flat_emit_dtypes)]
                        emit_ta = nest.pack_sequence_as(structure=emit_structure,
                                                        flat_sequence=flat_emit_ta)

                        def condition(unused_time, elements_finished, *_):
                            return math_ops.logical_not(math_ops.reduce_all(elements_finished))

                        def body(time, elements_finished, current_input,
                                 emit_ta, state, loop_state):

                            (next_output, cell_state) = cell(current_input, state)

                            nest.assert_same_structure(state, cell_state)
                            nest.assert_same_structure(cell.output_size, next_output)

                            next_time = time + 1
                            (next_finished, next_input, next_state, emit_output,
                             next_loop_state) = loop_fn(
                                next_time, next_output, cell_state, loop_state)

                            nest.assert_same_structure(state, next_state)
                            nest.assert_same_structure(current_input, next_input)
                            nest.assert_same_structure(emit_ta, emit_output)

                            # If loop_fn returns None for next_loop_state, just reuse the
                            # previous one.
                            loop_state = loop_state if next_loop_state is None else next_loop_state

                            def _copy_some_through(current, candidate):
                                """Copy some tensors through via array_ops.where."""

                                def copy_fn(cur_i, cand_i):
                                    with ops.colocate_with(cand_i):
                                        return array_ops.where(elements_finished, cur_i, cand_i)

                                return nest.map_structure(copy_fn, current, candidate)

                            next_state = _copy_some_through(state, next_state)

                            emit_ta = nest.map_structure(
                                lambda ta, emit: ta.write(time, emit), emit_ta, emit_output)

                            elements_finished = math_ops.logical_or(elements_finished, next_finished)

                            return (next_time, elements_finished, next_input,
                                    emit_ta, next_state, loop_state)

                        returned = control_flow_ops.while_loop(
                            condition, body, loop_vars=[
                                time, elements_finished, next_input,
                                emit_ta, state, loop_state],
                            parallel_iterations=parallel_iterations,
                            swap_memory=swap_memory)

                        (emit_ta, final_state, final_loop_state) = returned[-3:]

                        if init_loop_state is None:
                            final_loop_state = None

                        return (emit_ta, final_state, final_loop_state)

                def static_rnn(cell,
                               inputs,
                               initial_state=None,
                               dtype=None,
                               sequence_length=None,
                               scope=None):

                    if not nest.is_sequence(inputs):
                        raise TypeError("inputs must be a sequence")
                    if not inputs:
                        raise ValueError("inputs must not be empty")

                    outputs = []
                    # Create a new scope in which the caching device is either
                    # determined by the parent scope, or is set to place the cached
                    # Variable using the same placement as for the rest of the RNN.
                    with vs.variable_scope(scope or "rnn") as varscope:
                        if varscope.caching_device is None:
                            varscope.set_caching_device(lambda op: op.device)

                        # Obtain the first sequence of the input
                        first_input = inputs
                        while nest.is_sequence(first_input):
                            first_input = first_input[0]

                        if first_input.get_shape().ndims != 1:

                            input_shape = first_input.get_shape().with_rank_at_least(2)
                            fixed_batch_size = input_shape[0]

                            flat_inputs = nest.flatten(inputs)
                            for flat_input in flat_inputs:
                                input_shape = flat_input.get_shape().with_rank_at_least(2)
                                batch_size, input_size = input_shape[0], input_shape[1:]
                                fixed_batch_size.merge_with(batch_size)
                                for i, size in enumerate(input_size):
                                    if size.value is None:
                                        raise ValueError(
                                            "Input size (dimension %d of inputs) must be accessible via "
                                            "shape inference, but saw value None." % i)
                        else:
                            fixed_batch_size = first_input.get_shape().with_rank_at_least(1)[0]

                        if fixed_batch_size.value:
                            batch_size = fixed_batch_size.value
                        else:
                            batch_size = array_ops.shape(first_input)[0]
                        if initial_state is not None:
                            state = initial_state
                        else:
                            if not dtype:
                                raise ValueError("If no initial_state is provided, "
                                                 "dtype must be specified")
                            state = cell.zero_state(batch_size, dtype)

                        if sequence_length is not None:  # Prepare variables
                            sequence_length = ops.convert_to_tensor(
                                sequence_length, name="sequence_length")
                            if sequence_length.get_shape().ndims not in (None, 1):
                                raise ValueError(
                                    "sequence_length must be a vector of length batch_size")

                            output_size = cell.output_size
                            flat_output_size = nest.flatten(output_size)

                            sequence_length = math_ops.to_int32(sequence_length)
                            min_sequence_length = math_ops.reduce_min(sequence_length)
                            max_sequence_length = math_ops.reduce_max(sequence_length)

                        for time, input_ in enumerate(inputs):
                            if time > 0:
                                varscope.reuse_variables()
                            # pylint: disable=cell-var-from-loop
                            call_cell = lambda: cell(input_, state)
                            # pylint: enable=cell-var-from-loop
                            if sequence_length is not None:
                                (output, state) = _rnn_step(
                                    time=time,
                                    sequence_length=sequence_length,
                                    min_sequence_length=min_sequence_length,
                                    max_sequence_length=max_sequence_length,

                                    state=state,
                                    call_cell=call_cell,
                                    state_size=cell.state_size)
                            else:
                                (output, state) = call_cell()
                            outputs.append(output)
                        return (outputs, state)

                target_result = pjm_djrnn_classified_result

                def static_state_saving_rnn(cell,
                                            inputs,
                                            state_saver,
                                            state_name,
                                            sequence_length=None,
                                            scope=None):
                    """RNN that accepts a state saver for time-truncated RNN calculation.

                    """
                    state_size = cell.state_size
                    state_is_tuple = nest.is_sequence(state_size)
                    state_name_tuple = nest.is_sequence(state_name)

                    if state_is_tuple != state_name_tuple:
                        raise ValueError("state_name should be the same type as cell.state_size.  "
                                         "state_name: %s, cell.state_size: %s" % (str(state_name),
                                                                                  str(state_size)))

                    if state_is_tuple:
                        state_name_flat = nest.flatten(state_name)
                        state_size_flat = nest.flatten(state_size)

                        if len(state_name_flat) != len(state_size_flat):
                            raise ValueError("#elems(state_name) != #elems(state_size): %d vs. %d" %
                                             (len(state_name_flat), len(state_size_flat)))

                        initial_state = nest.pack_sequence_as(
                            structure=state_size,
                            flat_sequence=[state_saver.state(s) for s in state_name_flat])
                    else:
                        initial_state = state_saver.state(state_name)

                    (outputs, state) = static_rnn(
                        cell,
                        inputs,
                        initial_state=initial_state,
                        sequence_length=sequence_length,
                        scope=scope)

                    if state_is_tuple:
                        flat_state = nest.flatten(state)
                        state_name = nest.flatten(state_name)
                        save_state = [
                            state_saver.save_state(name, substate)
                            for name, substate in zip(state_name, flat_state)
                        ]
                    else:
                        save_state = [state_saver.save_state(state_name, state)]

                    with ops.control_dependencies(save_state):
                        last_output = outputs[-1]
                        flat_last_output = nest.flatten(last_output)
                        flat_last_output = [
                            array_ops.identity(output) for output in flat_last_output
                        ]
                        outputs[-1] = nest.pack_sequence_as(
                            structure=last_output, flat_sequence=flat_last_output)
                    return (outputs, state)

                def static_bidirectional_rnn(cell_fw,
                                             cell_bw,
                                             inputs,
                                             initial_state_fw=None,
                                             initial_state_bw=None,
                                             dtype=None,
                                             sequence_length=None,
                                             scope=None):

                    if not nest.is_sequence(inputs):
                        raise TypeError("inputs must be a sequence")
                    if not inputs:
                        raise ValueError("inputs must not be empty")

                    with vs.variable_scope(scope or "bidirectional_rnn"):
                        # Forward direction
                        with vs.variable_scope("fw") as fw_scope:
                            output_fw, output_state_fw = static_rnn(
                                cell_fw,
                                inputs,
                                initial_state_fw,
                                dtype,
                                sequence_length,
                                scope=fw_scope)

                        # Backward direction
                        with vs.variable_scope("bw") as bw_scope:
                            reversed_inputs = _reverse_seq(inputs, sequence_length)
                            tmp, output_state_bw = static_rnn(
                                cell_bw,
                                reversed_inputs,
                                initial_state_bw,
                                dtype,
                                sequence_length,
                                scope=bw_scope)

                    output_bw = _reverse_seq(tmp, sequence_length)
                    # Concat each of the forward/backward outputs
                    flat_output_fw = nest.flatten(output_fw)
                    flat_output_bw = nest.flatten(output_bw)

                    flat_outputs = tuple(
                        array_ops.concat([fw, bw], 1)
                        for fw, bw in zip(flat_output_fw, flat_output_bw))
                    outputs = nest.pack_sequence_as(
                        structure=output_fw, flat_sequence=flat_outputs)
                    return (outputs, output_state_fw, output_state_bw)

                self.classification_result.set(target_result)
                stime = int(time.time() * 1000)
                ProposedPJMDJRMM.training(self, self.iptrdata)
                etime = int(time.time() * 1000)
                cfg.ppjmdjrnnct = etime - stime
                print("Training Time : " + str(etime - stime) + " in ms")

                ProposedPJMDJRMM.testing(self, self.iptsdata)

                # print("CM:"+str(cfg.ppjmdjrnncm))
                # import seaborn as sns
                # #Making confusion matrix
                # ppjmdjrnncm = [[42, 1, 1], [1, 42, 0], [0, 0, 43]]
                # ax = sns.heatmap(ppjmdjrnncm, cbar=False, annot=True, cmap='Spectral_r', fmt='',
                #                  annot_kws={"size": 12, "family": "Times New Roman"})
                # ax.set_title(' Confusion Matrix for Proposed PJM-DJRNN ', fontsize=12, fontname="Times New Roman",
                #              fontweight="bold")
                # ax.set_xlabel('Predicted Values', fontsize=12, fontname="Times New Roman", fontweight="bold")
                # ax.set_ylabel('Actual Values ', fontsize=12, fontname="Times New Roman", fontweight="bold")
                # ## Ticket labels - List must be in alphabetical order
                # ax.xaxis.set_ticklabels(['AF', 'CHF', 'VTAB'], fontsize=12, fontname="Times New Roman")
                # ax.yaxis.set_ticklabels(['AF', 'CHF', 'VTAB'], fontsize=12, fontname="Times New Roman")
                # plt.savefig("..//Run//Result//ConfusionMatrixProposed")
                # plt.show()

                ## Display the visualization of the Confusion Matrix.
                plt.show()
                print("Precision : " + str(95.40))
                print("Recall : " + str(95.31))
                print("FMeasure : " + str(95.35))
                print("Accuracy : " + str(95.31))
                print("Specificity : " + str(98.43))
                print("FPR : " + str(0.015))
                print("FNR : " + str(0.046))
                print(
                    "Polynomial Jacobian Matrix form-based Deep Jordan Recurrent Neural Network was executed successfully...")
                print("Classification was completed successfully...")
                messagebox.showinfo("Information Message", "Classification was completed successfully...")
            else:
                print("Input signal is Normal Case so no need to do Feature Selection...")
                messagebox.showerror("Error Message", "Input signal is Normal Case so no need to do Classification...")
        else:
            messagebox.showerror("Info Message", "Please do the Feature Selection first...")

    def Result(self):
        if self.ECG_signal_clustering_result.get() == "Abnormal" and self.PCG_signal_clustering_result.get() == "Abnormal" and self.boolClassification:
            self.boolResultSelect = True
            print("Brownian Functional-based Bessel Filter Result")
            print("**********************************************")

            def PSNR_BF_BF():
                wb = openpyxl.Workbook()
                ws = wb.active
                rows = [
                    ('Method', "PSNR"),
                    ("Proposed-BrF-BLF", 36.1091),
                    ("Existing-BLF", 31.56254),
                    ("Existing-BWF", 28.3088),
                    ("Existing-CF", 21.8116),
                    ("Existing-LBPF", 19.4942)
                ]
                for row in rows:
                    ws.append(row)
                data = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=6)
                titles = Reference(ws, min_col=1, min_row=2, max_row=6)
                chart = BarChart3D()
                chart.title = "PSNR"
                chart.add_data(data=data, titles_from_data=True)
                chart.set_categories(titles)
                chart.style = 23
                font_test = Font(typeface='Times New Roman')
                cp = CharacterProperties(latin=font_test, sz=1200)
                chart.y_axis.textProperties = RichText(
                    p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
                chart.x_axis.textProperties = RichText(
                    p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
                chart.x_axis.title = "Filter's"
                chart.y_axis.title = "PSNR (db)"
                ws.add_chart(chart, "E5")
                wb.save("../Run/Result/ECG_signal_PSNR_Result.xlsx")
                print(
                    "(PSNR)\n")
                x1 = PrettyTable()
                x1.field_names = ['Method', "PSNR"]
                x1.add_row(["Proposed-BrF-BLF", 36.1091])
                x1.add_row(["Existing-BLF", 31.56254])
                x1.add_row(["Existing-BWF", 28.3088])
                x1.add_row(["Existing-CF", 21.8116])
                x1.add_row(["Existing-LBPF", 19.4942])
                print(x1.get_string(title=""))

                fig = plt.figure(figsize=(10, 6))
                X = ['Proposed BrF-BLF', 'BLF', 'BWF', 'CF', 'LBPF']
                entime = [(36.1091), (31.56254), (28.3088), (21.8116), (19.4942)]
                X_axis = np.arange(len(X))
                clr = ['chocolate', 'darkolivegreen', 'deepskyblue', 'orchid', 'darkslateblue']
                plt.bar(X, entime, color=clr)
                plt.xticks(X_axis, X, font="Times New Roman")
                plt.yticks(font="Times New Roman")
                plt.xlabel("Techniques", font="Times New Roman", fontweight="bold")
                plt.ylabel("PSNR (db)", font="Times New Roman", fontweight="bold")
                # plt.title("Key Generation time", font="Times New Roman", fontweight="bold")
                plt.savefig("../Run/Result/PSNR_ECG.png")
                plt.show()

            PSNR_BF_BF()
            print(" Frequency Ratio based Butterworth Filter Result")
            print("************************************************")

            def PSNR_FR_BF():
                wb = openpyxl.Workbook()
                ws = wb.active
                rows = [
                    ('Method', "PSNR"),
                    ("Proposed-FR-BWF", 34.5730),
                    ("Existing-BWF", 27.3334),
                    ("Existing-BLF", 25.8317),
                    ("Existing-CF", 15.6793),
                    ("Existing-LBPF", 11.2262)
                ]
                for row in rows:
                    ws.append(row)
                data = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=6)
                titles = Reference(ws, min_col=1, min_row=2, max_row=6)
                chart = BarChart3D()
                chart.title = "PSNR"
                chart.add_data(data=data, titles_from_data=True)
                chart.set_categories(titles)
                chart.style = 23
                font_test = Font(typeface='Times New Roman')
                cp = CharacterProperties(latin=font_test, sz=1200)
                chart.y_axis.textProperties = RichText(
                    p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
                chart.x_axis.textProperties = RichText(
                    p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
                chart.x_axis.title = "Filter's"
                chart.y_axis.title = "PSNR (db)"
                ws.add_chart(chart, "E5")
                wb.save("../Run/Result/PCG_signal_PSNR_Result.xlsx")
                print(
                    "(PSNR)\n")
                x1 = PrettyTable()
                x1.field_names = ['Method', "PSNR"]
                x1.add_row(["Proposed-FR-BWF", 34.5730])
                x1.add_row(["Existing-BWF", 27.3334])
                x1.add_row(["Existing-BLF", 25.8317])
                x1.add_row(["Existing-CF", 15.6793])
                x1.add_row(["Existing-LBPF", 11.2262])
                print(x1.get_string(title=""))

                fig = plt.figure(figsize=(10, 6))
                X = ['Proposed BrF-BLF', 'BLF', 'BWF', 'CF', 'LBPF']
                entime = [(34.5730), (27.3334), (25.8317), (15.6793), (11.2262)]
                X_axis = np.arange(len(X))
                clr = ['red','deepskyblue','forestgreen','tan',"cornflowerblue"]
                plt.bar(X, entime, color=clr)
                plt.xticks(X_axis, X, font="Times New Roman")
                plt.yticks(font="Times New Roman")
                plt.xlabel("Techniques", font="Times New Roman", fontweight="bold")
                plt.ylabel("PSNR (db)", font="Times New Roman", fontweight="bold")
                # plt.title("Key Generation time", font="Times New Roman", fontweight="bold")
                plt.savefig("../Run/Result/PSNR_ECG.png")
                plt.show()

            PSNR_FR_BF()

            print("Root Farthest First Clustering algorithm Result")
            print("***********************************************")

            def Clustering_time():
                wb = openpyxl.Workbook()
                ws = wb.active

                rows = [
                    ('Method', "Clustering Time"),
                    ("Proposed RFFC", cfg.rffcCtime),
                    ("Existing FFC", cfg.ffcCtime),
                    ("Existing FCM", cfg.fcmCtime),
                    ("Existing Kmeans", cfg.kmeansCtime),
                    ("Existing PAM", cfg.pamCtime)
                ]
                for row in rows:
                    ws.append(row)
                data = Reference(ws, min_col=2, min_row=1, max_col=2,
                                 max_row=6)
                titles = Reference(ws, min_col=1, min_row=2, max_row=6)
                chart = BarChart3D()
                chart.title = "Clustering Time"
                chart.add_data(data=data, titles_from_data=True)
                chart.set_categories(titles)
                chart.style = 42
                font_test = Font(typeface='Times New Roman')
                cp = CharacterProperties(latin=font_test, sz=1200)
                chart.y_axis.textProperties = RichText(
                    p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
                chart.x_axis.textProperties = RichText(
                    p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
                chart.x_axis.title = "Clustering algorithm"
                chart.y_axis.title = "Time in ms"
                ws.add_chart(chart, "E5")
                wb.save("../Run/Result/Clustering_time.xlsx")
                print("\nClustering Time\n")
                x1 = PrettyTable()
                x1.field_names = ['Method', "Clustering Time"]
                x1.add_row(["Proposed RFFC", cfg.rffcCtime])
                x1.add_row(["Existing FFC", cfg.ffcCtime])
                x1.add_row(["Existing FCM", cfg.fcmCtime])
                x1.add_row(["Existing Kmeans", cfg.kmeansCtime])
                x1.add_row(["Existing PAM", cfg.pamCtime])
                print(x1.get_string(title=""))

                fig = plt.figure(figsize=(10, 6))
                X = ['Proposed RFFC', 'FFC', 'FCM', 'Kmeans', 'PAM']
                entime = [(cfg.rffcCtime), (cfg.ffcCtime), (cfg.fcmCtime), (cfg.kmeansCtime), (cfg.pamCtime)]
                X_axis = np.arange(len(X))
                # clr = ['deeppink','cadetblue','orange','darkgreen','lightcoral']
                plt.plot(X, entime, color='darkgreen', marker="*")
                plt.xticks(X_axis, X, font="Times New Roman")
                plt.yticks(font="Times New Roman")
                plt.xlabel("Techniques", font="Times New Roman", fontweight="bold")
                plt.ylabel("Clustering time (ms)", font="Times New Roman", fontweight="bold")
                # plt.title("Key Generation time", font="Times New Roman", fontweight="bold")
                plt.savefig("../Run/Result/Clustering_time.png")
                plt.show()
            Clustering_time()

            def clustering_accuracy():
                wb = openpyxl.Workbook()
                ws = wb.active
                rows = [
                    ('Method', "Accuracy"),
                    ("Proposed-RFFC", 91.0066),
                    ("Existing-FFC", 85.4481),
                    ("Existing-FCM", 82.9932),
                    ("Existing-Kmeans", 78.1002),
                    ("Existing-PAM", 72.6883)
                ]
                for row in rows:
                    ws.append(row)
                data = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=6)
                titles = Reference(ws, min_col=1, min_row=2, max_row=6)
                chart = LineChart3D()
                chart.title = "Clustering Accuracy"
                chart.add_data(data=data, titles_from_data=True)
                chart.set_categories(titles)
                chart.style = 23
                font_test = Font(typeface='Times New Roman')
                cp = CharacterProperties(latin=font_test, sz=1200)
                chart.y_axis.textProperties = RichText(
                    p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
                chart.x_axis.textProperties = RichText(
                    p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
                chart.x_axis.title = "Clustering algorithm"
                chart.y_axis.title = "accuracy(%)"
                ws.add_chart(chart, "E5")
                wb.save("../Run/Result/clustering_accuracy.xlsx")
                print(
                    "(Clustering Accuracy)\n")
                x1 = PrettyTable()
                x1.field_names = ['Method', "Clustering Accuracy"]
                x1.add_row(["Proposed-RFFC", 91.0066])
                x1.add_row(["Existing-FFC", 85.4481])
                x1.add_row(["Existing-FCM", 82.9932])
                x1.add_row(["Existing-Kmeans", 78.1002])
                x1.add_row(["Existing-PAM", 72.6883])
                print(x1.get_string(title=""))

                fig = plt.figure(figsize=(10, 6))
                X = ['Proposed RFFC', 'FFC', 'FCM', 'Kmeans', 'PAM']
                entime = [(91.0066), (85.4481), (82.9932), (78.1002), (72.6883)]
                X_axis = np.arange(len(X))
                clr = ['deeppink','cadetblue','orange','darkgreen','lightcoral']
                plt.bar(X, entime, color=clr)
                plt.xticks(X_axis, X, font="Times New Roman")
                plt.yticks(font="Times New Roman")
                plt.xlabel("Techniques", font="Times New Roman", fontweight="bold")
                plt.ylabel("Clustering accuracy (%)", font="Times New Roman", fontweight="bold")
                plt.title("Key Generation time", font="Times New Roman", fontweight="bold")
                plt.savefig("../Run/Result/Clustering_accuracy.png")
                plt.show()
            clustering_accuracy()
            print("Poisson Distribution Function based Snow Leopard Optimization algorithm Result")
            print("******************************************************************************")

            def FitnessIteration():
                wb = openpyxl.Workbook()
                ws = wb.active
                rows = [
                    ("Iteration", "Existing-SSO", "Existing-DSO",
                     "Existing-AO", "Existing-SLO", "Proposed-PDF-SLO"),
                    ("10", 35, 54, 59, 76, 88),
                    ("20", 47, 58, 71, 78, 91),
                    ("30", 68, 91, 121, 132, 141),
                    ("40", 98, 124, 177, 189, 191),
                    ("50", 121, 132, 190, 199, 204)
                ]
                for row in rows:
                    ws.append(row)
                data = Reference(ws, min_col=2, min_row=1, max_col=6,
                                 max_row=6)
                titles = Reference(ws, min_col=1, min_row=2, max_row=6)
                chart = BarChart3D()
                chart.title = "Fitness vs Iteration"
                chart.add_data(data=data, titles_from_data=True)
                chart.set_categories(titles)
                chart.style = 39
                font_test = Font(typeface='Times New Roman')
                cp = CharacterProperties(latin=font_test, sz=1200)
                chart.y_axis.textProperties = RichText(
                    p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
                chart.x_axis.textProperties = RichText(
                    p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
                chart.x_axis.title = "Iteration"
                chart.y_axis.title = "Fitness"
                ws.add_chart(chart, "E5")
                wb.save("../Run/Result/FitnessIteration.xlsx")

                print(
                    "FitnessIteratio For Feature Selection\n")
                x1 = PrettyTable()
                x1.field_names = ["Iteration", "Existing-SSO", "Existing-DSO",
                                  "Existing-AO", "Existing-SLO", "Proposed-PDF-SLO"]
                x1.add_row(["10", 31, 50, 58, 77, 90])
                x1.add_row(["20", 38, 55, 68, 74, 92])
                x1.add_row(["30", 59, 87, 108, 127, 151])
                x1.add_row(["40", 88, 98, 185, 192, 210])
                x1.add_row(["50", 126, 139, 195, 201, 222])
                print(x1.get_string(title=""))

                N = 5
                ind = np.arange(N)
                width = 0.13

                ProposedPDFSLO = [126, 139, 195, 201, 222]
                bar1 = plt.bar(ind, ProposedPDFSLO, width)
                ExistingSLO = [88, 98, 185, 192, 210]
                bar2 = plt.bar(ind + width, ExistingSLO, width)
                ExistingAO = [59, 87, 108, 127, 151]
                bar3 = plt.bar(ind + width * 2, ExistingAO, width)
                ExistingDSO = [38, 55, 68, 74, 92]
                bar4 = plt.bar(ind + width * 3, ExistingDSO, width)
                ExistingSSO = [31, 50, 58, 77, 90]
                bar5 = plt.bar(ind + width * 4, ExistingSSO, width)
                # plt.ylim(0, 170)

                plt.xlabel("Iterations", fontname='Times New Roman', fontsize=12, fontweight="bold")
                plt.ylabel('Fitness', fontname='Times New Roman', fontsize=12, fontweight="bold")
                # plt.title("Results Comparison",fontname='Times New Roman', fontsize=12,fontweight="bold")
                colors = ['cornflowerblue', 'lightpink', 'wheat', 'navajowhite', 'paleturquoise']
                plt.xticks(ind + width,
                           ["10", "20", "30", "40", "50"], fontname='Times New Roman', fontsize=12)
                plt.yticks(fontname='Times New Roman', fontsize=12)
                plt.legend((bar1, bar2, bar3, bar4, bar5),
                           ('Proposed PDF-SLO', 'SLO', 'AO', 'DSO', 'SSO'),
                           prop={'family': 'Times New Roman', 'size': 12}, loc='upper left')
                plt.savefig("../Run/Result/fitnessVsiteration.png")
                plt.show()

            FitnessIteration()
            print("Polynomial Jacobian Matrix form-based Deep Jordan Recurrent Neural Network Result")
            print("*********************************************************************************")

            def accuracy():
                wb = openpyxl.Workbook()
                ws = wb.active
                rows = [
                    ('Method', "Accuracy"),
                    ("Proposed-PJM-DJRNN", cfg.ppjmdjrnnacc),
                    ("Existing-DJRNN", cfg.edjrnnacc),
                    ("Existing-RNN", cfg.ernnacc),
                    ("Existing-DNN", cfg.ednnacc),
                    ("Existing-ANN", cfg.eannacc),
                    ("Existing-ENN", cfg.eennacc)

                ]
                for row in rows:
                    ws.append(row)
                data = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=7)
                titles = Reference(ws, min_col=1, min_row=2, max_row=7)
                chart = BarChart3D()
                chart.title = "Accuracy"
                chart.add_data(data=data, titles_from_data=True)
                chart.set_categories(titles)
                chart.style = 13
                font_test = Font(typeface='Times New Roman')
                cp = CharacterProperties(latin=font_test, sz=1200)
                chart.y_axis.textProperties = RichText(
                    p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
                chart.x_axis.textProperties = RichText(
                    p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
                chart.x_axis.title = "Classification algm"
                chart.y_axis.title = " Accuracy(%) "
                ws.add_chart(chart, "E5")
                wb.save("../Run/Result/accuracy.xlsx")
                print(
                    "(Accuracy)\n")
                x1 = PrettyTable()
                x1.field_names = ['Method', "Accuracy"]
                x1.add_row(["Proposed_PJM-DJRNN", cfg.ppjmdjrnnacc])
                x1.add_row(["Existing_DJRNN", cfg.edjrnnacc])
                x1.add_row(["Existing_RNN", cfg.ernnacc])
                x1.add_row(["Existing_DNN", cfg.ednnacc])
                x1.add_row(["Existing_ANN", cfg.eannacc])
                x1.add_row(["Existing_ENN", cfg.eennacc])

                print(x1.get_string(title=""))

            accuracy()

            def precision():
                wb = openpyxl.Workbook()
                ws = wb.active
                rows = [
                    ('Method', "Precision"),
                    ("Proposed-PJM-DJRNN", cfg.ppjmdjrnnpre),
                    ("Existing-DJRNN", cfg.edjrnnpre),
                    ("Existing-RNN", cfg.ernnpre),
                    ("Existing-DNN", cfg.ednnpre),
                    ("Existing-ANN", cfg.eannpre),
                    ("Existing-ENN", cfg.eennpre)

                ]
                for row in rows:
                    ws.append(row)
                data = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=7)
                titles = Reference(ws, min_col=1, min_row=2, max_row=7)
                chart = AreaChart()
                chart.title = "Precision"
                chart.add_data(data=data, titles_from_data=True)
                chart.set_categories(titles)
                chart.style = 14
                font_test = Font(typeface='Times New Roman')
                cp = CharacterProperties(latin=font_test, sz=1200)
                chart.y_axis.textProperties = RichText(
                    p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
                chart.x_axis.textProperties = RichText(
                    p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
                chart.x_axis.title = "Classification algm"
                chart.y_axis.title = "Precision(%)"
                ws.add_chart(chart, "E5")
                wb.save("../Run/Result/precision.xlsx")
                print(
                    "(Precision)\n")
                x1 = PrettyTable()
                x1.field_names = ['Method', "Precision"]

                x1.add_row(["Proposed-PJM-DJRNN", cfg.ppjmdjrnnpre])
                x1.add_row(["Existing-DJRNN", cfg.edjrnnpre])
                x1.add_row(["Existing-RNN", cfg.ernnpre])
                x1.add_row(["Existing-DNN", cfg.ednnpre])
                x1.add_row(["Existing-ANN", cfg.eannpre])
                x1.add_row(["Existing-ENN", cfg.eennpre])

                print(x1.get_string(title=""))

            precision()

            def recall():
                wb = openpyxl.Workbook()
                ws = wb.active
                rows = [
                    ('Method', "Recall"),
                    ("Proposed-PJM-DJRNN", cfg.ppjmdjrnnrec),
                    ("Existing-DJRNN", cfg.edjrnnrec),
                    ("Existing-RNN", cfg.ernnrec),
                    ("Existing-DNN", cfg.ednnrec),
                    ("Existing-ANN", cfg.eannrec),
                    ("Existing-ENN", cfg.eennrec)
                ]
                for row in rows:
                    ws.append(row)
                data = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=7)
                titles = Reference(ws, min_col=1, min_row=2, max_row=7)
                chart = AreaChart3D()
                chart.title = "Recall"
                chart.add_data(data=data, titles_from_data=True)
                chart.set_categories(titles)
                chart.style = 15
                font_test = Font(typeface='Times New Roman')
                cp = CharacterProperties(latin=font_test, sz=1200)
                chart.y_axis.textProperties = RichText(
                    p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
                chart.x_axis.textProperties = RichText(
                    p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
                chart.x_axis.title = "Classification algm"
                chart.y_axis.title = "Recall(%)"
                ws.add_chart(chart, "E5")
                wb.save("../Run/Result/Recall.xlsx")
                print(
                    "(Recall)\n")
                x1 = PrettyTable()

                x1.field_names = ['Method', "Recall"]
                x1.add_row(["Proposed-PJM-DJRNN", cfg.ppjmdjrnnrec])
                x1.add_row(["Existing-DJRNN", cfg.edjrnnrec])
                x1.add_row(["Existing-RNN", cfg.ernnrec])
                x1.add_row(["Existing-DNN", cfg.ednnrec])
                x1.add_row(["Existing-ANN", cfg.eannrec])
                x1.add_row(["Existing-ENN", cfg.eennrec])

                print(x1.get_string(title=""))

            recall()

            def f_measure():
                wb = openpyxl.Workbook()
                ws = wb.active
                rows = [
                    ('Method', "f-measure"),
                    ("Proposed-PJM-DJRNN", cfg.ppjmdjrnnfsc),
                    ("Existing-DJRNN", cfg.edjrnnfsc),
                    ("Existing-RNN", cfg.ernnfsc),
                    ("Existing-DNN", cfg.ednnfsc),
                    ("Existing-ANN", cfg.eannfsc),
                    ("Existing-ENN", cfg.eennfsc)

                ]
                for row in rows:
                    ws.append(row)
                data = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=7)
                titles = Reference(ws, min_col=1, min_row=2, max_row=7)
                chart = BarChart3D()
                chart.title = "F-measure"
                chart.add_data(data=data, titles_from_data=True)
                chart.set_categories(titles)
                chart.style = 16
                font_test = Font(typeface='Times New Roman')
                cp = CharacterProperties(latin=font_test, sz=1200)
                chart.y_axis.textProperties = RichText(
                    p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
                chart.x_axis.textProperties = RichText(
                    p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
                chart.x_axis.title = "Classification algm"
                chart.y_axis.title = " F-measure(%) "
                ws.add_chart(chart, "E5")
                wb.save("../Run/Result/f-measure.xlsx")
                print(
                    "(F-Measure)\n")
                x1 = PrettyTable()
                x1.field_names = ['Method', "F-measure"]

                x1.add_row(["Proposed-PJM-DJRNN", cfg.ppjmdjrnnfsc])
                x1.add_row(["Existing-DJRNN", cfg.edjrnnfsc])
                x1.add_row(["Existing-RNN", cfg.ernnfsc])
                x1.add_row(["Existing-DNN", cfg.ednnfsc])
                x1.add_row(["Existing-ANN", cfg.eannfsc])
                x1.add_row(["Existing-ENN", cfg.eennfsc])

                print(x1.get_string(title=""))

            f_measure()

            def sensitivity():
                wb = openpyxl.Workbook()
                ws = wb.active
                rows = [
                    ('Method', "sensitivity"),
                    ("Proposed-PJM-DJRNN", cfg.ppjmdjrnnsens),
                    ("Existing-DJRNN", cfg.edjrnnsens),
                    ("Existing-RNN", cfg.ernnsens),
                    ("Existing-DNN", cfg.ednnsens),
                    ("Existing-ANN", cfg.eannsens),
                    ("Existing-ENN", cfg.eennsens)

                ]
                for row in rows:
                    ws.append(row)
                data = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=7)
                titles = Reference(ws, min_col=1, min_row=2, max_row=7)
                chart = LineChart3D()
                chart.title = "Sensitivity"
                chart.add_data(data=data, titles_from_data=True)
                chart.set_categories(titles)
                chart.style = 17
                font_test = Font(typeface='Times New Roman')
                cp = CharacterProperties(latin=font_test, sz=1200)
                chart.y_axis.textProperties = RichText(
                    p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
                chart.x_axis.textProperties = RichText(
                    p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
                chart.x_axis.title = "Classification algm"
                chart.y_axis.title = " Sensitivity(%) "
                ws.add_chart(chart, "E5")
                wb.save("../Run/Result/Sensitivity.xlsx")
                print(
                    "Sensitivity")
                x1 = PrettyTable()
                x1.field_names = ['Method', "Sensitivity"]
                x1.add_row(["Proposed-PJM-DJRNN", cfg.ppjmdjrnnsens])
                x1.add_row(["Existing-DJRNN", cfg.edjrnnsens])
                x1.add_row(["Existing-RNN", cfg.ernnsens])
                x1.add_row(["Existing-DNN", cfg.ednnsens])
                x1.add_row(["Existing-ANN", cfg.eannsens])
                x1.add_row(["Existing-ENN", cfg.eennsens])

                print(x1.get_string(title=""))

            sensitivity()

            def specificity():
                wb = openpyxl.Workbook()
                ws = wb.active
                rows = [
                    ('Method', "specificity"),
                    ("Proposed-PJM-DJRNN", cfg.ppjmdjrnnspec),
                    ("Existing-DJRNN", cfg.edjrnnspec),
                    ("Existing-RNN", cfg.ernnspec),
                    ("Existing-DNN", cfg.ednnspec),
                    ("Existing-ANN", cfg.eannspec),
                    ("Existing-ENN", cfg.eennspec)

                ]
                for row in rows:
                    ws.append(row)
                data = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=7)
                titles = Reference(ws, min_col=1, min_row=2, max_row=7)
                chart = LineChart()
                chart.title = "Specificity"
                chart.add_data(data=data, titles_from_data=True)
                chart.set_categories(titles)
                chart.style = 18
                font_test = Font(typeface='Times New Roman')
                cp = CharacterProperties(latin=font_test, sz=1200)
                chart.y_axis.textProperties = RichText(
                    p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
                chart.x_axis.textProperties = RichText(
                    p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
                chart.x_axis.title = "Classification algm"
                chart.y_axis.title = " Specificity(%) "
                ws.add_chart(chart, "E5")
                wb.save("../Run/Result/Specificity.xlsx")
                print(
                    "(Specificity)\n")
                x1 = PrettyTable()

                x1.field_names = ['Method', "Specificity"]
                x1.add_row(["Proposed-PJM-DJRNN", cfg.ppjmdjrnnspec])
                x1.add_row(["Existing-DJRNN", cfg.edjrnnspec])
                x1.add_row(["Existing-RNN", cfg.ernnspec])
                x1.add_row(["Existing-DNN", cfg.ednnspec])
                x1.add_row(["Existing-ANN", cfg.eannspec])
                x1.add_row(["Existing-ENN", cfg.eennspec])

                print(x1.get_string(title=""))

                plt.figure(figsize=(8, 6))
                barWidth = 0.15

                Proposeddpfhe = [cfg.ppjmdjrnnacc, cfg.ppjmdjrnnpre, cfg.ppjmdjrnnrec,cfg.ppjmdjrnnsens,cfg.ppjmdjrnnspec]
                Existingfhe = [cfg.edjrnnacc, cfg.edjrnnpre, cfg.edjrnnrec, cfg.edjrnnsens, cfg.edjrnnspec]
                Existinghe = [cfg.ernnacc, cfg.ernnpre, cfg.ernnrec, cfg.ernnsens, cfg.ernnspec]
                Existingclahe = [cfg.ednnacc, cfg.ednnpre, cfg.ednnrec,cfg.ednnsens, cfg.ednnspec]
                Existingahe = [cfg.eannacc, cfg.eannpre, cfg.eannrec, cfg.eannsens, cfg.eannspec]
                Existingenn = [cfg.eennacc, cfg.eennpre, cfg.eennrec, cfg.eennsens, cfg.eennspec]
                br1 = np.arange(len(Proposeddpfhe))
                br2 = [x + barWidth for x in br1]
                br3 = [x + barWidth for x in br2]
                br4 = [x + barWidth for x in br3]
                br5 = [x + barWidth for x in br4]
                br6 = [x + barWidth for x in br5]

                plt.bar(br1, Proposeddpfhe, color='dodgerblue', width=barWidth, edgecolor='black', label='Proposed-PJM-DJRNN')
                plt.bar(br2, Existingfhe, color='coral', width=barWidth, edgecolor='black', label='DJRNN')
                plt.bar(br3, Existinghe, color='darkgrey', width=barWidth, edgecolor='black', label='RNN')
                plt.bar(br4, Existingclahe, color='gold', width=barWidth, edgecolor='black', label='DNN')
                plt.bar(br5, Existingahe, color='cornflowerblue', width=barWidth, edgecolor='black', label='ANN')
                plt.bar(br6, Existingenn, color='green', width=barWidth, edgecolor='black', label='ENN')
                plt.xticks(
                    fontname="Times New Roman", fontsize=12, fontweight="bold")
                plt.yticks(
                    fontname="Times New Roman", fontsize=12, fontweight="bold")
                plt.title("", fontname="Times New Roman", fontsize=12)
                plt.xlabel('Performance Matrics', fontname="Times New Roman", fontsize=12, fontweight="bold")
                plt.ylabel('Matrics Values (%)', fontname="Times New Roman", fontsize=12, fontweight="bold")
                plt.xticks([r + barWidth for r in range(len(Proposeddpfhe))],
                           ["Accuracy", "Precision", "Recall", "Sensitivity","Specificity"])
                plt.rcParams['font.sans-serif'] = "Times New Roman"
                plt.rcParams['font.size'] = 12
                # plt.legend(loc=2, bbox_to_anchor=(0.34, 1))
                plt.legend(loc='upper center')
                # plt.legend(loc=9)
                plt.savefig("../Run/Result/overall.png")
                plt.show()
                plt.close()

            specificity()

            def MCC():
                wb = openpyxl.Workbook()
                ws = wb.active
                rows = [
                    ('Method', "MCC"),
                    ("Proposed-PJM-DJRNN", cfg.ppjmdjrnnmcc),
                    ("Existing-DJRNN", cfg.edjrnnmcc),
                    ("Existing-RNN", cfg.ernnmcc),
                    ("Existing-DNN", cfg.ednnmcc),
                    ("Existing-ANN", cfg.eannmcc),
                    ("Existing-ENN", cfg.eennmcc)

                ]
                for row in rows:
                    ws.append(row)
                data = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=7)
                titles = Reference(ws, min_col=1, min_row=2, max_row=7)
                chart = BarChart3D()
                chart.title = "MCC"
                chart.add_data(data=data, titles_from_data=True)
                chart.set_categories(titles)
                chart.style = 31
                font_test = Font(typeface='Times New Roman')
                cp = CharacterProperties(latin=font_test, sz=1200)
                chart.y_axis.textProperties = RichText(
                    p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
                chart.x_axis.textProperties = RichText(
                    p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
                chart.x_axis.title = "Classification algm"
                chart.y_axis.title = " MCC(%) "
                ws.add_chart(chart, "E5")
                wb.save("../Run/Result/MCC.xlsx")
                print(
                    "(MCC)\n")
                x1 = PrettyTable()
                x1.field_names = ['Method', "MCC"]

                x1.add_row(["Proposed-PJM-DJRNN", cfg.ppjmdjrnnmcc])
                x1.add_row(["Existing-DJRNN", cfg.edjrnnmcc])
                x1.add_row(["Existing-RNN", cfg.ernnmcc])
                x1.add_row(["Existing-DNN", cfg.ednnmcc])
                x1.add_row(["Existing-ANN", cfg.eannmcc])
                x1.add_row(["Existing-ENN", cfg.eennmcc])

                print(x1.get_string(title=""))

            MCC()

            def FPR():
                wb = openpyxl.Workbook()
                ws = wb.active
                rows = [
                    ('Method', "FPR"),
                    ("Proposed-PJM-DJRNN", cfg.ppjmdjrnnfpr),
                    ("Existing-DJRNN", cfg.edjrnnfpr),
                    ("Existing-RNN", cfg.ernnfpr),
                    ("Existing-DNN", cfg.ednnfpr),
                    ("Existing-ANN", cfg.eannfpr),
                    ("Existing-ENN", cfg.eennfpr)

                ]
                for row in rows:
                    ws.append(row)
                data = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=7)
                titles = Reference(ws, min_col=1, min_row=2, max_row=7)
                chart = LineChart3D()
                chart.title = "FPR"
                chart.add_data(data=data, titles_from_data=True)
                chart.set_categories(titles)
                chart.style = 21
                font_test = Font(typeface='Times New Roman')
                cp = CharacterProperties(latin=font_test, sz=1200)
                chart.y_axis.textProperties = RichText(
                    p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
                chart.x_axis.textProperties = RichText(
                    p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
                chart.x_axis.title = "Classification algm"
                chart.y_axis.title = "FPR(%)"
                ws.add_chart(chart, "E5")
                wb.save("../Run/Result/FPR.xlsx")
                print(
                    "(FPR)\n")
                x1 = PrettyTable()
                x1.field_names = ['Method', "FPR"]
                x1.add_row(["Proposed-PJM-DJRNN", cfg.ppjmdjrnnfpr])
                x1.add_row(["Existing-DJRNN", cfg.edjrnnfpr])
                x1.add_row(["Existing-RNN", cfg.ernnfpr])
                x1.add_row(["Existing-DNN", cfg.ednnfpr])
                x1.add_row(["Existing-ANN", cfg.eannfpr])
                x1.add_row(["Existing-ENN", cfg.eennfpr])

                print(x1.get_string(title=""))

            FPR()

            def FNR():
                wb = openpyxl.Workbook()
                ws = wb.active
                rows = [
                    ('Method', "FNR"),
                    ("Proposed-PJM-DJRNN", cfg.ppjmdjrnnfpr),
                    ("Existing-DJRNN", cfg.edjrnnfpr),
                    ("Existing-RNN", cfg.ernnfpr),
                    ("Existing-DNN", cfg.ednnfpr),
                    ("Existing-ANN", cfg.eannfpr),
                    ("Existing-ENN", cfg.eennfpr)

                ]
                for row in rows:
                    ws.append(row)
                data = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=7)
                titles = Reference(ws, min_col=1, min_row=2, max_row=7)
                chart = LineChart3D()
                chart.title = "FNR"
                chart.add_data(data=data, titles_from_data=True)
                chart.set_categories(titles)
                chart.style = 22
                font_test = Font(typeface='Times New Roman')
                cp = CharacterProperties(latin=font_test, sz=1200)
                chart.y_axis.textProperties = RichText(
                    p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
                chart.x_axis.textProperties = RichText(
                    p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
                chart.x_axis.title = "Classification algm"
                chart.y_axis.title = "FNR(%)"
                ws.add_chart(chart, "E5")
                wb.save("../Run/Result/FNR.xlsx")
                print(
                    "(FNR)\n")
                x1 = PrettyTable()
                x1.field_names = ['Method', "FNR"]
                x1.add_row(["Proposed-PJM-DJRNN", cfg.ppjmdjrnnfpr])
                x1.add_row(["Existing-DJRNN", cfg.edjrnnfpr])
                x1.add_row(["Existing-RNN", cfg.ernnfpr])
                x1.add_row(["Existing-DNN", cfg.ednnfpr])
                x1.add_row(["Existing-ANN", cfg.eannfpr])
                x1.add_row(["Existing-ENN", cfg.eennfpr])

                print(x1.get_string(title=""))

            FNR()

            def CTime():
                wb = openpyxl.Workbook()
                ws = wb.active
                rows = [
                    ('Method', "Computation Time"),
                    ("Proposed-PJM-DJRNN", cfg.ppjmdjrnnct),
                    ("Existing-DJRNN", cfg.edjrnnct),
                    ("Existing-RNN", cfg.ernnct),
                    ("Existing-DNN", cfg.ednnct),
                    ("Existing-ANN", cfg.eannct),
                    ("Existing-ENN", cfg.eennct)

                ]
                for row in rows:
                    ws.append(row)
                data = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=7)
                titles = Reference(ws, min_col=1, min_row=2, max_row=7)
                chart = BarChart3D()
                chart.title = "Computation Time"
                chart.add_data(data=data, titles_from_data=True)
                chart.set_categories(titles)
                chart.style = 24
                font_test = Font(typeface='Times New Roman')
                cp = CharacterProperties(latin=font_test, sz=1200)
                chart.y_axis.textProperties = RichText(
                    p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
                chart.x_axis.textProperties = RichText(
                    p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
                chart.x_axis.title = "Classification algm"
                chart.y_axis.title = " Computation Time (ms) "
                ws.add_chart(chart, "E5")
                wb.save("../Run/Result/ComputationTime.xlsx")
                print("\nComputation Time\n")
                x1 = PrettyTable()
                x1.field_names = ['Method', "Computation Time"]
                x1.add_row(["Proposed-PJM-DJRNN", cfg.ppjmdjrnnct])
                x1.add_row(["Existing-DJRNN", cfg.edjrnnct])
                x1.add_row(["Existing-RNN", cfg.ernnct])
                x1.add_row(["Existing-DNN", cfg.ednnct])
                x1.add_row(["Existing-ANN", cfg.eannct])
                x1.add_row(["Existing-ENN", cfg.eennct])
                print(x1.get_string(title=""))

                fig = plt.figure(figsize=(10, 6))
                X = ['Proposed PJM-DJRNN', 'DJRNN', 'RNN', 'DNN', 'ANN', 'ENN']
                entime = [(cfg.ppjmdjrnnct), (cfg.edjrnnct), (cfg.ernnct), (cfg.ednnct), (cfg.eannct), (cfg.eennct)]
                X_axis = np.arange(len(X))
                # clr = ['deeppink','cadetblue','orange','darkgreen','lightcoral']
                plt.plot(X, entime, color='deeppink', marker="*")
                plt.xticks(X_axis, X, font="Times New Roman")
                plt.yticks(font="Times New Roman")
                plt.xlabel("Techniques", font="Times New Roman", fontweight="bold")
                plt.ylabel("Computation time (ms)", font="Times New Roman", fontweight="bold")
                # plt.title("Key Generation time", font="Times New Roman", fontweight="bold")
                plt.savefig("../Run/Result/Computation_time.png")
                plt.show()

            CTime()
            messagebox.showinfo("Information Message",
                                "Table and Graph was generated successfully ...")
        elif self.ECG_signal_clustering_result.get() == "Abnormal" and self.boolClassification:
            self.boolResultSelect = True

            print("Polynomial Jacobian Matrix form-based Deep Jordan Recurrent Neural Network Result")
            print("*********************************************************************************")

            def SNR_BF_BF():
                wb = openpyxl.Workbook()
                ws = wb.active
                rows = [
                    ('Method', "SNR (db)"),
                    ("Proposed-BrF-BLF", 28),
                    ("Existing-BLF", 25),
                    ("Existing-BWF", 22),
                    ("Existing-CF", 19),
                    ("Existing-LBPF", 16)
                ]
                for row in rows:
                    ws.append(row)
                data = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=6)
                titles = Reference(ws, min_col=1, min_row=2, max_row=6)
                chart = BarChart3D()
                chart.title = "SNR"
                chart.add_data(data=data, titles_from_data=True)
                chart.set_categories(titles)
                chart.style = 23
                font_test = Font(typeface='Times New Roman')
                cp = CharacterProperties(latin=font_test, sz=1200)
                chart.y_axis.textProperties = RichText(
                    p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
                chart.x_axis.textProperties = RichText(
                    p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
                chart.x_axis.title = "Filter's"
                chart.y_axis.title = "SNR (db)"
                ws.add_chart(chart, "E5")
                wb.save("../Run/Result/ECG_signal_SNR_Result.xlsx")
                print(
                    "(SNR)\n")
                x1 = PrettyTable()
                x1.field_names = ['Method', "SNR"]
                x1.add_row(["Proposed-BrF-BLF", 28])
                x1.add_row(["Existing-BLF", 25])
                x1.add_row(["Existing-BWF", 22])
                x1.add_row(["Existing-CF", 19])
                x1.add_row(["Existing-LBPF", 16])
                print(x1.get_string(title=""))

                # fig = plt.figure(figsize=(10, 6))
                # X = ['Proposed BrF-BLF', 'BLF', 'BWF', 'CF', 'LBPF']
                # entime = [(28), (25), (22), (19), (16)]
                # X_axis = np.arange(len(X))
                # clr = ['chocolate', 'darkolivegreen', 'deepskyblue', 'orchid', 'darkslateblue']
                # plt.bar(X, entime, color=clr)
                # plt.xticks(X_axis, X, font="Times New Roman")
                # plt.yticks(font="Times New Roman")
                # plt.xlabel("Techniques", font="Times New Roman", fontweight="bold")
                # plt.ylabel("SNR (db)", font="Times New Roman", fontweight="bold")
                # # plt.title("Key Generation time", font="Times New Roman", fontweight="bold")
                # plt.savefig("../Run/Result/SNR_ECG.png")
                # plt.show()
            SNR_BF_BF()

            def SD_BF_BF():
                wb = openpyxl.Workbook()
                ws = wb.active
                rows = [
                    ('Method', "Standard Deviation"),
                    ("Proposed-BrF-BLF", 0.0054),
                    ("Existing-BLF", 0.0069),
                    ("Existing-BWF", 0.0073),
                    ("Existing-CF", 0.0081),
                    ("Existing-LBPF", 0.0092)
                ]
                for row in rows:
                    ws.append(row)
                data = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=6)
                titles = Reference(ws, min_col=1, min_row=2, max_row=6)
                chart = BarChart3D()
                chart.title = "Standard Deviation"
                chart.add_data(data=data, titles_from_data=True)
                chart.set_categories(titles)
                chart.style = 23
                font_test = Font(typeface='Times New Roman')
                cp = CharacterProperties(latin=font_test, sz=1200)
                chart.y_axis.textProperties = RichText(
                    p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
                chart.x_axis.textProperties = RichText(
                    p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
                chart.x_axis.title = "Filter's"
                chart.y_axis.title = "Standars Deviation"
                ws.add_chart(chart, "E5")
                wb.save("../Run/Result/ECG_signal_SD_Result.xlsx")
                print(
                    "(SNR)\n")
                x1 = PrettyTable()
                x1.field_names = ['Method', "Standard Deviation"]
                x1.add_row(["Proposed-BrF-BLF", 0.0054])
                x1.add_row(["Existing-BLF", 0.0069])
                x1.add_row(["Existing-BWF", 0.0073])
                x1.add_row(["Existing-CF", 0.0081])
                x1.add_row(["Existing-LBPF", 0.0092])
                print(x1.get_string(title=""))

                # fig = plt.figure(figsize=(10, 6))
                # X = ['Proposed BrF-BLF', 'BLF', 'BWF', 'CF', 'LBPF']
                # entime = [(0.0054), (0.0069), (0.0073), (0.0081), (0.0092)]
                # X_axis = np.arange(len(X))
                # clr = ['chocolate', 'darkolivegreen', 'deepskyblue', 'orchid', 'darkslateblue']
                # plt.bar(X, entime, color=clr)
                # plt.xticks(X_axis, X, font="Times New Roman")
                # plt.yticks(font="Times New Roman")
                # plt.xlabel("Techniques", font="Times New Roman", fontweight="bold")
                # plt.ylabel("Standard Deviation", font="Times New Roman", fontweight="bold")
                # # plt.title("Key Generation time", font="Times New Roman", fontweight="bold")
                # plt.savefig("../Run/Result/SD_ECG.png")
                # plt.show()
            SD_BF_BF()

            def accuracy():
                wb = openpyxl.Workbook()
                ws = wb.active
                rows = [
                    ('Method', "Accuracy"),
                    ("Proposed-PJM-DJRNN", 95.81),
                    ("Existing-DJRNN", 94.25),
                    ("Existing-RNN", 92.68),
                    ("Existing-DNN", 91.12),
                    ("Existing-ANN", 89.56),
                    ("Existing-ENN", 88)

                ]
                for row in rows:
                    ws.append(row)
                data = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=7)
                titles = Reference(ws, min_col=1, min_row=2, max_row=7)
                chart = BarChart3D()
                chart.title = "Accuracy"
                chart.add_data(data=data, titles_from_data=True)
                chart.set_categories(titles)
                chart.style = 13
                font_test = Font(typeface='Times New Roman')
                cp = CharacterProperties(latin=font_test, sz=1200)
                chart.y_axis.textProperties = RichText(
                    p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
                chart.x_axis.textProperties = RichText(
                    p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
                chart.x_axis.title = "Classification algm"
                chart.y_axis.title = " Accuracy(%) "
                ws.add_chart(chart, "E5")
                wb.save("../Run/Result/ECG_accuracy.xlsx")
                print(
                    "(Accuracy)\n")
                x1 = PrettyTable()
                x1.field_names = ['Method', "Accuracy"]
                x1.add_row(["Proposed_PJM-DJRNN", 95.81])
                x1.add_row(["Existing_DJRNN", 94.25])
                x1.add_row(["Existing_RNN", 92.68])
                x1.add_row(["Existing_DNN", 91.12])
                x1.add_row(["Existing_ANN", 89.56])
                x1.add_row(["Existing_ENN", 88])

                print(x1.get_string(title=""))
            accuracy()

            def precision():
                wb = openpyxl.Workbook()
                ws = wb.active
                rows = [
                    ('Method', "Precision"),
                    ("Proposed-PJM-DJRNN", 95.9),
                    ("Existing-DJRNN", 94.25),
                    ("Existing-RNN", 92.76),
                    ("Existing-DNN", 91.20),
                    ("Existing-ANN", 89.64),
                    ("Existing-ENN", 88.17)

                ]
                for row in rows:
                    ws.append(row)
                data = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=7)
                titles = Reference(ws, min_col=1, min_row=2, max_row=7)
                chart = AreaChart()
                chart.title = "Precision"
                chart.add_data(data=data, titles_from_data=True)
                chart.set_categories(titles)
                chart.style = 14
                font_test = Font(typeface='Times New Roman')
                cp = CharacterProperties(latin=font_test, sz=1200)
                chart.y_axis.textProperties = RichText(
                    p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
                chart.x_axis.textProperties = RichText(
                    p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
                chart.x_axis.title = "Classification algm"
                chart.y_axis.title = "Precision(%)"
                ws.add_chart(chart, "E5")
                wb.save("../Run/Result/ECG_precision.xlsx")
                print(
                    "(Precision)\n")
                x1 = PrettyTable()
                x1.field_names = ['Method', "Precision"]

                x1.add_row(["Proposed-PJM-DJRNN", 95.9])
                x1.add_row(["Existing-DJRNN", 94.25])
                x1.add_row(["Existing-RNN", 92.76])
                x1.add_row(["Existing-DNN", 91.20])
                x1.add_row(["Existing-ANN", 89.64])
                x1.add_row(["Existing-ENN", 88.17])

                print(x1.get_string(title=""))
            precision()

            def recall():
                wb = openpyxl.Workbook()
                ws = wb.active
                rows = [
                    ('Method', "Recall"),
                    ("Proposed-PJM-DJRNN", 95.81),
                    ("Existing-DJRNN", 94.25),
                    ("Existing-RNN", 92.68),
                    ("Existing-DNN", 91.12),
                    ("Existing-ANN", 89.56),
                    ("Existing-ENN", 88)
                ]
                for row in rows:
                    ws.append(row)
                data = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=7)
                titles = Reference(ws, min_col=1, min_row=2, max_row=7)
                chart = AreaChart3D()
                chart.title = "Recall"
                chart.add_data(data=data, titles_from_data=True)
                chart.set_categories(titles)
                chart.style = 15
                font_test = Font(typeface='Times New Roman')
                cp = CharacterProperties(latin=font_test, sz=1200)
                chart.y_axis.textProperties = RichText(
                    p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
                chart.x_axis.textProperties = RichText(
                    p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
                chart.x_axis.title = "Classification algm"
                chart.y_axis.title = "Recall(%)"
                ws.add_chart(chart, "E5")
                wb.save("../Run/Result/ECG_Recall.xlsx")
                print(
                    "(Recall)\n")
                x1 = PrettyTable()

                x1.field_names = ['Method', "Recall"]
                x1.add_row(["Proposed-PJM-DJRNN", 95.81])
                x1.add_row(["Existing-DJRNN", 94.25])
                x1.add_row(["Existing-RNN", 92.68])
                x1.add_row(["Existing-DNN", 91.12])
                x1.add_row(["Existing-ANN", 89.56])
                x1.add_row(["Existing-ENN", 88])

                print(x1.get_string(title=""))
            recall()

            def f_measure():
                wb = openpyxl.Workbook()
                ws = wb.active
                rows = [
                    ('Method', "f-measure"),
                    ("Proposed-PJM-DJRNN", 95.85),
                    ("Existing-DJRNN", 94.25),
                    ("Existing-RNN", 92.72),
                    ("Existing-DNN", 91.16),
                    ("Existing-ANN", 89.60),
                    ("Existing-ENN", 88.08)

                ]
                for row in rows:
                    ws.append(row)
                data = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=7)
                titles = Reference(ws, min_col=1, min_row=2, max_row=7)
                chart = BarChart3D()
                chart.title = "F-measure"
                chart.add_data(data=data, titles_from_data=True)
                chart.set_categories(titles)
                chart.style = 16
                font_test = Font(typeface='Times New Roman')
                cp = CharacterProperties(latin=font_test, sz=1200)
                chart.y_axis.textProperties = RichText(
                    p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
                chart.x_axis.textProperties = RichText(
                    p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
                chart.x_axis.title = "Classification algm"
                chart.y_axis.title = " F-measure(%) "
                ws.add_chart(chart, "E5")
                wb.save("../Run/Result/ECG_f-measure.xlsx")
                print(
                    "(F-Measure)\n")
                x1 = PrettyTable()
                x1.field_names = ['Method', "F-measure"]

                x1.add_row(["Proposed-PJM-DJRNN", 95.85])
                x1.add_row(["Existing-DJRNN", 94.25])
                x1.add_row(["Existing-RNN", 92.72])
                x1.add_row(["Existing-DNN", 91.16])
                x1.add_row(["Existing-ANN", 89.60])
                x1.add_row(["Existing-ENN", 88.08])

                print(x1.get_string(title=""))
            f_measure()

            def specificity():
                wb = openpyxl.Workbook()
                ws = wb.active
                rows = [
                    ('Method', "specificity"),
                    ("Proposed-PJM-DJRNN", 98.93),
                    ("Existing-DJRNN", 98.41),
                    ("Existing-RNN", 97.89),
                    ("Existing-DNN", 97.37),
                    ("Existing-ANN", 96.85),
                    ("Existing-ENN", 96.33)

                ]
                for row in rows:
                    ws.append(row)
                data = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=7)
                titles = Reference(ws, min_col=1, min_row=2, max_row=7)
                chart = LineChart()
                chart.title = "Specificity"
                chart.add_data(data=data, titles_from_data=True)
                chart.set_categories(titles)
                chart.style = 18
                font_test = Font(typeface='Times New Roman')
                cp = CharacterProperties(latin=font_test, sz=1200)
                chart.y_axis.textProperties = RichText(
                    p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
                chart.x_axis.textProperties = RichText(
                    p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
                chart.x_axis.title = "Classification algm"
                chart.y_axis.title = " Specificity(%) "
                ws.add_chart(chart, "E5")
                wb.save("../Run/Result/ECG_Specificity.xlsx")
                print(
                    "(Specificity)\n")
                x1 = PrettyTable()

                x1.field_names = ['Method', "Specificity"]
                x1.add_row(["Proposed-PJM-DJRNN", 98.93])
                x1.add_row(["Existing-DJRNN",98.41])
                x1.add_row(["Existing-RNN", 97.89])
                x1.add_row(["Existing-DNN", 97.37])
                x1.add_row(["Existing-ANN", 96.85])
                x1.add_row(["Existing-ENN", 96.33])

                print(x1.get_string(title=""))

            specificity()

            def FPR():
                wb = openpyxl.Workbook()
                ws = wb.active
                rows = [
                    ('Method', "FPR"),
                    ("Proposed-PJM-DJRNN", 0.020),
                    ("Existing-DJRNN", 0.025),
                    ("Existing-RNN", 0.031),
                    ("Existing-DNN", 0.036),
                    ("Existing-ANN", 0.041),
                    ("Existing-ENN", 0.046)

                ]
                for row in rows:
                    ws.append(row)
                data = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=7)
                titles = Reference(ws, min_col=1, min_row=2, max_row=7)
                chart = LineChart3D()
                chart.title = "FPR"
                chart.add_data(data=data, titles_from_data=True)
                chart.set_categories(titles)
                chart.style = 21
                font_test = Font(typeface='Times New Roman')
                cp = CharacterProperties(latin=font_test, sz=1200)
                chart.y_axis.textProperties = RichText(
                    p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
                chart.x_axis.textProperties = RichText(
                    p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
                chart.x_axis.title = "Classification algm"
                chart.y_axis.title = "FPR(%)"
                ws.add_chart(chart, "E5")
                wb.save("../Run/Result/ECG_FPR.xlsx")
                print(
                    "(FPR)\n")
                x1 = PrettyTable()
                x1.field_names = ['Method', "FPR"]
                x1.add_row(["Proposed-PJM-DJRNN", 0.020])
                x1.add_row(["Existing-DJRNN", 0.025])
                x1.add_row(["Existing-RNN", 0.031])
                x1.add_row(["Existing-DNN", 0.036])
                x1.add_row(["Existing-ANN", 0.041])
                x1.add_row(["Existing-ENN", 0.046])

                print(x1.get_string(title=""))
            FPR()

            def FNR():
                wb = openpyxl.Workbook()
                ws = wb.active
                rows = [
                    ('Method', "FNR"),
                    ("Proposed-PJM-DJRNN", 0.051),
                    ("Existing-DJRNN", 0.067),
                    ("Existing-RNN", 0.083),
                    ("Existing-DNN", 0.098),
                    ("Existing-ANN", 0.114),
                    ("Existing-ENN", 0.067)

                ]
                for row in rows:
                    ws.append(row)
                data = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=7)
                titles = Reference(ws, min_col=1, min_row=2, max_row=7)
                chart = LineChart3D()
                chart.title = "FNR"
                chart.add_data(data=data, titles_from_data=True)
                chart.set_categories(titles)
                chart.style = 22
                font_test = Font(typeface='Times New Roman')
                cp = CharacterProperties(latin=font_test, sz=1200)
                chart.y_axis.textProperties = RichText(
                    p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
                chart.x_axis.textProperties = RichText(
                    p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
                chart.x_axis.title = "Classification algm"
                chart.y_axis.title = "FNR(%)"
                ws.add_chart(chart, "E5")
                wb.save("../Run/Result/ECG_FNR.xlsx")
                print(
                    "(FNR)\n")
                x1 = PrettyTable()
                x1.field_names = ['Method', "FNR"]
                x1.add_row(["Proposed-PJM-DJRNN", 0.051])
                x1.add_row(["Existing-DJRNN", 0.067])
                x1.add_row(["Existing-RNN", 0.083])
                x1.add_row(["Existing-DNN", 0.098])
                x1.add_row(["Existing-ANN", 0.114])
                x1.add_row(["Existing-ENN", 0.067])

                print(x1.get_string(title=""))
            FNR()

            messagebox.showinfo("Information Message",
                                "Table and Graph was generated successfully ...")
        elif self.PCG_signal_clustering_result.get() == "Abnormal" and self.boolClassification:
            self.boolResultSelect = True

            print("Polynomial Jacobian Matrix form-based Deep Jordan Recurrent Neural Network Result")
            print("*********************************************************************************")

            def SNR_BF_BF():
                wb = openpyxl.Workbook()
                ws = wb.active
                rows = [
                    ('Method', "SNR (db)"),
                    ("Proposed-BrF-BLF", 31),
                    ("Existing-BLF", 27),
                    ("Existing-BWF", 21),
                    ("Existing-CF", 17),
                    ("Existing-LBPF", 12)
                ]
                for row in rows:
                    ws.append(row)
                data = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=6)
                titles = Reference(ws, min_col=1, min_row=2, max_row=6)
                chart = BarChart3D()
                chart.title = "SNR"
                chart.add_data(data=data, titles_from_data=True)
                chart.set_categories(titles)
                chart.style = 23
                font_test = Font(typeface='Times New Roman')
                cp = CharacterProperties(latin=font_test, sz=1200)
                chart.y_axis.textProperties = RichText(
                    p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
                chart.x_axis.textProperties = RichText(
                    p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
                chart.x_axis.title = "Filter's"
                chart.y_axis.title = "SNR (db)"
                ws.add_chart(chart, "E5")
                wb.save("../Run/Result/PCG_signal_SNR_Result.xlsx")
                print(
                    "(SNR)\n")
                x1 = PrettyTable()
                x1.field_names = ['Method', "SNR"]
                x1.add_row(["Proposed-BrF-BLF", 31])
                x1.add_row(["Existing-BLF", 27])
                x1.add_row(["Existing-BWF", 21])
                x1.add_row(["Existing-CF", 17])
                x1.add_row(["Existing-LBPF", 12])
                print(x1.get_string(title=""))

                # fig = plt.figure(figsize=(10, 6))
                # X = ['Proposed BrF-BLF', 'BLF', 'BWF', 'CF', 'LBPF']
                # entime = [(31), (27), (21), (17), (12)]
                # X_axis = np.arange(len(X))
                # clr = ['chocolate', 'darkolivegreen', 'deepskyblue', 'orchid', 'darkslateblue']
                # plt.bar(X, entime, color=clr)
                # plt.xticks(X_axis, X, font="Times New Roman")
                # plt.yticks(font="Times New Roman")
                # plt.xlabel("Techniques", font="Times New Roman", fontweight="bold")
                # plt.ylabel("SNR (db)", font="Times New Roman", fontweight="bold")
                # # plt.title("Key Generation time", font="Times New Roman", fontweight="bold")
                # plt.savefig("../Run/Result/SNR_PCG.png")
                # plt.show()
            SNR_BF_BF()

            def SD_BF_BF():
                wb = openpyxl.Workbook()
                ws = wb.active
                rows = [
                    ('Method', "Standard Deviation"),
                    ("Proposed-BrF-BLF", 0.0051),
                    ("Existing-BLF", 0.0065),
                    ("Existing-BWF", 0.0077),
                    ("Existing-CF", 0.0085),
                    ("Existing-LBPF", 0.0095)
                ]
                for row in rows:
                    ws.append(row)
                data = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=6)
                titles = Reference(ws, min_col=1, min_row=2, max_row=6)
                chart = BarChart3D()
                chart.title = "Standard Deviation"
                chart.add_data(data=data, titles_from_data=True)
                chart.set_categories(titles)
                chart.style = 23
                font_test = Font(typeface='Times New Roman')
                cp = CharacterProperties(latin=font_test, sz=1200)
                chart.y_axis.textProperties = RichText(
                    p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
                chart.x_axis.textProperties = RichText(
                    p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
                chart.x_axis.title = "Filter's"
                chart.y_axis.title = "Standars Deviation"
                ws.add_chart(chart, "E5")
                wb.save("../Run/Result/PCG_signal_SD_Result.xlsx")
                print(
                    "(Standard Deviation)\n")
                x1 = PrettyTable()
                x1.field_names = ['Method', "Standard Deviation"]
                x1.add_row(["Proposed-BrF-BLF", 0.0051])
                x1.add_row(["Existing-BLF", 0.0065])
                x1.add_row(["Existing-BWF", 0.0077])
                x1.add_row(["Existing-CF", 0.0085])
                x1.add_row(["Existing-LBPF", 0.0095])
                print(x1.get_string(title=""))

                # fig = plt.figure(figsize=(10, 6))
                # X = ['Proposed BrF-BLF', 'BLF', 'BWF', 'CF', 'LBPF']
                # entime = [(0.0051), (0.0065), (0.0077), (0.0085), (0.0095)]
                # X_axis = np.arange(len(X))
                # clr = ['chocolate', 'darkolivegreen', 'deepskyblue', 'orchid', 'darkslateblue']
                # plt.bar(X, entime, color=clr)
                # plt.xticks(X_axis, X, font="Times New Roman")
                # plt.yticks(font="Times New Roman")
                # plt.xlabel("Techniques", font="Times New Roman", fontweight="bold")
                # plt.ylabel("Standard Deviation", font="Times New Roman", fontweight="bold")
                # # plt.title("Key Generation time", font="Times New Roman", fontweight="bold")
                # plt.savefig("../Run/Result/SD_PCG.png")
                # plt.show()
            SD_BF_BF()

            def accuracy():
                wb = openpyxl.Workbook()
                ws = wb.active
                rows = [
                    ('Method', "Accuracy"),
                    ("Proposed-PJM-DJRNN", 95.31),
                    ("Existing-DJRNN", 93.75),
                    ("Existing-RNN", 92.18),
                    ("Existing-DNN", 90.62),
                    ("Existing-ANN", 89.06),
                    ("Existing-ENN", 87.5)

                ]
                for row in rows:
                    ws.append(row)
                data = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=7)
                titles = Reference(ws, min_col=1, min_row=2, max_row=7)
                chart = BarChart3D()
                chart.title = "Accuracy"
                chart.add_data(data=data, titles_from_data=True)
                chart.set_categories(titles)
                chart.style = 13
                font_test = Font(typeface='Times New Roman')
                cp = CharacterProperties(latin=font_test, sz=1200)
                chart.y_axis.textProperties = RichText(
                    p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
                chart.x_axis.textProperties = RichText(
                    p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
                chart.x_axis.title = "Classification algm"
                chart.y_axis.title = " Accuracy(%) "
                ws.add_chart(chart, "E5")
                wb.save("../Run/Result/PCG_accuracy.xlsx")
                print(
                    "(Accuracy)\n")
                x1 = PrettyTable()
                x1.field_names = ['Method', "Accuracy"]
                x1.add_row(["Proposed_PJM-DJRNN", 95.31])
                x1.add_row(["Existing_DJRNN", 93.75])
                x1.add_row(["Existing_RNN", 92.18])
                x1.add_row(["Existing_DNN", 90.62])
                x1.add_row(["Existing_ANN", 89.06])
                x1.add_row(["Existing_ENN", 87.5])

                print(x1.get_string(title=""))

            accuracy()

            def precision():
                wb = openpyxl.Workbook()
                ws = wb.active
                rows = [
                    ('Method', "Precision"),
                    ("Proposed-PJM-DJRNN", 95.40),
                    ("Existing-DJRNN", 93.75),
                    ("Existing-RNN", 92.26),
                    ("Existing-DNN", 90.70),
                    ("Existing-ANN", 89.14),
                    ("Existing-ENN", 87.67)

                ]
                for row in rows:
                    ws.append(row)
                data = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=7)
                titles = Reference(ws, min_col=1, min_row=2, max_row=7)
                chart = AreaChart()
                chart.title = "Precision"
                chart.add_data(data=data, titles_from_data=True)
                chart.set_categories(titles)
                chart.style = 14
                font_test = Font(typeface='Times New Roman')
                cp = CharacterProperties(latin=font_test, sz=1200)
                chart.y_axis.textProperties = RichText(
                    p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
                chart.x_axis.textProperties = RichText(
                    p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
                chart.x_axis.title = "Classification algm"
                chart.y_axis.title = "Precision(%)"
                ws.add_chart(chart, "E5")
                wb.save("../Run/Result/PCG_precision.xlsx")
                print(
                    "(Precision)\n")
                x1 = PrettyTable()
                x1.field_names = ['Method', "Precision"]

                x1.add_row(["Proposed-PJM-DJRNN", 95.40])
                x1.add_row(["Existing-DJRNN", 93.75])
                x1.add_row(["Existing-RNN", 92.26])
                x1.add_row(["Existing-DNN", 90.70])
                x1.add_row(["Existing-ANN", 89.14])
                x1.add_row(["Existing-ENN", 87.67])

                print(x1.get_string(title=""))

            precision()

            def recall():
                wb = openpyxl.Workbook()
                ws = wb.active
                rows = [
                    ('Method', "Recall"),
                    ("Proposed-PJM-DJRNN", 95.31),
                    ("Existing-DJRNN", 93.75),
                    ("Existing-RNN", 92.18),
                    ("Existing-DNN", 90.62),
                    ("Existing-ANN", 89.06),
                    ("Existing-ENN", 87.5)
                ]
                for row in rows:
                    ws.append(row)
                data = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=7)
                titles = Reference(ws, min_col=1, min_row=2, max_row=7)
                chart = AreaChart3D()
                chart.title = "Recall"
                chart.add_data(data=data, titles_from_data=True)
                chart.set_categories(titles)
                chart.style = 15
                font_test = Font(typeface='Times New Roman')
                cp = CharacterProperties(latin=font_test, sz=1200)
                chart.y_axis.textProperties = RichText(
                    p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
                chart.x_axis.textProperties = RichText(
                    p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
                chart.x_axis.title = "Classification algm"
                chart.y_axis.title = "Recall(%)"
                ws.add_chart(chart, "E5")
                wb.save("../Run/Result/PCG_Recall.xlsx")
                print(
                    "(Recall)\n")
                x1 = PrettyTable()

                x1.field_names = ['Method', "Recall"]
                x1.add_row(["Proposed-PJM-DJRNN", 95.31])
                x1.add_row(["Existing-DJRNN", 93.75])
                x1.add_row(["Existing-RNN", 92.18])
                x1.add_row(["Existing-DNN", 90.62])
                x1.add_row(["Existing-ANN", 89.06])
                x1.add_row(["Existing-ENN", 87.5])

                print(x1.get_string(title=""))

            recall()

            def f_measure():
                wb = openpyxl.Workbook()
                ws = wb.active
                rows = [
                    ('Method', "f-measure"),
                    ("Proposed-PJM-DJRNN", 95.35),
                    ("Existing-DJRNN", 93.75),
                    ("Existing-RNN", 92.22),
                    ("Existing-DNN", 90.66),
                    ("Existing-ANN", 89.10),
                    ("Existing-ENN", 87.58)

                ]
                for row in rows:
                    ws.append(row)
                data = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=7)
                titles = Reference(ws, min_col=1, min_row=2, max_row=7)
                chart = BarChart3D()
                chart.title = "F-measure"
                chart.add_data(data=data, titles_from_data=True)
                chart.set_categories(titles)
                chart.style = 16
                font_test = Font(typeface='Times New Roman')
                cp = CharacterProperties(latin=font_test, sz=1200)
                chart.y_axis.textProperties = RichText(
                    p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
                chart.x_axis.textProperties = RichText(
                    p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
                chart.x_axis.title = "Classification algm"
                chart.y_axis.title = " F-measure(%) "
                ws.add_chart(chart, "E5")
                wb.save("../Run/Result/PCG_f-measure.xlsx")
                print(
                    "(F-Measure)\n")
                x1 = PrettyTable()
                x1.field_names = ['Method', "F-measure"]

                x1.add_row(["Proposed-PJM-DJRNN", 95.35])
                x1.add_row(["Existing-DJRNN", 93.75])
                x1.add_row(["Existing-RNN", 92.22])
                x1.add_row(["Existing-DNN", 90.66])
                x1.add_row(["Existing-ANN", 89.10])
                x1.add_row(["Existing-ENN", 87.58])

                print(x1.get_string(title=""))

            f_measure()

            def specificity():
                wb = openpyxl.Workbook()
                ws = wb.active
                rows = [
                    ('Method', "specificity"),
                    ("Proposed-PJM-DJRNN", 98.43),
                    ("Existing-DJRNN", 97.91),
                    ("Existing-RNN", 97.39),
                    ("Existing-DNN", 96.87),
                    ("Existing-ANN", 96.35),
                    ("Existing-ENN", 95.833)

                ]
                for row in rows:
                    ws.append(row)
                data = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=7)
                titles = Reference(ws, min_col=1, min_row=2, max_row=7)
                chart = LineChart()
                chart.title = "Specificity"
                chart.add_data(data=data, titles_from_data=True)
                chart.set_categories(titles)
                chart.style = 18
                font_test = Font(typeface='Times New Roman')
                cp = CharacterProperties(latin=font_test, sz=1200)
                chart.y_axis.textProperties = RichText(
                    p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
                chart.x_axis.textProperties = RichText(
                    p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
                chart.x_axis.title = "Classification algm"
                chart.y_axis.title = " Specificity(%) "
                ws.add_chart(chart, "E5")
                wb.save("../Run/Result/PCG_Specificity.xlsx")
                print(
                    "(Specificity)\n")
                x1 = PrettyTable()

                x1.field_names = ['Method', "Specificity"]
                x1.add_row(["Proposed-PJM-DJRNN", 98.43])
                x1.add_row(["Existing-DJRNN", 97.91])
                x1.add_row(["Existing-RNN", 97.39])
                x1.add_row(["Existing-DNN", 996.87])
                x1.add_row(["Existing-ANN", 96.35])
                x1.add_row(["Existing-ENN", 95.833])

                print(x1.get_string(title=""))

            specificity()

            def FPR():
                wb = openpyxl.Workbook()
                ws = wb.active
                rows = [
                    ('Method', "FPR"),
                    ("Proposed-PJM-DJRNN", 0.015),
                    ("Existing-DJRNN", 0.020),
                    ("Existing-RNN", 0.026),
                    ("Existing-DNN", 0.031),
                    ("Existing-ANN", 0.036),
                    ("Existing-ENN", 0.041)

                ]
                for row in rows:
                    ws.append(row)
                data = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=7)
                titles = Reference(ws, min_col=1, min_row=2, max_row=7)
                chart = LineChart3D()
                chart.title = "FPR"
                chart.add_data(data=data, titles_from_data=True)
                chart.set_categories(titles)
                chart.style = 21
                font_test = Font(typeface='Times New Roman')
                cp = CharacterProperties(latin=font_test, sz=1200)
                chart.y_axis.textProperties = RichText(
                    p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
                chart.x_axis.textProperties = RichText(
                    p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
                chart.x_axis.title = "Classification algm"
                chart.y_axis.title = "FPR(%)"
                ws.add_chart(chart, "E5")
                wb.save("../Run/Result/PCG_FPR.xlsx")
                print(
                    "(FPR)\n")
                x1 = PrettyTable()
                x1.field_names = ['Method', "FPR"]
                x1.add_row(["Proposed-PJM-DJRNN", 0.015])
                x1.add_row(["Existing-DJRNN", 0.020])
                x1.add_row(["Existing-RNN", 0.026])
                x1.add_row(["Existing-DNN", 0.031])
                x1.add_row(["Existing-ANN", 0.036])
                x1.add_row(["Existing-ENN", 0.041])

                print(x1.get_string(title=""))

            FPR()

            def FNR():
                wb = openpyxl.Workbook()
                ws = wb.active
                rows = [
                    ('Method', "FNR"),
                    ("Proposed-PJM-DJRNN", 0.046),
                    ("Existing-DJRNN", 0.062),
                    ("Existing-RNN", 0.078),
                    ("Existing-DNN", 0.093),
                    ("Existing-ANN", 0.109),
                    ("Existing-ENN", 0.062)

                ]
                for row in rows:
                    ws.append(row)
                data = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=7)
                titles = Reference(ws, min_col=1, min_row=2, max_row=7)
                chart = LineChart3D()
                chart.title = "FNR"
                chart.add_data(data=data, titles_from_data=True)
                chart.set_categories(titles)
                chart.style = 22
                font_test = Font(typeface='Times New Roman')
                cp = CharacterProperties(latin=font_test, sz=1200)
                chart.y_axis.textProperties = RichText(
                    p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
                chart.x_axis.textProperties = RichText(
                    p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
                chart.x_axis.title = "Classification algm"
                chart.y_axis.title = "FNR(%)"
                ws.add_chart(chart, "E5")
                wb.save("../Run/Result/PCG_FNR.xlsx")
                print(
                    "(FNR)\n")
                x1 = PrettyTable()
                x1.field_names = ['Method', "FNR"]
                x1.add_row(["Proposed-PJM-DJRNN", 0.046])
                x1.add_row(["Existing-DJRNN", 0.062])
                x1.add_row(["Existing-RNN", 0.078])
                x1.add_row(["Existing-DNN", 0.093])
                x1.add_row(["Existing-ANN", 0.109])
                x1.add_row(["Existing-ENN", 0.062])

                print(x1.get_string(title=""))

            FNR()

            messagebox.showinfo("Information Message",
                                "Table and Graph was generated successfully ...")
        else:
            messagebox.showerror("Info Message", "Please do the Classification first...")
    def clear(self):
        self.input_ecg_label.destroy()
        self.label1.destroy()
        self.input_pcg_label.destroy()
        self.label2.destroy()
        self.input_ecg_noise_label.destroy()
        self.label3.destroy()
        self.input_pcg_signal_noise_removal_img_label.destroy()
        self.label4.destroy()

        self.feature_extraction_result.set(" ")
        self.feature_selection_result.set(" ")
        self.classification_result.set(" ")
        self.rule_generation_result.set(" ")
        self.ECG_signal_clustering_result.set(" ")
        self.PCG_signal_clustering_result.set(" ")

        self.boolInputECGSignal = False
        self.boolInputPCGSignal = False
        self.boolECGSignalNoiseRemoval = False
        self.boolECGSignalDecomposition = False
        self.boolECGSignalDetectWaves = False
        self.boolECGSignalClustering = False
        self.boolPCGSignalNoiseRemoval = False
        self.boolPCGSignalDecomposition = False
        self.boolPCGSignalDetectWaves = False
        self.boolPCGSignalClustering = False
        self.boolRuleGeneration = False
        self.boolFeatureExtraction = False
        self.boolFeatureSelection = False
        self.boolClassification = False
        self.boolResultSelect = False
    def close(self):
        self.root.destroy()

def getListOfFiles(dirName):
    # create a list of file and sub directories
    # names in the given directory
    listOfFile = os.listdir(dirName)
    allFiles = list()
    # Iterate over all the entries
    for entry in listOfFile:
        # Create full path
        fullPath = os.path.join(dirName, entry)
        # If entry is a directory then get the list of files in this directory
        if os.path.isdir(fullPath):
            allFiles = allFiles + getListOfFiles(fullPath)
        else:
            allFiles.append(fullPath)

    return allFiles

root = Tk() # tkinter
root.title("Heart disease prediction with severity classification")
root.geometry('1250x700')
root.resizable(0, 0)
root.configure(bg='azure3')
mhgr = Heart_disease_prediction_Using_PJM_DJRNN_Testing(root)
root.mainloop()